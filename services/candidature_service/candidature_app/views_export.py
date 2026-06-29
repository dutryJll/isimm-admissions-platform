
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from io import BytesIO
from django.db.models import F

from .models import Candidature, MembreCommission
from .exports import ExportPDFService, ExportExcelService

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
except ImportError:
    colors = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_candidatures(request):
    """
    Export candidatures for responsable.
    
    Query parameters:
    - scope: 'specialite' or 'master' (required)
    - format: 'pdf' or 'xlsx' (required)
    - specialite: specialty name (required if scope='specialite')
    
    Returns: File download (PDF or Excel)
    """
    role = getattr(request.user, 'role', None)
    
    # Check permissions
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response(
            {'error': 'Permission refusée'},
            status=status.HTTP_403_FORBIDDEN
        )

    # Get parameters
    scope = request.query_params.get('scope', 'master').lower()
    format_type = request.query_params.get('format', 'pdf').lower()
    specialite_filter = request.query_params.get('specialite', '').strip()

    # Validate parameters
    if scope not in ['specialite', 'master']:
        return Response(
            {'error': 'scope doit être "specialite" ou "master"'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if format_type not in ['pdf', 'xlsx']:
        return Response(
            {'error': 'format doit être "pdf" ou "xlsx"'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Get master IDs accessible to this user
    master_ids = None
    if role in ['responsable_commission', 'commission']:
        master_ids = list(
            MembreCommission.objects.filter(
                user=request.user,
                actif=True,
                commission__actif=True
            ).values_list('commission__master_id', flat=True)
        )
        if not master_ids:
            return Response(
                {'error': 'Aucun master accessible'},
                status=status.HTTP_403_FORBIDDEN
            )

    # Query candidatures
    candidatures_qs = Candidature.objects.select_related(
        'candidat', 'master'
    ).filter(
        statut__in=['preselectionne', 'selectionne', 'inscrit']
    ).order_by(
        F('score').desc(nulls_last=True),
        '-date_soumission'
    )

    # Filter by master
    if master_ids is not None:
        candidatures_qs = candidatures_qs.filter(master_id__in=master_ids)

    # Filter by specialty if scope is 'specialite'
    if scope == 'specialite' and specialite_filter:
        candidatures_qs = candidatures_qs.filter(
            specialite__icontains=specialite_filter
        )

    # Convert to list for processing
    candidatures = list(candidatures_qs)

    if not candidatures:
        return Response(
            {'error': 'Aucune candidature à exporter'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Generate file based on format
    if format_type == 'pdf':
        return _generate_pdf_export(candidatures, scope, specialite_filter)
    else:  # xlsx
        return _generate_xlsx_export(candidatures, scope, specialite_filter)


def _generate_pdf_export(candidatures, scope, specialite):
    """Generate PDF export file"""
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            topMargin=1*cm,
            bottomMargin=1*cm,
            leftMargin=0.5*cm,
            rightMargin=0.5*cm
        )

        elements = []
        styles = getSampleStyleSheet()

        # Title
        titre_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=15,
            alignment=1
        )

        titre_text = f"Export Candidatures - {scope.upper()}"
        if specialite:
            titre_text += f" ({specialite})"

        titre = Paragraph(titre_text, titre_style)
        elements.append(titre)
        elements.append(Spacer(1, 10))

        # Table data
        headers = [
            'Rang',
            'N° Candidature',
            'Nom Candidat',
            'Spécialité',
            'Score',
            'Email',
            'Statut'
        ]

        data = [headers]
        for idx, cand in enumerate(candidatures, 1):
            data.append([
                str(idx),
                str(cand.numero or ''),
                str(cand.candidat.get_full_name() if cand.candidat else ''),
                str(cand.specialite or cand.master.specialite if cand.master else ''),
                f"{float(cand.score or 0):.2f}",
                str(cand.candidat.email if cand.candidat else ''),
                str(cand.statut or ''),
            ])

        # Create table with styling
        col_widths = [1.5*cm, 2.5*cm, 5*cm, 4*cm, 2*cm, 4*cm, 2.5*cm]
        table = Table(data, colWidths=col_widths)

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="candidatures_{scope}_{candidatures[0].master_id or "all"}.pdf"'
        )
        return response

    except Exception as e:
        return Response(
            {'error': f'Erreur lors de la génération du PDF: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def _generate_xlsx_export(candidatures, scope, specialite):
    """Generate Excel export file"""
    try:
        if not openpyxl:
            return Response(
                {'error': 'openpyxl non installé'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Candidatures'

        # Headers
        headers = [
            'Rang',
            'N° Candidature',
            'Nom Candidat',
            'Email',
            'Spécialité',
            'Score',
            'Statut',
            'Date Soumission'
        ]

        # Style headers
        header_fill = PatternFill(
            start_color='1e293b',
            end_color='1e293b',
            fill_type='solid'
        )
        header_font = Font(bold=True, color='ffffff')

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, candidature in enumerate(candidatures, 2):
            worksheet.cell(row=row_idx, column=1).value = row_idx - 1
            worksheet.cell(row=row_idx, column=2).value = candidature.numero or ''
            worksheet.cell(row=row_idx, column=3).value = (
                candidature.candidat.get_full_name() if candidature.candidat else ''
            )
            worksheet.cell(row=row_idx, column=4).value = (
                candidature.candidat.email if candidature.candidat else ''
            )
            worksheet.cell(row=row_idx, column=5).value = (
                candidature.specialite or (
                    candidature.master.specialite if candidature.master else ''
                )
            )
            worksheet.cell(row=row_idx, column=6).value = float(candidature.score or 0)
            worksheet.cell(row=row_idx, column=7).value = candidature.statut or ''
            worksheet.cell(row=row_idx, column=8).value = (
                str(candidature.date_soumission) if candidature.date_soumission else ''
            )

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # Create response
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="candidatures_{scope}_{candidatures[0].master_id or "all"}.xlsx"'
        )
        return response

    except Exception as e:
        return Response(
            {'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
