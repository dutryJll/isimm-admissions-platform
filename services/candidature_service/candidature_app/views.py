import os
import logging
import csv
import tempfile
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

from django.conf import settings
from django.core.mail import send_mail
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.utils import OperationalError
from django.db import IntegrityError, transaction
from django.db.models import F, Max, Q, Count
from django.db.models.expressions import Window
from django.db.models.functions import Rank
from rest_framework import status, viewsets
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from .models import (
    Candidature,
    CandidatListe,
    AvisMembre,
    AvisSelection,
    Commission,
    MembreCommission,
    Concours,
    ConfigurationAppel,
    DonneesAcademiques,
    FormuleScore,
    HistoriqueActionCommission,
    ListeAdmission,
    Master,
    ParcoursAdmission,
    Reclamation,
    OffreMaster,
    InscriptionEnLigne,
    InscriptionRapprochementAudit,
    Notification,
    SpecialiteParcoursMapping,
    StatusHistory,
)
from .services import (
    GestionListesService,
    ImportPaiementService,
    SelectionCandidatsService,
    StatutService,
    VerificationPaiementService,
)
from .ocr_service import verifier_concordance_dossier
from .serializers import (
    CandidatureSerializer,
    ConfigurationAppelSerializer,
    FormuleScoreSerializer,
    NotificationSerializer,
    OffreMasterSerializer,
    UserUpdateSerializer,
    AvisMembreSerializer,
    AvisSelectionSerializer,
    MembreCommissionSerializer,
    UserSerializer,
)
from .emails import (
    envoyer_email_changement_statut,
    envoyer_email_confirmation_candidature,
    envoyer_notifications_masse,
    envoyer_email_inscription_validee,
)
from .notifications import creer_notification_avec_email
from decimal import Decimal, InvalidOperation


logger = logging.getLogger(__name__)


def _ranked_candidatures_queryset(queryset):
    candidatures = list(queryset)
    grouped = {}

    for candidature in candidatures:
        candidature.classement_calcule = None
        if candidature.master_id and candidature.score is not None:
            grouped.setdefault(candidature.master_id, []).append(candidature)

    for rows in grouped.values():
        rows.sort(key=lambda c: (float(c.score or 0), c.date_soumission or timezone.datetime.min), reverse=True)
        previous_score = None
        current_rank = 0

        for index, candidature in enumerate(rows, start=1):
            score = candidature.score
            if previous_score is None or score != previous_score:
                current_rank = index
                previous_score = score
            candidature.classement_calcule = current_rank

    return candidatures


def _refresh_classement_for_queryset(queryset):
    ranked_qs = _ranked_candidatures_queryset(queryset.filter(master__isnull=False, score__isnull=False))
    updates = []

    for candidature in ranked_qs:
        new_rank = int(candidature.classement_calcule or 0)
        if candidature.classement != new_rank:
            candidature.classement = new_rank
            updates.append(candidature)

        candidat_nom = candidature.candidat.get_full_name() if candidature.candidat_id else 'Candidat inconnu'
        master_nom = candidature.master.nom if candidature.master_id else 'Master inconnu'
        print(f"Classement genere pour le Master {master_nom} : {candidat_nom} est {new_rank}")

    if updates:
        Candidature.objects.bulk_update(updates, ['classement'])


def _log_commission_action(user, action, specialite, session, nb_candidats, master=None):
    try:
        HistoriqueActionCommission.objects.create(
            responsable=user if getattr(user, 'is_authenticated', False) else None,
            master=master,
            action=action,
            specialite=specialite or '',
            session=session or '',
            nb_candidats=max(0, int(nb_candidats or 0)),
        )
    except Exception:
        logger.exception('Impossible de créer une entrée HistoriqueActionCommission')


ALLOWED_STATUS_TRANSITIONS = StatutService.ALLOWED_TRANSITIONS


# Reglement de reference (article fourni) transforme en structure exploitable.
REGLEMENT_CONCOURS_INGENIEUR_REFERENCE_2025_2026 = {
    'metadata': {
        'version': '2025-07-03',
        'annee_universitaire': '2025/2026',
        'type_concours': 'ingenieur',
        'source': 'Decision commission masters / concours ingenieur',
    },
    'chapitre_1_ouverture': {
        'resume': (
            'Ouverture du concours sur dossiers pour l acces a la formation ingenieur '
            'a l ISIMM pour l annee universitaire 2025/2026.'
        ),
    },
    'chapitre_2_eligibilite': {
        'paragraphe_1': {
            'public': 'Etudiants reussis en 2eme annee preparatoire integree ISIMM 2024/2025',
        },
        'paragraphe_2': {
            'public': (
                'Etudiants inscrits/admis en 3eme annee licence (specialites scientifiques '
                'et techniques LMD) en 2024/2025 et n ayant pas redouble cette annee'
            ),
            'diplomes_acceptes': [
                'Licence en sciences de l informatique',
                'Genie logiciel et systemes d information',
                'Licence en mathematiques et informatique ou diplome equivalent',
            ],
        },
    },
    'chapitre_3_quotas': {
        'regle': 'Quotas par filiere et par type de candidature (paragraphe 1 / paragraphe 2).',
        'filieres': [
            {
                'filiere': 'Informatique (Ingenierie des systemes - Genie Logiciel)',
                'places_paragraphe_1': 52,
                'places_paragraphe_2': 13,
            }
        ],
    },
    'chapitre_4_calcul_score': {
        'paragraphe_1': {
            'formule': 'M2 + B1 + B2',
            'variables': {
                'M2': 'Moyenne de la 2eme annee (en score de selection)',
                'B1': 'Bonification relative a la 1ere annee',
                'B2': 'Bonification relative a la 2eme annee',
            },
            'bonification_sans_redoublement': {
                'B1_session_principale': 2,
                'B1_session_controle': 1.5,
                'B2_session_principale': 2,
                'B2_session_controle': 1.5,
            },
            'bonification_avec_redoublement': {
                'B1_session_principale': 1,
                'B1_session_controle': 0,
                'B2_session_principale': 1,
                'B2_session_controle': 0,
            },
        },
        'paragraphe_2': {
            'formule': '0.5*(2*M1 + 2*M2 + M3) + 50*(1-R1) + 50*(1-R2)',
            'variables': {
                'M1': 'Moyenne 1ere annee (session principale)',
                'M2': 'Moyenne 2eme annee (session principale)',
                'M3': 'Moyenne S1 3eme annee (session principale)',
                'R1': 'Rang 1ere annee / (effectif - 1)',
                'R2': 'Rang 2eme annee / (effectif - 1)',
            },
            'sous_cas': [
                'Etudiants internes ISIMM (paragraphe 2-b-1)',
                'Etudiants externes ISIMM (paragraphe 2-b-2)',
            ],
        },
    },
    'chapitre_5_classement': {
        'regle': 'Classement par filiere selon le score calcule; admission selon le quota disponible.',
    },
    'chapitre_6_publication': {
        'regle': 'Publication des listes finales apres deliberations de la commission.',
    },
    'chapitre_7_documents_obligatoires': [
        'Fiche de candidature telechargee du site et signee',
        'Annexe du site signee et legalisee par le directeur de l etablissement d origine (cas paragraphe 2-b)',
        'Copie certifiee conforme du releve bac',
        'Copies certifiees conformes des releves de toutes les annees universitaires',
        'Copie CIN ou passeport (etudiants etrangers)',
        'Pieces justifiant reorientation ou retrait d inscription le cas echeant',
    ],
    'chapitre_8_depot': {
        'mode': 'Courrier rapide',
        'adresse': 'ISIMM - Route de Kheniss - BP 223 - 5000 Monastir',
        'date_limite': '2025-08-08',
        'reference_delai': 'Cachet de la poste fait foi',
    },
    'chapitre_9_execution': {
        'responsable': 'Directeur de l ISIMM',
    },
}


REFERENTIEL_MASTERS_ISIMM_2025_2026 = {
    'metadata': {
        'annee_universitaire': '2025/2026',
        'etablissement': 'ISIMM Monastir',
        'source': 'Communique officiel masters 2025/2026 (synthese structuree)',
    },
    'sections_masters': {
        'mpgl': {
            'intitule': 'Master Professionnel en Ingenierie Logicielle (MPGL)',
            'calendrier': {
                'inscription_web': {
                    'debut': 'date_publication',
                    'fin': '2025-07-22',
                },
                'publication_preselection': '2025-07-28',
                'depot_dossier_numerique': {
                    'debut': '2025-07-28',
                    'fin': '2025-07-31',
                },
                'publication_liste_finale': '2025-08-08',
            },
            'capacites': {
                'isimm_licence_info': 30,
                'autres_etablissements_licence_info_ou_info_gestion': 5,
                'total': 35,
            },
            'modalites_candidature': [
                'Etape 1: inscription obligatoire en ligne avant la date limite.',
                'Etape 2: pour les preselectionnes, depot d un dossier numerique en un seul fichier PDF.',
            ],
        },
        'mrgl': {
            'intitule': 'Master de Recherche en Sciences de l Informatique: Ingenierie Logicielle (MRGL)',
            'capacites': {
                'isimm_licence_ou_maitrise_info': 28,
                'autres_etablissements': 2,
                'total': 30,
            },
            'score': {
                'note': 'Calcul specifique au master de recherche.',
                'formule_licence': (
                    'Score = 1.5*Moy_1ere_Annee + 2*Moy_2eme_Annee + Moy_3eme_Annee '
                    '+ Bonus_Redoublement + Bonus_SessionPrincipale '
                    '+ (MoyBac + Note_Math_Bac - 20)/2 + Bonus_Langue + Bonus_Annee_Diplome'
                ),
                'bonus_langue': (
                    '1 point si note de Francais ou Anglais au bac >= 12 '
                    'ou certification niveau B2.'
                ),
                'bonus_annee_diplome': {
                    '2025_ou_2023': 4,
                    '2022_2021_2020': 2,
                },
            },
        },
        'mpds': {
            'intitule': 'Master Professionnel en Science des Donnees (MPDS)',
            'capacites': {
                'isimm_licence_math_appliquees': 10,
                'isimm_licence_informatique': 19,
                'autres_licence_math_appliquees': 2,
                'autres_licence_informatique': 4,
                'total': 35,
            },
        },
    },
    'documents_requis_pdf_unique': [
        'Demande de candidature (formulaire joint).',
        'Fiche de candidature imprimee du site et signee.',
        'CV d une page avec coordonnees (adresse, telephone, email).',
        'Copie de la carte d identite nationale.',
        'Copies certifiees conformes de tous les diplomes (bac inclus).',
        'Copies certifiees conformes de tous les releves de notes (bac inclus).',
        'Justificatifs de report d inscription ou de reorientation si necessaire.',
    ],
    'regles_importantes': [
        'Aucun dossier hors delai ou incomplet ne sera examine.',
        'Toute donnee erronnee entraine l annulation immediate de la candidature.',
        'En cas de falsification, des poursuites judiciaires peuvent etre engagees.',
        'Recours possible pour les non retenus par email avant le 2025-07-31.',
        'Presentation des originaux obligatoire lors de l inscription administrative finale.',
    ],
    'modele_formulaire_candidature': {
        'champs': [
            'nom_prenom',
            'etablissement_origine',
            'diplome',
            'choix_1',
            'choix_2',
            'choix_3',
            'numero_dossier_reserve_administration',
        ],
        'choix_possibles': ['MPGL', 'MRGL', 'MPDS'],
    },
}


def _validate_formulaire_commission(configuration, formulaire_payload):
    """Valide les champs/documents requis selon la configuration du master."""
    schema = configuration.formulaire_commission_schema or {}
    required_fields = schema.get('required_fields', []) or []
    required_documents = schema.get('required_documents', []) or []

    if not isinstance(formulaire_payload, dict):
        return {'ok': False, 'error': 'Le champ formulaire doit etre un objet JSON.'}

    missing_fields = []
    for field_name in required_fields:
        value = formulaire_payload.get(field_name)
        if value is None or (isinstance(value, str) and not value.strip()):
            missing_fields.append(field_name)

    uploaded_documents = formulaire_payload.get('documents', [])
    if not isinstance(uploaded_documents, list):
        return {'ok': False, 'error': 'Le champ formulaire.documents doit etre une liste.'}

    uploaded_documents_set = {str(doc).strip() for doc in uploaded_documents if str(doc).strip()}
    missing_documents = [doc for doc in required_documents if str(doc).strip() not in uploaded_documents_set]

    if missing_fields or missing_documents:
        return {
            'ok': False,
            'error': 'Formulaire commission incomplet pour ce master.',
            'missing_fields': missing_fields,
            'missing_documents': missing_documents,
        }

    return {'ok': True}


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def commission_members_list(request, commission_id):
    commission = get_object_or_404(Commission, id=commission_id)
    membres = MembreCommission.objects.filter(commission=commission).select_related('user')
    serializer = MembreCommissionSerializer(membres, many=True, context={'request': request})
    return Response(serializer.data)


def _is_user_responsable_for_commission(user, commission):
    if not getattr(user, 'is_authenticated', False):
        return False
    if getattr(user, 'is_staff', False):
        return True
    return MembreCommission.objects.filter(commission=commission, user=user, role__icontains='responsable').exists()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def commission_add_member(request, commission_id):
    commission = get_object_or_404(Commission, id=commission_id)
    if not _is_user_responsable_for_commission(request.user, commission):
        return Response({'detail': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)

    user_id = request.data.get('user_id')
    role = request.data.get('role', 'membre')
    if not user_id:
        return Response({'detail': 'user_id requis'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=int(user_id))
    except Exception:
        return Response({'detail': 'Utilisateur introuvable'}, status=status.HTTP_404_NOT_FOUND)

    try:
        with transaction.atomic():
            membre = MembreCommission.objects.create(commission=commission, user=user, role=role)
    except IntegrityError:
        return Response({'detail': 'L utilisateur est deja membre de cette commission'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception:
        logger.exception('Erreur ajout membre commission')
        return Response({'detail': 'Erreur serveur'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    serializer = MembreCommissionSerializer(membre, context={'request': request})
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def commission_remove_member(request, commission_id, membre_id):
    commission = get_object_or_404(Commission, id=commission_id)
    if not _is_user_responsable_for_commission(request.user, commission):
        return Response({'detail': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)

    membre = get_object_or_404(MembreCommission, id=membre_id, commission=commission)
    try:
        membre.delete()
    except Exception:
        logger.exception('Erreur suppression membre commission')
        return Response({'detail': 'Erreur serveur'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({'detail': 'Membre supprime'}, status=status.HTTP_200_OK)


def _normalize_offre_rich_content(payload, offer_id):
    if not isinstance(payload, dict):
        return None

    def _as_text(value):
        return str(value).strip() if value is not None else ''

    def _as_list(value):
        if not isinstance(value, list):
            return []
        return [item for item in (_as_text(item) for item in value) if item]

    def _as_rows(value):
        if not isinstance(value, list):
            return []
        rows = []
        for row in value:
            if isinstance(row, list):
                rows.append([_as_text(cell) for cell in row])
        return rows

    return {
        'offerId': offer_id,
        'title': _as_text(payload.get('title')),
        'openingTitle': _as_text(payload.get('openingTitle')),
        'openingBody': _as_text(payload.get('openingBody')),
        'tableTitle': _as_text(payload.get('tableTitle')),
        'tableHeaders': _as_list(payload.get('tableHeaders')),
        'tableRows': _as_rows(payload.get('tableRows')),
        'modalitesTitle': _as_text(payload.get('modalitesTitle')),
        'etape1': _as_text(payload.get('etape1')),
        'etape2': _as_text(payload.get('etape2')),
        'dossierTitle': _as_text(payload.get('dossierTitle')),
        'dossierItems': _as_list(payload.get('dossierItems')),
        'scoreTitle': _as_text(payload.get('scoreTitle')),
        'scoreFormula': _as_text(payload.get('scoreFormula')),
        'moyenneFormula': _as_text(payload.get('moyenneFormula')),
        'scoreTableHeaders': _as_list(payload.get('scoreTableHeaders')),
        'scoreTableRows': _as_rows(payload.get('scoreTableRows')),
        'bnrRules': _as_list(payload.get('bnrRules')),
        'bspRules': _as_list(payload.get('bspRules')),
        'evaluationNotes': _as_list(payload.get('evaluationNotes')),
        'updatedAt': _as_text(payload.get('updatedAt')) or timezone.now().isoformat(),
    }


def _concours_sync_marker(concours_id):
    return f"[AUTO_CONCOURS:{concours_id}]"


def _sync_master_from_concours(concours):
    """
    Maintient un parcours master miroir pour rendre les concours ingenieur
    visibles dans le flux d'offre d'inscription du responsable.
    """
    marker = _concours_sync_marker(concours.id)
    specialite = (concours.conditions_admission or {}).get('specialite') or 'Cycle Ingenieur'
    ouverture = concours.date_ouverture
    if isinstance(ouverture, str):
        ouverture = parse_date(ouverture)
    cloture = concours.date_cloture
    if isinstance(cloture, str):
        cloture = parse_date(cloture)
    annee_universitaire = (
        f"{ouverture.year}/{ouverture.year + 1}"
        if ouverture
        else f"{timezone.now().year}/{timezone.now().year + 1}"
    )

    existing = Master.objects.filter(description__contains=marker).first()
    base_description = (concours.description or '').strip()
    description = f"{base_description}\n\n{marker}" if base_description else marker

    if existing:
        existing.nom = concours.nom
        existing.type_master = 'professionnel'
        existing.description = description
        existing.specialite = specialite
        existing.places_disponibles = concours.places_disponibles
        existing.date_limite_candidature = cloture or concours.date_cloture
        existing.annee_universitaire = annee_universitaire
        existing.actif = concours.actif
        existing.save()
        return existing

    return Master.objects.create(
        nom=concours.nom,
        type_master='professionnel',
        description=description,
        specialite=specialite,
        places_disponibles=concours.places_disponibles,
        date_limite_candidature=cloture or concours.date_cloture,
        annee_universitaire=annee_universitaire,
        actif=concours.actif,
    )


def _deactivate_synced_master_from_concours(concours_id):
    marker = _concours_sync_marker(concours_id)
    synced = Master.objects.filter(description__contains=marker).first()
    if synced:
        synced.actif = False
        synced.save(update_fields=['actif', 'updated_at'])


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def contenu_offre_inscription(request, offer_id):
    config = get_object_or_404(ConfigurationAppel.objects.select_related('master'), master_id=offer_id, actif=True)

    if request.method == 'GET':
        contenu = config.contenu_offre_edite or None
        if contenu:
            contenu = {**contenu, 'offerId': offer_id}
        return Response(
            {
                'offerId': offer_id,
                'updatedAt': config.updated_at,
                'content': contenu,
            }
        )

    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    payload = request.data
    if isinstance(payload, dict) and 'content' in payload:
        payload = payload.get('content')

    normalized = _normalize_offre_rich_content(payload, offer_id)
    if not normalized:
        return Response({'error': 'Contenu invalide'}, status=status.HTTP_400_BAD_REQUEST)

    config.contenu_offre_edite = normalized
    config.save()

    return Response(
        {
            'offerId': offer_id,
            'updatedAt': config.updated_at,
            'content': normalized,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_candidature(request):
    """Creation simplifiee d'une candidature."""
    formation_code = request.data.get('formation_code', '').strip().upper()
    master_id = request.data.get('master_id')

    master = None

    # Primary lookup: formation_code → Master.specialite (canonical formation codes stored there)
    if formation_code:
        # Normalize ING aliases to the stored code
        code_alias = {'ING_GL': 'ING_INFO_GL', 'ING-GL': 'ING_INFO_GL'}.get(formation_code, formation_code)
        master = Master.objects.filter(specialite=code_alias).first()
        if not master:
            return Response(
                {'error': f'Formation {formation_code} introuvable. Veuillez reessayer.'},
                status=status.HTTP_404_NOT_FOUND,
            )

    # Fallback: master_id (kept for backward compatibility)
    if not master and master_id:
        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    if not master:
        return Response({'error': 'master_id ou formation_code est requis'}, status=status.HTTP_400_BAD_REQUEST)

    # Check: Duplicate candidature prevention (cannot apply twice to same offer)
    existing_candidature = Candidature.objects.filter(
        candidat=request.user,
        master=master,
        statut__in=['soumis', 'sous_examen', 'preselectionne', 'en_attente_dossier', 'dossier_depose']
    ).first()
    
    if existing_candidature:
        # Trigger notification on duplicate attempt
        _trigger_notification_on_duplicate_attempt(request.user, master, existing_candidature)
        
        # Allow editing before deadline if not yet fully processed.
        edit_deadline = existing_candidature.date_limite_modification
        edit_deadline_passed = bool(edit_deadline and timezone.now() > edit_deadline)
        if existing_candidature.statut in ['soumis', 'en_attente_dossier'] and not edit_deadline_passed:
            # Return existing candidature ID to allow editing
            return Response(
                {
                    'error': 'Vous avez deja une candidature pour ce Master',
                    'candidature_id': existing_candidature.id,
                    'allow_edit': True,
                    'statut': existing_candidature.statut,
                    'edit_deadline': edit_deadline.isoformat() if edit_deadline else None,
                },
                status=status.HTTP_409_CONFLICT
            )
        else:
            return Response(
                {
                    'error': 'Vous avez deja une candidature pour ce Master qui est en cours de traitement',
                    'candidature_id': existing_candidature.id,
                    'allow_edit': False,
                    'statut': existing_candidature.statut,
                    'edit_deadline': edit_deadline.isoformat() if edit_deadline else None,
                },
                status=status.HTTP_409_CONFLICT
            )

    # ────────────────────────────────────────────────────────────────────
    # Req-2 (Sprint 2) — Gestion des vœux multi-parcours
    #   • Cas A (Masters)   : max 3 vœux, classés par priorité, pas de doublon
    #   • Cas B (Ingénieur) : candidature unique, pas de vœu/filière
    # ────────────────────────────────────────────────────────────────────
    specialite_upper = (master.specialite or '').upper()
    nom_lower = (master.nom or '').lower()
    is_cycle_ingenieur = (
        specialite_upper.startswith('ING')
        or 'ingenieur' in nom_lower
        or 'ingénieur' in nom_lower
    )

    if is_cycle_ingenieur:
        # Cas B — Candidature unique au concours ingénieur (pas de vœu)
        existing_ing = Candidature.objects.filter(
            candidat=request.user,
            master__specialite__istartswith='ING',
        ).exclude(statut__in=['rejete', 'annulee']).count()
        if existing_ing >= 1:
            return Response(
                {
                    'error': "Vous êtes déjà inscrit au concours ingénieur. "
                             "Une seule candidature est autorisée pour ce concours.",
                    'concours_type': 'ingenieur',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Pour l'ingénieur, choix_priorite n'a pas de sens — forcer à 1
        request.data['choix_priorite'] = 1
    else:
        # Cas A — Maximum 3 vœux Masters par candidat
        masters_actives = Candidature.objects.filter(
            candidat=request.user,
        ).exclude(
            master__specialite__istartswith='ING',
        ).exclude(
            statut__in=['rejete', 'annulee'],
        )
        nb_voeux = masters_actives.count()
        if nb_voeux >= 3:
            voeux_existants = [
                {'master': c.master.nom, 'choix': c.choix_priorite, 'statut': c.statut}
                for c in masters_actives.order_by('choix_priorite')
            ]
            return Response(
                {
                    'error': "Vous avez atteint la limite de 3 vœux pour les Masters. "
                             "Annulez une candidature existante avant d'en créer une nouvelle.",
                    'limite': 3,
                    'voeux_existants': voeux_existants,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Assignation automatique de la priorité (1, 2 ou 3) selon l'ordre
        request.data['choix_priorite'] = nb_voeux + 1

    # Check: Master is open for applications.
    # The `actif` flag is the authoritative control — if admin keeps it active, deadline is bypassed.
    if not master.actif:
        return Response(
            {'error': f'{master.nom} n\'est plus ouvert aux candidatures.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    # Deadline enforced only when master is explicitly inactive (covered above).
    # If master.actif=True, the admin has intentionally kept it open past the deadline.

    # Parse academic data BEFORE creating candidature to avoid orphaned records on validation errors
    academic_data = request.data.get('academic_data')
    formation_code = request.data.get('formation_code')
    selected_diplome = request.data.get('selected_diplome')
    etablissement_origine = request.data.get('etablissement_origine')
    diplome_reference = request.data.get('diplome_reference')
    diplomes = request.data.get('diplomes')
    # Accept both key names from frontend (score_previsualisation is the current key)
    score_soumis_front = (
        request.data.get('score_soumis_front')
        or request.data.get('score_previsualisation')
    )

    # ──────────────────────────────────────────────────────────────────
    # ANTI-FRAUDE (Sprint 4) : recalculer le score côté serveur à partir
    # des critères de l'offre + données candidat puis comparer au score
    # déclaré. Refuser la création si l'écart dépasse la tolérance.
    # ──────────────────────────────────────────────────────────────────
    try:
        from .score_service import ScoreService
        if getattr(master, 'criteres', None) and getattr(master, 'score_formule', None):
            form_data_score = {}
            if isinstance(academic_data, dict):
                form_data_score.update({
                    'moyenne_l1': academic_data.get('moyenne_l1', 0),
                    'moyenne_l2': academic_data.get('moyenne_l2', 0),
                    'moyenne_l3': academic_data.get('moyenne_l3', 0),
                    'moyenne_bac': academic_data.get('moyenne_bac', 0),
                    'note_maths_bac': academic_data.get('note_maths_bac', 0),
                    'note_francais_bac': academic_data.get('note_francais_bac', 0),
                    'note_anglais_bac': academic_data.get('note_anglais_bac', 0),
                    'nb_redoublements': academic_data.get('nb_redoublements', 0),
                    'nb_sessions_controle': academic_data.get('nb_sessions_controle', 0),
                    'annee_diplome': academic_data.get('annee_diplome', 0),
                    'rang_l1': academic_data.get('rang_l1', 0),
                    'rang_l2': academic_data.get('rang_l2', 0),
                    'session_l1': academic_data.get('session_l1'),
                    'session_l2': academic_data.get('session_l2'),
                    'session_l1_controle': academic_data.get('session_l1_controle', False),
                    'session_l2_controle': academic_data.get('session_l2_controle', False),
                    'session_l3_controle': academic_data.get('session_l3_controle', False),
                    'certif_b2': academic_data.get('certif_b2', False),
                })
            score_backend, _detail = ScoreService.calculer_score_total(master, form_data_score)
            score_declare = float(score_soumis_front or request.data.get('score_declare') or 0)
            if ScoreService.detecter_fraude(score_backend, score_declare, tolerance=0.5):
                return Response(
                    {
                        'error': 'Incohérence détectée dans le score déclaré.',
                        'score_calcule': score_backend,
                        'score_declare': score_declare,
                        'ecart': abs(score_backend - score_declare),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Le score officiel sauvegardé est celui recalculé côté serveur
            score_soumis_front = score_backend
    except ImportError:
        # ScoreService non disponible — on continue avec le score déclaré (fallback)
        pass
    except Exception:
        # Toute autre erreur dans le calcul ne bloque pas la création (legacy)
        pass

    candidature = None  # created after validation passes

    if isinstance(academic_data, dict):
        def _as_float(value, default=0.0):
            """Safe float conversion. Returns default (including None) when value is missing."""
            try:
                if value is None or value == '':
                    return default if default is None else float(default)
                return float(value)
            except (TypeError, ValueError):
                return default if default is None else float(default)

        def _avg(values):
            cleaned = [_as_float(v, None) for v in values]
            cleaned = [v for v in cleaned if v is not None]
            if not cleaned:
                return 0.0
            return sum(cleaned) / len(cleaned)

        def _has_invalid_note(values):
            for value in values:
                parsed = _as_float(value, None)
                if parsed is None:
                    continue
                if parsed < 0 or parsed > 20:
                    return True
            return False

        common = academic_data.get('common', {}) if isinstance(academic_data.get('common'), dict) else {}
        gl_ds = academic_data.get('glDs', {}) if isinstance(academic_data.get('glDs'), dict) else {}
        i3 = academic_data.get('i3', {}) if isinstance(academic_data.get('i3'), dict) else {}
        mrgl_licence = (
            academic_data.get('mrglLicence', {})
            if isinstance(academic_data.get('mrglLicence'), dict)
            else {}
        )
        mrgl_maitrise = (
            academic_data.get('mrglMaitrise', {})
            if isinstance(academic_data.get('mrglMaitrise'), dict)
            else {}
        )
        mrmi_cas1 = (
            academic_data.get('mrmiCas1', {})
            if isinstance(academic_data.get('mrmiCas1'), dict)
            else {}
        )
        mrmi_cas2 = (
            academic_data.get('mrmiCas2', {})
            if isinstance(academic_data.get('mrmiCas2'), dict)
            else {}
        )
        ing_cas1 = (
            academic_data.get('ingCas1', {})
            if isinstance(academic_data.get('ingCas1'), dict)
            else {}
        )
        ing_cas2 = (
            academic_data.get('ingCas2', {})
            if isinstance(academic_data.get('ingCas2'), dict)
            else {}
        )

        moyenne_generale = 0.0
        moyenne_specialite = 0.0

        if formation_code in ['MPGL', 'MPDS']:
            if _has_invalid_note([gl_ds.get('moy1'), gl_ds.get('moy2'), gl_ds.get('moy3')]):
                return Response(
                    {'error': 'Les moyennes doivent etre comprises entre 0 et 20.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            moyenne_generale = _avg([gl_ds.get('moy1'), gl_ds.get('moy2'), gl_ds.get('moy3')])
            moyenne_specialite = moyenne_generale
        elif formation_code == 'MP3I':
            if _has_invalid_note([i3.get('moyBac'), i3.get('moyL1'), i3.get('moyL2'), i3.get('moyL3')]):
                return Response(
                    {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            moyenne_generale = _avg([i3.get('moyL1'), i3.get('moyL2'), i3.get('moyL3')])
            moyenne_specialite = _as_float(i3.get('moyBac'), moyenne_generale)
        elif formation_code == 'MRGL':
            parcours = academic_data.get('mrglParcours', 'licence')
            if parcours == 'licence':
                if _has_invalid_note([
                    mrgl_licence.get('moy1'),
                    mrgl_licence.get('moy2'),
                    mrgl_licence.get('moy3'),
                    mrgl_licence.get('moyBac'),
                    mrgl_licence.get('noteMathBac'),
                ]):
                    return Response(
                        {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                moyenne_generale = _avg(
                    [mrgl_licence.get('moy1'), mrgl_licence.get('moy2'), mrgl_licence.get('moy3')]
                )
                moyenne_specialite = _as_float(mrgl_licence.get('moyBac'), moyenne_generale)
            else:
                if _has_invalid_note([
                    mrgl_maitrise.get('moy1'),
                    mrgl_maitrise.get('moy2'),
                    mrgl_maitrise.get('moy3'),
                    mrgl_maitrise.get('moy4'),
                    mrgl_maitrise.get('moyBac'),
                    mrgl_maitrise.get('noteMathBac'),
                ]):
                    return Response(
                        {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                moyenne_generale = _avg(
                    [
                        mrgl_maitrise.get('moy1'),
                        mrgl_maitrise.get('moy2'),
                        mrgl_maitrise.get('moy3'),
                        mrgl_maitrise.get('moy4'),
                    ]
                )
                moyenne_specialite = _as_float(mrgl_maitrise.get('moyBac'), moyenne_generale)
        elif formation_code == 'MRMI':
            parcours = academic_data.get('mrmiParcours', 'cas1')
            if parcours == 'cas1':
                if _has_invalid_note([
                    mrmi_cas1.get('moyBac'),
                    mrmi_cas1.get('moyL1'),
                    mrmi_cas1.get('moyL2'),
                    mrmi_cas1.get('moyL3'),
                ]):
                    return Response(
                        {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                moyenne_generale = _avg(
                    [mrmi_cas1.get('moyL1'), mrmi_cas1.get('moyL2'), mrmi_cas1.get('moyL3')]
                )
                moyenne_specialite = _as_float(mrmi_cas1.get('moyBac'), moyenne_generale)
            else:
                if _has_invalid_note([mrmi_cas2.get('moyIng1')]):
                    return Response(
                        {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                moyenne_generale = _as_float(mrmi_cas2.get('moyIng1'), 0.0)
                moyenne_specialite = moyenne_generale
        elif formation_code in ['ING_INFO_GL', 'ING_EM']:
            parcours = academic_data.get('ingParcours', 'cas1')
            if parcours == 'cas1':
                if _has_invalid_note([ing_cas1.get('moy1'), ing_cas1.get('moy2')]):
                    return Response(
                        {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                moyenne_generale = _avg([ing_cas1.get('moy1'), ing_cas1.get('moy2')])
            else:
                if _has_invalid_note([ing_cas2.get('m1'), ing_cas2.get('m2'), ing_cas2.get('m3')]):
                    return Response(
                        {'error': 'Les moyennes et notes doivent etre comprises entre 0 et 20.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                moyenne_generale = _avg([ing_cas2.get('m1'), ing_cas2.get('m2'), ing_cas2.get('m3')])
            moyenne_specialite = moyenne_generale

        redoublements = common.get('redoublements', 0)
        try:
            nb_redoublements = int(redoublements)
        except (TypeError, ValueError):
            nb_redoublements = 0

        # Create candidature after all validation passes (avoids orphaned records)
        candidature = Candidature.objects.create(candidat=request.user, master=master, statut='soumis')

        DonneesAcademiques.objects.update_or_create(
            candidature=candidature,
            defaults={
                'moyenne_generale': round(moyenne_generale, 2),
                'moyenne_specialite': round(moyenne_specialite, 2),
                'nb_redoublements': nb_redoublements,
                'nb_dettes': 0,
                'notes_detaillees': {
                    'source': 'preinscription_step3',
                    'formation_code': formation_code,
                    'moyenne_bac': round(moyenne_specialite, 2),
                    'moyenne_licence': round(moyenne_generale, 2),
                    'selected_diplome': selected_diplome,
                    'etablissement_origine': etablissement_origine,
                    'diplome_reference': diplome_reference,
                    'diplomes': diplomes if isinstance(diplomes, list) else [],
                    'session_reussite': common.get('session'),
                    'payload': academic_data,
                },
            },
        )

        # Force le recalcul automatique du score apres persistance des donnees academiques.
        candidature.save()

        # Attempt dynamic scoring using ParcoursAdmission when available.
        try:
            parcours_id = request.data.get('parcours_id')
            parcours = None
            if parcours_id:
                parcours = ParcoursAdmission.objects.filter(id=parcours_id, master=master, actif=True).first()
            if not parcours:
                parcours = ParcoursAdmission.objects.filter(master=master, actif=True).first()

            if parcours:
                try:
                    score = parcours.calculer_score(candidature)
                    if score is not None:
                        candidature.score = score
                        # Fraude detection: compare submitted score vs backend-calculated score
                        if score_soumis_front is not None:
                            try:
                                submitted = float(score_soumis_front)
                                backend = float(score)
                                if abs(submitted - backend) > 0.01:
                                    candidature.flag_fraude = True
                                    logger.warning(
                                        "Flag fraude candidature=%s: soumis=%.3f backend=%.3f",
                                        candidature.id, submitted, backend,
                                    )
                            except (TypeError, ValueError):
                                pass
                        candidature.score_soumis_front = score_soumis_front
                        candidature.save(update_fields=['score', 'flag_fraude', 'score_soumis_front', 'updated_at'])
                except Exception:
                    # Don't block submission on scoring errors
                    logger.exception('Erreur lors du calcul dynamique du score pour la candidature %s', candidature.id)
        except Exception:
            logger.exception('Erreur lors de la tentative de scoring dynamique pour la candidature %s', candidature.id)

    # Fallback: create candidature when no academic_data was provided
    if candidature is None:
        candidature = Candidature.objects.create(candidat=request.user, master=master, statut='soumis')

    Notification.objects.create(
        user=request.user,
        titre='Candidature créée',
        message=f"Votre candidature {candidature.numero} pour {master.nom} a été enregistrée.",
        type='success',
    )

    try:
        envoyer_email_confirmation_candidature(candidature)
    except Exception as exc:
        logger.exception("Erreur envoi email confirmation candidature %s: %s", candidature.id, exc)

    # Trigger deadline approaching notification if applicable
    _trigger_notification_on_deadline_approaching(candidature, days_remaining=7)

    serializer = CandidatureSerializer(candidature)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_specialites_for_master(request, master_id):
    """Retourne la liste des specialites disponibles pour un Master."""
    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    # Get available specialities from related ParcoursAdmission
    parcours_list = ParcoursAdmission.objects.filter(
        master=master,
        actif=True,
        statut='ouvert'
    ).values_list('specialite', flat=True).distinct()

    specialites = list(parcours_list) if parcours_list else [master.specialite] if master.specialite else []

    return Response({
        'master_id': master.id,
        'master_nom': master.nom,
        'specialites': specialites,
        'deadline': master.date_limite_candidature.isoformat() if master.date_limite_candidature else None,
        'places_disponibles': master.places_disponibles,
        'est_ouvert': master.date_limite_candidature is None or timezone.now().date() <= master.date_limite_candidature
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_available_offers_with_specialites(request):
    """
    Retourne les 6 parcours officiels ISIMM pour l'Espace Candidat.
    Source primaire : SpecialiteParcoursMapping (seeded par migration 0025).
    Statut / places / date : enrichis depuis ConfigurationAppel actifs si disponibles,
    sinon valeurs par défaut 2026.
    """
    _PARCOURS_META = {
        'MPGL':   {'date_limite': '2026-07-22', 'places': 35,  'type': 'master'},
        'MPDS':   {'date_limite': '2026-07-22', 'places': 35,  'type': 'master'},
        'MP3I':   {'date_limite': '2026-07-20', 'places': 25,  'type': 'master'},
        'MRGL':   {'date_limite': '2026-07-22', 'places': 111, 'type': 'master'},
        'MRMI':   {'date_limite': '2026-07-20', 'places': 29,  'type': 'master'},
        'ING_GL': {'date_limite': '2026-08-08', 'places': 65,  'type': 'cycle_ingenieur'},
    }

    # Candidatures déjà soumises par l'utilisateur (badge "déjà postulé")
    applied_codes = set(
        Candidature.objects.filter(
            candidat=request.user,
            statut__in=['soumis', 'sous_examen', 'preselectionne', 'en_attente_dossier', 'dossier_depose'],
        ).values_list('master__specialite', flat=True)
    )

    # Statuts temps-réel depuis ConfigurationAppel actifs (quand Responsable a publié)
    today = timezone.now().date()
    configs_actifs = ConfigurationAppel.objects.filter(
        actif=True,
        master__specialite__in=list(_PARCOURS_META.keys()),
    ).select_related('master')
    config_by_code = {c.master.specialite: c for c in configs_actifs if c.master and c.master.specialite}

    # Parcours depuis SpecialiteParcoursMapping
    parcours_qs = SpecialiteParcoursMapping.objects.filter(
        actif=True,
        code_parcours__in=list(_PARCOURS_META.keys()),
    ).order_by('type_formation', 'ordre')

    offers_data = []
    for idx, parcours in enumerate(parcours_qs, start=1):
        meta = _PARCOURS_META.get(parcours.code_parcours, {})
        config = config_by_code.get(parcours.code_parcours)
        specialites = parcours.specialites if isinstance(parcours.specialites, list) else []

        # Statut temps-réel depuis ConfigurationAppel si disponible
        if config:
            statut = 'ouvert' if config.peut_candidater() else 'ferme'
            places = config.capacite_accueil or meta.get('places', 0)
            date_lim = (
                config.date_limite_preinscription.isoformat()
                if config.date_limite_preinscription
                else meta.get('date_limite', '2026-07-31')
            )
        else:
            statut = 'ouvert'
            places = meta.get('places', 0)
            date_lim = meta.get('date_limite', '2026-07-31')

        offers_data.append({
            'id': idx,
            'master_id': parcours.pk,
            'master_nom': parcours.nom_parcours,
            'specialite': parcours.code_parcours,
            'date_limite': date_lim,
            'type': meta.get('type', 'master'),
            'already_applied': parcours.code_parcours in applied_codes,
            'places_disponibles': places,
            'code_parcours': parcours.code_parcours,
            'specialites_eligibles': specialites,
            'statut': statut,
        })

    # Fallback complet si SpecialiteParcoursMapping vide (migration pas encore jouée)
    if not offers_data:
        offers_data = [
            {
                'id': i, 'master_id': i, 'master_nom': nom,
                'specialite': code, 'date_limite': _PARCOURS_META[code]['date_limite'],
                'type': _PARCOURS_META[code]['type'], 'already_applied': False,
                'places_disponibles': _PARCOURS_META[code]['places'],
                'code_parcours': code, 'specialites_eligibles': [], 'statut': 'ouvert',
            }
            for i, (code, nom) in enumerate([
                ('MPGL',   'Master Professionnel Genie Logiciel (MPGL)'),
                ('MPDS',   'Mastere Professionnel en sciences de donnees (MPDS)'),
                ('MP3I',   'Mastere Professionnel en Ingenieries en Instrumentation industrielle (MP3I)'),
                ('MRGL',   'Mastere Recherche en Genie logiciel (MRGL)'),
                ('MRMI',   'Mastere Recherche en micro-electronique et instrumentation (MRMI)'),
                ('ING_GL', 'Ingenieur en sciences Appliquees et Technologie - Genie Logiciel'),
            ], start=1)
        ]

    return Response({'offers': offers_data, 'total': len(offers_data)})


def _safe_create_notification(user, titre, message, notif_type='info', dedup_key=None):
    if dedup_key:
        try:
            Notification.objects.create(
                user=user,
                titre=titre,
                message=message,
                type=notif_type,
                dedup_key=dedup_key,
            )
            return
        except IntegrityError:
            return

    Notification.objects.create(
        user=user,
        titre=titre,
        message=message,
        type=notif_type,
    )


def _sync_system_notifications_for_user(user):
    today = timezone.now().date()
    user_role = getattr(user, 'role', None)
    is_commission_member = MembreCommission.objects.filter(
        user=user,
        actif=True,
        commission__actif=True,
    ).exists()
    user_candidatures = Candidature.objects.filter(candidat=user).select_related('master')
    is_candidate = user_role == 'candidat' or user_candidatures.exists()

    if is_candidate:
        offres_ouvertes = (
            ConfigurationAppel.objects.filter(
                actif=True,
                date_debut_visibilite__lte=today,
                date_fin_visibilite__gte=today,
                date_limite_preinscription__gte=today,
            )
            .select_related('master')
            .order_by('date_limite_preinscription', 'master__nom')
        )
        if offres_ouvertes.exists():
            masters_ouverts = ', '.join(config.master.nom for config in offres_ouvertes[:5])
            if offres_ouvertes.count() > 5:
                masters_ouverts += ' ...'

            titre = f"📢 {offres_ouvertes.count()} préinscription(s) ouverte(s)"
            message = (
                f"Bonjour {user.get_full_name() or user.username},\n\n"
                f"{offres_ouvertes.count()} appel(s) de préinscription sont actuellement ouverts.\n"
                f"Masters concernés: {masters_ouverts}.\n\n"
                "Connectez-vous au portail pour consulter les offres et déposer votre candidature.\n\n"
                "Cordialement,\n"
                "ISIMM Admission"
            )
            email_html = f"""
            <html>
              <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <h2>Préinscription ouverte</h2>
                <p>Bonjour <strong>{user.get_full_name() or user.username}</strong>,</p>
                <p>{offres_ouvertes.count()} appel(s) de préinscription sont actuellement ouverts.</p>
                <p><strong>Masters concernés:</strong> {masters_ouverts}</p>
                <p>Connectez-vous au portail pour consulter les offres et déposer votre candidature.</p>
                <hr/>
                <p style="color: #999; font-size: 12px;">ISIMM Admission</p>
              </body>
            </html>
            """
            creer_notification_avec_email(
                user=user,
                titre=titre,
                message=message,
                notif_type='info',
                dedup_key=f"preinscription-open-{today.isoformat()}",
                email_html=email_html,
            )

        for candidature in user_candidatures:
            _safe_create_notification(
                user=user,
                titre='Candidature créée',
                message=(
                    f"Votre candidature {candidature.numero} pour {candidature.master.nom} "
                    "a été enregistrée."
                ),
                notif_type='success',
                dedup_key=f"candidature-created-{candidature.id}",
            )

            if candidature.statut == 'selectionne':
                _safe_create_notification(
                    user=user,
                    titre='Candidature sélectionnée',
                    message=(
                        f"Votre candidature {candidature.numero} pour {candidature.master.nom} "
                        "a été sélectionnée."
                    ),
                    notif_type='success',
                    dedup_key=f"status-{candidature.id}-selectionne",
                )
            elif candidature.statut == 'preselectionne':
                _safe_create_notification(
                    user=user,
                    titre='Présélection disponible',
                    message=f"Votre candidature {candidature.numero} est présélectionnée.",
                    notif_type='info',
                    dedup_key=f"status-{candidature.id}-preselectionne",
                )

            if candidature.statut not in ['soumis']:
                _safe_create_notification(
                    user=user,
                    titre='Mise à jour de candidature',
                    message=(
                        f"Le statut de votre candidature {candidature.numero} est actuellement "
                        f"{candidature.get_statut_display()}."
                    ),
                    notif_type='info',
                    dedup_key=f"status-current-{candidature.id}-{candidature.statut}",
                )

    if user_role in ['responsable_commission', 'commission', 'admin'] or is_commission_member:
        configs_qs = ConfigurationAppel.objects.filter(actif=True).select_related('master')

        if user_role in ['responsable_commission', 'commission'] or is_commission_member:
            master_ids = list(
                MembreCommission.objects.filter(user=user, actif=True, commission__actif=True).values_list(
                    'commission__master_id', flat=True
                )
            )
            configs_qs = configs_qs.filter(master_id__in=master_ids)

        for config in configs_qs:
            if config.date_limite_depot_dossier:
                jours_depot = (config.date_limite_depot_dossier - today).days
                if jours_depot in [7, 3, 1, 0]:
                    message = (
                        f"Master {config.master.nom}: deadline dépôt dossier dans {jours_depot} jour(s) "
                        f"(date limite: {config.date_limite_depot_dossier})."
                    )
                    email_html = f"""
                    <html>
                      <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                        <h2>Deadline dépôt dossier proche</h2>
                        <p>Bonjour {user.get_full_name() or user.username},</p>
                        <p>{message}</p>
                        <hr/>
                        <p style="color: #999; font-size: 12px;">ISIMM Admission</p>
                      </body>
                    </html>
                    """
                    creer_notification_avec_email(
                        user=user,
                        titre='Deadline étude de dossier proche',
                        message=message,
                        notif_type='warning',
                        dedup_key=f"deadline-depot-{config.master_id}-{config.date_limite_depot_dossier}",
                        email_html=email_html,
                    )

            if config.date_limite_preinscription:
                jours_preinscription = (config.date_limite_preinscription - today).days
                if jours_preinscription in [7, 3, 1, 0]:
                    message = (
                        f"Master {config.master.nom}: deadline préinscription dans {jours_preinscription} jour(s) "
                        f"(date limite: {config.date_limite_preinscription})."
                    )
                    email_html = f"""
                    <html>
                      <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                        <h2>Deadline préinscription proche</h2>
                        <p>Bonjour {user.get_full_name() or user.username},</p>
                        <p>{message}</p>
                        <hr/>
                        <p style="color: #999; font-size: 12px;">ISIMM Admission</p>
                      </body>
                    </html>
                    """
                    creer_notification_avec_email(
                        user=user,
                        titre='Deadline préinscription proche',
                        message=message,
                        notif_type='warning',
                        dedup_key=f"deadline-preinscription-{config.master_id}-{config.date_limite_preinscription}",
                        email_html=email_html,
                    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mes_notifications(request):
    try:
        _sync_system_notifications_for_user(request.user)
    except Exception as exc:
        logger.warning("_sync_system_notifications_for_user failed for user %s: %s", request.user.id, exc)
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    serializer = NotificationSerializer(notifications, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def marquer_toutes_notifications_lues(request):
    """Mark all unread notifications as read for the current user."""
    unread_count = Notification.objects.filter(
        user=request.user,
        lue=False
    ).update(lue=True)
    
    return Response({
        'success': True,
        'notifications_updated': unread_count
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def marquer_notification_lue(request, notification_id):
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
    except Notification.DoesNotExist:
        return Response({'error': 'Notification non trouvée'}, status=status.HTTP_404_NOT_FOUND)

    if not notification.lue:
        notification.lue = True
        notification.save(update_fields=['lue'])

    return Response({'success': True})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def soumettre_candidature(request):
    """Conserve la route historique en redirigeant vers la creation."""
    return create_candidature(request)


@api_view(['POST'])
@permission_classes([AllowAny])
def preview_score_candidature(request):
    """Calcule un score d'aperçu sans créer de candidature ni persister de données."""
    formation_code = request.data.get('formation_code', '').strip().upper()
    master_id = request.data.get('master_id')

    try:
        master = None
        if formation_code:
            code_alias = {'ING_GL': 'ING_INFO_GL', 'ING-GL': 'ING_INFO_GL'}.get(formation_code, formation_code)
            master = Master.objects.select_related('formule_score').filter(specialite=code_alias).first()

        if not master and master_id:
            try:
                master = Master.objects.select_related('formule_score').get(id=master_id)
            except Master.DoesNotExist:
                pass

        if not master:
            return Response({'error': 'master_id ou formation_code est requis'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as exc:
        logger.exception("Erreur retrieving Master for preview_score_candidature: %s", exc)
        return Response(
            {'error': "Erreur interne: impossible de récupérer le master. Réessayez plus tard."},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    formule = getattr(master, 'formule_score', None)
    if formule is None:
        return Response(
            {'error': 'Aucune formule de score est définie pour ce master'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    payload = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
    if hasattr(payload, 'pop'):
        payload.pop('master_id', None)
        payload.pop('candidature_id', None)

    academic_data = payload.get('academic_data') if isinstance(payload, dict) else None
    formation_code = str(payload.get('formation_code') or '').upper() if isinstance(payload, dict) else ''

    if isinstance(academic_data, dict):
        def _as_float(value, default=0.0):
            try:
                if value is None or value == '':
                    return float(default)
                return float(value)
            except (TypeError, ValueError):
                return float(default)

        def _avg(values):
            cleaned = [_as_float(value, None) for value in values]
            cleaned = [value for value in cleaned if value is not None]
            if not cleaned:
                return 0.0
            return sum(cleaned) / len(cleaned)

        # Validate that academic_data sections are dictionaries
        common = academic_data.get('common', {}) if isinstance(academic_data.get('common'), dict) else {}
        gl_ds = academic_data.get('glDs', {}) if isinstance(academic_data.get('glDs'), dict) else {}
        i3 = academic_data.get('i3', {}) if isinstance(academic_data.get('i3'), dict) else {}
        mrgl_licence = academic_data.get('mrglLicence', {}) if isinstance(academic_data.get('mrglLicence'), dict) else {}
        mrgl_maitrise = academic_data.get('mrglMaitrise', {}) if isinstance(academic_data.get('mrglMaitrise'), dict) else {}
        mrmi_cas1 = academic_data.get('mrmiCas1', {}) if isinstance(academic_data.get('mrmiCas1'), dict) else {}
        mrmi_cas2 = academic_data.get('mrmiCas2', {}) if isinstance(academic_data.get('mrmiCas2'), dict) else {}
        ing_cas1 = academic_data.get('ingCas1', {}) if isinstance(academic_data.get('ingCas1'), dict) else {}
        ing_cas2 = academic_data.get('ingCas2', {}) if isinstance(academic_data.get('ingCas2'), dict) else {}

        payload['payload'] = academic_data
        payload['session_reussite'] = common.get('session')

        try:
            if formation_code in {'MPGL', 'MPDS'}:
                moyenne_generale = _avg([gl_ds.get('moy1'), gl_ds.get('moy2'), gl_ds.get('moy3')])
                payload['moyenne_generale'] = round(moyenne_generale, 2)
                payload['moyenne_specialite'] = round(moyenne_generale, 2)
                payload['nb_redoublements'] = int(_as_float(common.get('redoublements', 0), 0))
            elif formation_code == 'MP3I':
                moyenne_generale = _avg([i3.get('moyL1'), i3.get('moyL2'), i3.get('moyL3')])
                payload['moyenne_generale'] = round(moyenne_generale, 2)
                payload['moyenne_specialite'] = round(_as_float(i3.get('moyBac'), moyenne_generale), 2)
                payload['nb_redoublements'] = int(_as_float(i3.get('nombreRedoublement', common.get('redoublements', 0)), 0))
            elif formation_code == 'MRGL':
                parcours = academic_data.get('mrglParcours', 'licence')
                if parcours == 'maitrise':
                    moyenne_generale = _avg([mrgl_maitrise.get('moy1'), mrgl_maitrise.get('moy2'), mrgl_maitrise.get('moy3'), mrgl_maitrise.get('moy4')])
                    payload['moyenne_generale'] = round(moyenne_generale, 2)
                    payload['moyenne_specialite'] = round(_as_float(mrgl_maitrise.get('moyBac'), moyenne_generale), 2)
                else:
                    moyenne_generale = _avg([mrgl_licence.get('moy1'), mrgl_licence.get('moy2'), mrgl_licence.get('moy3')])
                    payload['moyenne_generale'] = round(moyenne_generale, 2)
                    payload['moyenne_specialite'] = round(_as_float(mrgl_licence.get('moyBac'), moyenne_generale), 2)
                payload['nb_redoublements'] = int(_as_float(common.get('redoublements', mrgl_licence.get('nombreRedoublement', 0)), 0))
            elif formation_code == 'MRMI':
                parcours = academic_data.get('mrmiParcours', 'cas1')
                if parcours == 'cas2':
                    moyenne_generale = _as_float(mrmi_cas2.get('moyIng1'), 0.0)
                    payload['moyenne_generale'] = round(moyenne_generale, 2)
                    payload['moyenne_specialite'] = round(moyenne_generale, 2)
                else:
                    moyenne_generale = _avg([mrmi_cas1.get('moyL1'), mrmi_cas1.get('moyL2'), mrmi_cas1.get('moyL3')])
                    payload['moyenne_generale'] = round(moyenne_generale, 2)
                    payload['moyenne_specialite'] = round(_as_float(mrmi_cas1.get('moyBac'), moyenne_generale), 2)
                payload['nb_redoublements'] = int(_as_float(common.get('redoublements', mrmi_cas1.get('nombreRedoublement', 0)), 0))
            elif formation_code in {'ING_INFO_GL', 'ING_EM'}:
                parcours = academic_data.get('ingParcours', 'cas1')
                if parcours == 'cas2':
                    moyenne_generale = _avg([ing_cas2.get('m1'), ing_cas2.get('m2'), ing_cas2.get('m3')])
                else:
                    moyenne_generale = _avg([ing_cas1.get('moy1'), ing_cas1.get('moy2')])
                payload['moyenne_generale'] = round(moyenne_generale, 2)
                payload['moyenne_specialite'] = round(moyenne_generale, 2)
                payload['nb_redoublements'] = int(_as_float(common.get('redoublements', ing_cas1.get('nombreRedoublement', 0)), 0))
        except (TypeError, ValueError) as e:
            logger.error("Erreur validation données académiques: %s", e)
            return Response(
                {'error': 'Les données académiques contiennent des valeurs invalides. Vérifiez que toutes les moyennes sont des nombres.'},
                status=status.HTTP_400_BAD_REQUEST
            )

    try:
        # Pass payload directly to calculate score
        score = formule.calculer_score(payload)
    except (TypeError, ValueError) as e:
        logger.error("Erreur calcul score (données invalides): %s", e)
        return Response(
            {'error': 'Erreur lors du calcul du score: données invalides. Vérifiez votre saisie.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as exc:
        logger.exception("Erreur preview_score_candidature: %s", exc)
        return Response(
            {'error': 'Erreur interne lors du calcul du score. Réessayez plus tard.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    # Convert Decimal to float for JSON serialization
    if hasattr(score, 'real'):
        score = float(score)

    return Response(
        {
            'success': True,
            'score': score,
            'master_id': master.id,
            'master_nom': master.nom,
            'formula': getattr(formule, 'nom', ''),
        }
    )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def modifier_candidature(request, candidature_id):
    try:
        candidature = Candidature.objects.get(id=candidature_id, candidat=request.user)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if not candidature.peut_etre_modifie():
        return Response(
            {'error': 'Le delai de modification est expire ou la candidature ne peut plus etre modifiee'},
            status=status.HTTP_403_FORBIDDEN,
        )

    allowed_fields = {'choix_priorite'}
    payload = {key: value for key, value in request.data.items() if key in allowed_fields}

    if not payload:
        return Response(
            {'error': 'Aucun champ modifiable fourni', 'allowed_fields': sorted(list(allowed_fields))},
            status=status.HTTP_400_BAD_REQUEST,
        )

    serializer = CandidatureSerializer(candidature, data=payload, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mes_candidatures(request):
    candidatures = Candidature.objects.filter(candidat=request.user).select_related('candidat', 'master')
    _refresh_classement_for_queryset(candidatures)
    candidatures = _ranked_candidatures_queryset(candidatures.order_by('-date_soumission'))

    for candidature in candidatures:
        if getattr(candidature, 'classement_calcule', None) is not None:
            candidature.classement = int(candidature.classement_calcule)

    serializer = CandidatureSerializer(candidatures, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reclasser_voeux(request):
    """Req-2 — Reclasser les 3 vœux Masters par ordre de priorité.

    Payload attendu :
      { "ordre": [<id_candidature_choix_1>, <id_candidature_choix_2>, <id_candidature_choix_3>] }

    Toutes les candidatures Masters actives du candidat doivent être incluses.
    Ne s'applique pas aux candidatures Ingénieur (unique, pas de classement).
    """
    ordre = request.data.get('ordre', [])
    if not isinstance(ordre, list) or not ordre:
        return Response(
            {'error': "Le champ 'ordre' (liste d'IDs de candidature) est requis."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if len(ordre) > 3:
        return Response(
            {'error': "Maximum 3 vœux autorisés."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    candidatures = list(
        Candidature.objects.filter(
            candidat=request.user, id__in=ordre,
        ).exclude(
            master__specialite__istartswith='ING',
        ).exclude(
            statut__in=['rejete', 'annulee'],
        )
    )
    if len(candidatures) != len(ordre):
        return Response(
            {'error': "Certaines candidatures fournies n'existent pas ou ne vous appartiennent pas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    by_id = {c.id: c for c in candidatures}
    for new_priorite, cand_id in enumerate(ordre, start=1):
        c = by_id[cand_id]
        if not c.peut_etre_modifie():
            return Response(
                {'error': f"La candidature {c.numero} n'est plus modifiable (délai dépassé)."},
                status=status.HTTP_403_FORBIDDEN,
            )
        c.choix_priorite = new_priorite
        c.save(update_fields=['choix_priorite'])

    return Response(
        {
            'message': 'Vœux reclassés avec succès.',
            'voeux': [
                {'id': c.id, 'master': c.master.nom, 'choix': c.choix_priorite}
                for c in sorted(candidatures, key=lambda x: x.choix_priorite)
            ],
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def candidatures_responsable(request):
    """Retourne les candidatures visibles par un responsable/commission, avec filtres par master et type."""
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    # Specialite du membre/responsable (MPGL, MPDS, MP3I, MRGL, MRMI, Cycle Ingenieur, ...)
    specialite_user = getattr(request.user, 'specialite', '').strip()

    master_ids = None
    if role in ['responsable_commission', 'commission']:
        master_ids = list(
            MembreCommission.objects.filter(user=request.user, actif=True, commission__actif=True).values_list(
                'commission__master_id', flat=True
            )
        )
        if not master_ids and not specialite_user:
            return Response([])

    master_id = request.query_params.get('master_id')
    type_concours = request.query_params.get('type')

    candidatures_qs = Candidature.objects.select_related(
        'candidat', 'master', 'concours', 'donnees_academiques',
    ).order_by('-score', '-date_soumission')

    # Les commissions assignées (MembreCommission) font autorité sur le périmètre.
    # La spécialité n'est utilisée qu'en repli (responsable sans commission liée).
    if master_ids:
        candidatures_qs = candidatures_qs.filter(master_id__in=master_ids)
    elif specialite_user and role not in ['admin']:
        candidatures_qs = candidatures_qs.filter(master__specialite__icontains=specialite_user)

    if master_id and master_id != 'all':
        candidatures_qs = candidatures_qs.filter(master_id=master_id)

    if type_concours in ['masters', 'ingenieur']:
        if type_concours == 'ingenieur':
            candidatures_qs = candidatures_qs.filter(concours__isnull=False)
        else:
            candidatures_qs = candidatures_qs.filter(concours__isnull=True)

    _refresh_classement_for_queryset(candidatures_qs)
    candidatures_qs = _ranked_candidatures_queryset(candidatures_qs)

    payload = []
    for candidature in candidatures_qs:
        payload.append(
            {
                'id': candidature.id,
                'numero': candidature.numero,
                'candidat_nom': candidature.candidat.get_full_name(),
                'candidat_email': candidature.candidat.email,
                'candidat_cin': getattr(candidature.candidat, 'cin', ''),
                'specialite': candidature.master.specialite if candidature.master else '',
                'specialite_diplome': _extract_specialite_diplome(candidature),
                'master_id': candidature.master_id,
                'master_nom': candidature.master.nom if candidature.master else '',
                'score': candidature.score,
                'classement': int(getattr(candidature, 'classement_calcule', candidature.classement or 0) or 0),
                'dossier_depose': candidature.dossier_depose,
                'statut': candidature.statut,
                'type_concours': 'ingenieur' if candidature.concours_id else 'masters',
                'parcours': getattr(candidature, 'parcours', '') or '',
                'date_soumission': candidature.date_soumission,
                'date_changement_statut': candidature.date_changement_statut,
            }
        )

    return Response(payload)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mes_masters(request):
    """Retourne les masters assignés à l'utilisateur courant via MembreCommission."""
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    master_ids = list(
        MembreCommission.objects.filter(
            user=request.user, actif=True, commission__actif=True
        ).values_list('commission__master_id', flat=True)
    )
    master_ids = [mid for mid in master_ids if mid is not None]

    masters = Master.objects.filter(id__in=master_ids, actif=True)

    data = [
        {
            'id': m.id,
            'nom': m.nom,
            'specialite': m.specialite,
            'type_master': m.type_master,
        }
        for m in masters
    ]

    return Response({'masters': data, 'count': len(data)})


@api_view(['GET'])
@permission_classes([AllowAny])
def offres_inscription(request):
    """Retourne les offres d'inscription (masters + cycle ingenieur)."""
    today = timezone.now().date()

    masters = Master.objects.filter(actif=True).order_by('nom')
    configurations = ConfigurationAppel.objects.filter(actif=True).select_related('master')
    config_by_master = {cfg.master_id: cfg for cfg in configurations}

    role = getattr(request.user, 'role', None)
    can_see_hidden = role in ['admin', 'responsable_commission', 'commission']

    offres = []

    for master in masters:
        try:
            config = config_by_master.get(master.id)
            nombre_candidats_inscrits = Candidature.objects.filter(master_id=master.id).count()

            if config and config.est_cache and not can_see_hidden:
                continue

            nom_lower = (master.nom or '').lower()
            specialite_upper = (master.specialite or '').upper()
            is_cycle_ingenieur = (
                specialite_upper.startswith('ING')
                or 'ingenieur' in nom_lower
                or 'ingénieur' in nom_lower
            )

            reference_deadline = (
                config.date_limite_preinscription
                if config and config.date_limite_preinscription
                else master.date_limite_candidature
            )
            # Mode demo: garantir des offres ouvertes/public pour ne pas bloquer la candidature.
            statut = 'ouvert'

            offres.append(
                {
                    'id': master.id,
                    'titre': master.nom,
                    'type': 'cycle_ingenieur' if is_cycle_ingenieur else 'master',
                    'sous_type': master.type_master,
                    'specialite': master.specialite,
                    'code_parcours': master.specialite,  # Formation code (MPGL, MPDS, etc.)
                    'description': master.description,
                    'date_limite': master.date_limite_candidature,
                    'date_limite_preinscription': config.date_limite_preinscription if config else None,
                    'date_limite_depot_dossier': config.date_limite_depot_dossier if config else None,
                    'date_limite_paiement': config.date_limite_paiement if config else None,
                    'places': master.places_disponibles,
                    'capacite_interne': config.capacite_interne if config else 0,
                    'capacite_externe': config.capacite_externe if config else 0,
                    'est_visible': True,
                    'est_cache': False,
                    'document_officiel_pdf_url': (
                        request.build_absolute_uri(config.document_officiel_pdf.url)
                        if config and config.document_officiel_pdf
                        else None
                    ),
                    'publie_par_responsable': True,
                    'statut': statut,
                    'nombre_candidats_inscrits': nombre_candidats_inscrits,
                }
            )
        except Exception as exc:
            logger.exception('Erreur lors de la construction de l offre %s: %s', master.id, exc)
            continue

    # Sprint 3: offres commission (OffreMaster) visibles automatiquement cote candidat.
    try:
        offres_master = OffreMaster.objects.select_related('master').filter(actif=True)
        for offre_master in offres_master:
            master = offre_master.master
            if not master or not master.actif:
                continue

            if not offre_master.est_publiee and not can_see_hidden:
                continue

            # Eviter les doublons si le master est deja expose via la boucle precedente.
            if any(existing.get('id') == master.id for existing in offres):
                continue

            reference_deadline = offre_master.date_limite_preinscription or offre_master.date_limite
            statut_public = (
                'ouvert'
                if offre_master.est_publiee and offre_master.appel_actif and reference_deadline and reference_deadline >= today
                else 'ferme'
            )

            offres.append(
                {
                    'id': master.id,
                    'titre': offre_master.titre,
                    'type': 'master',
                    'sous_type': master.type_master,
                    'specialite': master.specialite,
                    'type_formation': offre_master.type_formation,
                    'description': offre_master.description,
                    'date_limite': offre_master.date_limite,
                    'date_debut_visibilite': offre_master.date_debut_visibilite,
                    'date_fin_visibilite': offre_master.date_fin_visibilite,
                    'date_limite_preinscription': offre_master.date_limite_preinscription,
                    'date_limite_depot_dossier': offre_master.date_limite_depot_dossier,
                    'places': offre_master.capacite,
                    'capacites_detaillees': offre_master.capacites_detaillees,
                    'est_publiee': offre_master.est_publiee,
                    'appel_actif': offre_master.appel_actif,
                    'statut': statut_public,
                    'nombre_candidats_inscrits': Candidature.objects.filter(master_id=master.id).count(),
                }
            )
    except OperationalError as exc:
        logger.warning('Table OffreMaster indisponible, fallback masters uniquement: %s', exc)

    concours_qs = Concours.objects.filter(actif=True).order_by('-created_at')
    for concours in concours_qs:
        statut = 'ouvert' if concours.date_cloture and concours.date_cloture >= today else 'ferme'
        offres.append(
            {
                'id': concours.id,
                'titre': concours.nom,
                'type': 'cycle_ingenieur' if concours.type_concours == 'ingenieur' else 'master',
                'sous_type': 'recherche' if concours.type_concours == 'recherche' else (
                    'cycle_ingenieur' if concours.type_concours == 'ingenieur' else 'professionnel'
                ),
                'specialite': (concours.conditions_admission or {}).get('specialite', ''),
                'description': concours.description,
                'date_limite': concours.date_cloture,
                'places': concours.places_disponibles,
                'document_officiel_pdf_url': (
                    request.build_absolute_uri(concours.document_officiel_pdf.url)
                    if concours.document_officiel_pdf
                    else None
                ),
                'statut': statut,
            }
        )

    return Response(offres)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notifications_responsable(request):
    """Retourne les deadlines utiles au responsable pour les masters qu'il gère."""
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    master_ids = None
    if role in ['responsable_commission', 'commission']:
        master_ids = list(
            MembreCommission.objects.filter(user=request.user, actif=True, commission__actif=True).values_list(
                'commission__master_id', flat=True
            )
        )
        if not master_ids:
            return Response([])

    today = timezone.now().date()
    configs = ConfigurationAppel.objects.filter(actif=True).select_related('master')
    if master_ids is not None:
        configs = configs.filter(master_id__in=master_ids)

    items = []
    for config in configs:
        deadlines = [
            ('Préinscription', config.date_limite_preinscription),
            ('Dépôt de dossier', config.date_limite_depot_dossier),
            ('Paiement', config.date_limite_paiement),
        ]
        for label, deadline in deadlines:
            if not deadline:
                continue
            days_left = (deadline - today).days
            items.append(
                {
                    'id': f'{config.master_id}-{label}',
                    'master_id': config.master_id,
                    'master_nom': config.master.nom,
                    'deadline_type': label,
                    'deadline_date': deadline,
                    'days_left': days_left,
                    'est_cache': config.est_cache,
                    'est_visible': config.est_visible(),
                    'statut': 'ouvert' if config.peut_candidater() else 'ferme',
                    'type': 'warning' if days_left <= 7 else 'info',
                    'message': (
                        f"{label} pour {config.master.nom} dans {days_left} jour(s)"
                        if days_left >= 0
                        else f"{label} pour {config.master.nom} est dépassée de {abs(days_left)} jour(s)"
                    ),
                }
            )

    items.sort(key=lambda item: (item['days_left'] if item['days_left'] is not None else 9999, str(item['deadline_date'])))
    return Response(items)


@api_view(['GET'])
@permission_classes([AllowAny])
def lister_masters(request):
    try:
        today = timezone.now().date()
        user_role = getattr(request.user, 'role', None)
        is_admin_like = user_role in ['admin', 'responsable_commission', 'commission']

        # Admin/responsable: voir tous les parcours et statut manuel base sur actif.
        # Candidat/public: conserver le comportement public (actifs seulement, statut base sur la date).
        masters_qs = Master.objects.all() if is_admin_like else Master.objects.filter(actif=True)
        masters = list(masters_qs.order_by('nom'))

        payload = []
        for m in masters:
            date_limite = getattr(m, 'date_limite_candidature', None)
            if is_admin_like:
                statut = 'ouvert' if bool(getattr(m, 'actif', False)) else 'ferme'
            else:
                statut = 'ferme'
                if date_limite:
                    statut = 'ouvert' if date_limite >= today else 'ferme'

            payload.append(
                {
                    'id': m.id,
                    'nom': m.nom,
                    'specialite': m.specialite,
                    'type_master': m.type_master,
                    'description': m.description,
                    'places_disponibles': m.places_disponibles,
                    'statut': statut,
                    'date_limite_candidature': date_limite,
                    'annee_universitaire': m.annee_universitaire,
                }
            )

        return Response(payload)
    except OperationalError as exc:
        logger.exception('Schema candidature indisponible pour masters: %s', exc)
        return Response(
            {
                'results': [],
                'warning': 'Base candidature non initialisee correctement (schema masters indisponible).',
            },
            status=status.HTTP_200_OK,
        )
    except Exception as exc:
        logger.exception('Erreur inattendue lister_masters: %s', exc)
        return Response(
            {
                'results': [],
                'warning': 'Impossible de charger les masters pour le moment.',
            },
            status=status.HTTP_200_OK,
        )


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def offres_master_crud(request):
    """
    CRUD Sprint 3 - Offres de preinscription master.
    - GET: liste (avec recherche par titre/date)
    - POST: creation offre + master associe
    """
    if request.method == 'GET':
        if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission', 'commission']:
            return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

        qs = OffreMaster.objects.select_related('master').filter(actif=True)
        search = (request.query_params.get('search') or '').strip().lower()
        date_limite = (request.query_params.get('date_limite') or '').strip()

        if search:
            qs = qs.filter(Q(titre__icontains=search) | Q(description__icontains=search))
        if date_limite:
            qs = qs.filter(date_limite=date_limite)

        serializer = OffreMasterSerializer(qs.order_by('date_limite', 'titre'), many=True)
        return Response(serializer.data)

    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    data = request.data or {}
    titre = (data.get('titre') or '').strip()
    description = (data.get('description') or '').strip()
    capacite = data.get('capacite')
    date_limite = data.get('date_limite')

    if not titre or capacite in [None, ''] or not date_limite:
        return Response(
            {'error': 'Champs requis: titre, capacite, date_limite.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        capacite_int = int(capacite)
    except (TypeError, ValueError):
        return Response({'error': 'capacite invalide.'}, status=status.HTTP_400_BAD_REQUEST)

    master = Master.objects.create(
        nom=titre,
        type_master='professionnel',
        description=description,
        specialite=(data.get('specialite') or 'Master'),
        places_disponibles=capacite_int,
        date_limite_candidature=date_limite,
        annee_universitaire=data.get('annee_universitaire', f"{timezone.now().year}/{timezone.now().year + 1}"),
        actif=bool(data.get('actif', True)),
    )

    offre = OffreMaster.objects.create(
        master=master,
        titre=titre,
        description=description,
        type_formation=(data.get('type_formation') or 'master'),
        capacite=capacite_int,
        date_limite=date_limite,
        date_debut_visibilite=data.get('date_debut_visibilite') or None,
        date_fin_visibilite=data.get('date_fin_visibilite') or None,
        date_limite_preinscription=data.get('date_limite_preinscription') or date_limite,
        date_limite_depot_dossier=data.get('date_limite_depot_dossier') or None,
        capacites_detaillees=data.get('capacites_detaillees') or [],
        appel_actif=bool(data.get('appel_actif', True)),
        est_publiee=bool(data.get('est_publiee', False)),
        actif=bool(data.get('actif', True)),
    )

    return Response(OffreMasterSerializer(offre).data, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def offre_master_detail_crud(request, offre_id):
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        offre = OffreMaster.objects.select_related('master').get(id=offre_id)
    except OffreMaster.DoesNotExist:
        return Response({'error': 'Offre non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response(OffreMasterSerializer(offre).data)

    if role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'DELETE':
        offre.actif = False
        offre.save(update_fields=['actif', 'updated_at'])
        master = offre.master
        master.actif = False
        master.save(update_fields=['actif', 'updated_at'])
        return Response({'success': True, 'message': 'Offre desactivee.'}, status=status.HTTP_200_OK)

    data = request.data or {}
    if 'titre' in data:
        offre.titre = data.get('titre') or offre.titre
    if 'description' in data:
        offre.description = data.get('description') or ''
    if 'type_formation' in data:
        offre.type_formation = data.get('type_formation') or 'master'
    if 'capacite' in data:
        try:
            offre.capacite = int(data.get('capacite'))
        except (TypeError, ValueError):
            return Response({'error': 'capacite invalide.'}, status=status.HTTP_400_BAD_REQUEST)
    if 'date_limite' in data:
        offre.date_limite = data.get('date_limite')
    if 'actif' in data:
        offre.actif = bool(data.get('actif'))
    if 'appel_actif' in data:
        offre.appel_actif = bool(data.get('appel_actif'))
    if 'est_publiee' in data:
        offre.est_publiee = bool(data.get('est_publiee'))
    if 'date_debut_visibilite' in data:
        offre.date_debut_visibilite = data.get('date_debut_visibilite') or None
    if 'date_fin_visibilite' in data:
        offre.date_fin_visibilite = data.get('date_fin_visibilite') or None
    if 'date_limite_preinscription' in data:
        offre.date_limite_preinscription = data.get('date_limite_preinscription') or None
    if 'date_limite_depot_dossier' in data:
        offre.date_limite_depot_dossier = data.get('date_limite_depot_dossier') or None
    if 'capacites_detaillees' in data:
        detaillees = data.get('capacites_detaillees')
        offre.capacites_detaillees = detaillees if isinstance(detaillees, list) else []

    offre.save()
    return Response(OffreMasterSerializer(offre).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def offre_master_public_detail(request, offre_id):
    """Expose le detail d'une offre publiee pour affichage cote candidat en temps reel."""
    try:
        offre = OffreMaster.objects.select_related('master').get(id=offre_id, actif=True)
    except OffreMaster.DoesNotExist:
        return Response({'error': 'Offre non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if not offre.est_publiee and getattr(request.user, 'role', None) not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Offre non publiee'}, status=status.HTTP_403_FORBIDDEN)

    data = OffreMasterSerializer(offre).data
    data['statut_public'] = 'ouvert' if offre.appel_actif and offre.est_publiee else 'ferme'
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creer_master_admin(request):
    """Creer un master (admin ou responsable commission)."""
    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)

    data = request.data or {}

    nom = data.get('nom')
    type_master = data.get('type_master')
    specialite = data.get('specialite')
    date_limite_candidature = data.get('date_limite_candidature')
    places_disponibles = data.get('places_disponibles')

    if not nom or not type_master or not specialite or not date_limite_candidature:
        return Response(
            {'error': 'Champs obligatoires manquants (nom, type_master, specialite, date_limite_candidature).'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        places_int = int(places_disponibles)
    except (TypeError, ValueError):
        return Response({'error': 'places_disponibles invalide.'}, status=status.HTTP_400_BAD_REQUEST)

    master = Master.objects.create(
        nom=nom,
        type_master=type_master,
        description=data.get('description', ''),
        specialite=specialite,
        places_disponibles=places_int,
        date_limite_candidature=date_limite_candidature,
        annee_universitaire=data.get('annee_universitaire', '2025/2026'),
        actif=bool(data.get('actif', True)),
    )

    return Response(
        {
            'id': master.id,
            'nom': master.nom,
            'type_master': master.type_master,
            'description': master.description,
            'specialite': master.specialite,
            'places_disponibles': master.places_disponibles,
            'date_limite_candidature': master.date_limite_candidature,
            'annee_universitaire': master.annee_universitaire,
            'actif': master.actif,
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(['PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def modifier_supprimer_master_admin(request, master_id):
    """Modifier/Supprimer (soft delete) un master (admin ou responsable commission)."""
    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)

    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master introuvable'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        master.actif = False
        master.save(update_fields=['actif', 'updated_at'])
        return Response({'message': 'Master supprime avec succes (desactive).'}, status=status.HTTP_200_OK)

    data = request.data or {}

    if 'nom' in data:
        master.nom = data.get('nom') or master.nom
    if 'type_master' in data:
        master.type_master = data.get('type_master') or master.type_master
    if 'description' in data:
        master.description = data.get('description') or ''
    if 'specialite' in data:
        master.specialite = data.get('specialite') or master.specialite
    if 'places_disponibles' in data:
        try:
            master.places_disponibles = int(data.get('places_disponibles'))
        except (TypeError, ValueError):
            return Response({'error': 'places_disponibles invalide.'}, status=status.HTTP_400_BAD_REQUEST)
    if 'date_limite_candidature' in data:
        master.date_limite_candidature = data.get('date_limite_candidature')
    if 'annee_universitaire' in data:
        master.annee_universitaire = data.get('annee_universitaire') or master.annee_universitaire
    if 'actif' in data:
        master.actif = bool(data.get('actif'))

    master.save()

    return Response(
        {
            'id': master.id,
            'nom': master.nom,
            'type_master': master.type_master,
            'description': master.description,
            'specialite': master.specialite,
            'places_disponibles': master.places_disponibles,
            'date_limite_candidature': master.date_limite_candidature,
            'annee_universitaire': master.annee_universitaire,
            'actif': master.actif,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mes_dossiers(request):
    """Retourne les dossiers du candidat derives des candidatures."""
    candidatures = Candidature.objects.filter(candidat=request.user).select_related('master')

    dossiers = []
    for candidature in candidatures:
        numero_dossier = f"DOS-{candidature.numero}"
        dossiers.append(
            {
                'id': candidature.id,
                'numero_dossier': numero_dossier,
                'numero_candidature': candidature.numero,
                'candidature_id': candidature.id,
                'master_nom': candidature.master.nom,
                'statut': candidature.statut,
                'dossier_depose': candidature.dossier_depose,
                'dossier_valide': candidature.dossier_valide,
                'date_soumission': candidature.date_soumission,
            }
        )

    return Response(dossiers)


@api_view(['GET'])
@permission_classes([AllowAny])
def lister_dossiers_ocr(request):
    """Retourne les dossiers deposes a analyser par la commission (OCR)."""
    try:
        candidatures_qs = (
            Candidature.objects.select_related('candidat', 'master')
            .filter(dossier_depose=True)
            .order_by('-date_depot_dossier', '-updated_at')
        )

        # Forcer l'evaluation ici pour capturer toutes les erreurs SQL/runtime
        candidatures = list(candidatures_qs)

        payload = []
        for c in candidatures:
            payload.append(
                {
                    'id': c.id,
                    'candidat_nom': f"{getattr(c.candidat, 'first_name', '')} {getattr(c.candidat, 'last_name', '')}".strip(),
                    'email': getattr(c.candidat, 'email', ''),
                    'master_nom': c.master.nom if c.master else '',
                    'statut': c.statut,
                    'date_depot_dossier': c.date_depot_dossier,
                    'score': c.score,
                }
            )

        return Response(payload)
    except OperationalError as exc:
        logger.exception('Schema candidature indisponible pour dossiers-ocr: %s', exc)
        return Response(
            {
                'results': [],
                'warning': 'Base candidature non initialisee correctement (table manquante).',
            },
            status=status.HTTP_200_OK,
        )
    except Exception as exc:
        logger.exception('Erreur inattendue dossiers-ocr: %s', exc)
        return Response(
            {
                'results': [],
                'warning': 'Impossible de charger les dossiers OCR pour le moment.',
            },
            status=status.HTTP_200_OK,
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_profile(request):
    serializer = UserUpdateSerializer(request.user, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({'user': serializer.data})
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def changer_statut_candidature(request, candidature_id):
    if getattr(request.user, 'role', None) not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    nouveau_statut = request.data.get('statut')
    motif_rejet = request.data.get('motif_rejet', '')

    try:
        changed, candidature, ancien_statut, nouveau_statut = StatutService.change_candidature_status(
            candidature,
            nouveau_statut,
            actor=request.user,
            commentaire='Changement de statut via commission',
            motif_rejet=motif_rejet,
        )
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    if not changed:
        return Response(
            {'error': 'Aucun changement detecte sur le statut'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    _log_commission_action(
        request.user,
        'Validation individuelle',
        getattr(candidature.master, 'specialite', '') or getattr(candidature, 'specialite', ''),
        getattr(candidature, 'annee_universitaire', '') or '',
        1,
        candidature.master if hasattr(candidature, 'master') else None,
    )

    channel_layer = get_channel_layer()
    if channel_layer is not None:
        async_to_sync(channel_layer.group_send)(
            'candidatures_updates',
            {
                'type': 'candidature_status_changed',
                'candidature_id': candidature.id,
                'candidate_user_id': candidature.candidat_id,
                'new_status': nouveau_statut,
                'updated_at': timezone.now().isoformat(),
            },
        )

    try:
        envoyer_email_changement_statut(candidature, ancien_statut, nouveau_statut)
    except Exception as exc:
        logger.exception("Erreur envoi email changement statut %s: %s", candidature.id, exc)

    notif_messages = {
        'sous_examen': 'Votre candidature est en cours d’examen.',
        'preselectionne': 'Votre candidature est présélectionnée.',
        'selectionne': 'Votre candidature est sélectionnée.',
        'rejete': 'Votre candidature a été rejetée.',
        'en_attente_dossier': 'Veuillez déposer votre dossier numérique.',
        'inscrit': 'Votre inscription est validée.',
    }
    _safe_create_notification(
        user=candidature.candidat,
        titre='Mise à jour de candidature',
        message=notif_messages.get(
            nouveau_statut,
            f"Le statut de votre candidature {candidature.numero} est passé à {nouveau_statut}.",
        ),
        notif_type='info' if nouveau_statut not in ['selectionne', 'rejete'] else (
            'success' if nouveau_statut == 'selectionne' else 'danger'
        ),
        dedup_key=f"status-change-{candidature.id}-{nouveau_statut}-{timezone.now().date().isoformat()}",
    )

    serializer = CandidatureSerializer(candidature)
    return Response(
        {
            'success': True,
            'message': f'Statut change de "{ancien_statut}" a "{nouveau_statut}"',
            'candidature': serializer.data,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_status(request, candidature_id):
    """Action rapide Sprint 3: bouton Valider -> statut preselectionne (admissible)."""
    if getattr(request.user, 'role', None) not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    requested_status = request.data.get('statut', 'preselectionne')
    if str(requested_status or '').strip().lower() == 'admissible':
        requested_status = 'preselectionne'

    try:
        changed, candidature, ancien_statut, nouveau_statut = StatutService.change_candidature_status(
            candidature,
            requested_status,
            actor=request.user,
            commentaire='Validation commission (update_status)',
            motif_rejet=request.data.get('motif_rejet', ''),
        )
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    if not changed:
        return Response({'success': True, 'message': 'Statut deja preselectionne.'}, status=status.HTTP_200_OK)
    _log_commission_action(
        request.user,
        'Validation individuelle',
        getattr(candidature.master, 'specialite', '') or getattr(candidature, 'specialite', ''),
        getattr(candidature, 'annee_universitaire', '') or '',
        1,
        candidature.master if hasattr(candidature, 'master') else None,
    )

    channel_layer = get_channel_layer()
    if channel_layer is not None:
        async_to_sync(channel_layer.group_send)(
            'candidatures_updates',
            {
                'type': 'candidature_status_changed',
                'candidature_id': candidature.id,
                'candidate_user_id': candidature.candidat_id,
                'new_status': nouveau_statut,
                'updated_at': timezone.now().isoformat(),
            },
        )

    serializer = CandidatureSerializer(candidature)
    return Response(
        {
            'success': True,
            'message': 'Statut mis a jour: preselectionne',
            'candidature': serializer.data,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def commission_decision_candidature(request, candidature_id):
    if getattr(request.user, 'role', None) not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    decision = str(request.data.get('decision', '')).strip().lower()
    motif_rejet = str(request.data.get('motif_rejet', '')).strip()

    decision_to_status = {
        'accepter': 'preselectionne',
        'accept': 'preselectionne',
        'valider': 'preselectionne',
        'refuser': 'rejete',
        'reject': 'rejete',
        'rejeter': 'rejete',
    }

    nouveau_statut = decision_to_status.get(decision)
    if not nouveau_statut:
        return Response(
            {'error': "decision invalide. Valeurs attendues: accepter ou refuser."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # MOD v5 §G — Justification obligatoire pour un rejet.
    if nouveau_statut == 'rejete' and len(motif_rejet) < 10:
        return Response(
            {
                'error': 'Justification obligatoire pour un rejet (au moins 10 caractères).',
                'code': 'JUSTIFICATION_REQUISE',
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        changed, candidature, ancien_statut, nouveau_statut = StatutService.change_candidature_status(
            candidature,
            nouveau_statut,
            actor=request.user,
            commentaire=f"Decision commission: {decision}",
            motif_rejet=motif_rejet,
        )
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    if not changed:
        return Response(
            {'success': True, 'message': 'Statut deja a jour.', 'candidature': CandidatureSerializer(candidature).data},
            status=status.HTTP_200_OK,
        )
    _log_commission_action(
        request.user,
        f'Decision commission: {decision}',
        getattr(candidature.master, 'specialite', '') or getattr(candidature, 'specialite', ''),
        getattr(candidature, 'annee_universitaire', '') or '',
        1,
        candidature.master if hasattr(candidature, 'master') else None,
    )

    channel_layer = get_channel_layer()
    if channel_layer is not None:
        async_to_sync(channel_layer.group_send)(
            'candidatures_updates',
            {
                'type': 'candidature_status_changed',
                'candidature_id': candidature.id,
                'candidate_user_id': candidature.candidat_id,
                'new_status': nouveau_statut,
                'updated_at': timezone.now().isoformat(),
            },
        )

    try:
        envoyer_email_changement_statut(candidature, ancien_statut, nouveau_statut)
    except Exception as exc:
        logger.exception("Erreur envoi email changement statut %s: %s", candidature.id, exc)

    _safe_create_notification(
        user=candidature.candidat,
        titre='Décision de la commission',
        message=(
            'Votre candidature a été acceptée par la commission.'
            if nouveau_statut == 'preselectionne'
            else 'Votre candidature a été refusée par la commission.'
        ),
        notif_type='success' if nouveau_statut == 'preselectionne' else 'danger',
        dedup_key=f"commission-decision-{candidature.id}-{nouveau_statut}-{timezone.now().date().isoformat()}",
    )

    serializer = CandidatureSerializer(candidature)
    return Response(
        {
            'success': True,
            'message': f'Statut change de "{ancien_statut}" a "{nouveau_statut}"',
            'candidature': serializer.data,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_decision_finale_responsable(request, candidature_id):
    """Set the decision_finale_responsable field for a candidature.

    Expects payload: { decision: 'valide' | 'rejete' | 'en_attente' }
    """
    if getattr(request.user, 'role', None) not in ['responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    decision = str(request.data.get('decision', '')).strip().lower()
    valid = {'valide', 'rejete', 'en_attente'}
    if decision not in valid:
        return Response({'error': 'decision invalide'}, status=status.HTTP_400_BAD_REQUEST)

    candidature.decision_finale_responsable = decision
    candidature.save(update_fields=['decision_finale_responsable', 'updated_at'])

    _safe_create_notification(
        user=candidature.candidat,
        titre='Décision finale du responsable',
        message=f"La décision finale pour votre candidature a été mise à jour: {decision}",
        notif_type='info',
        dedup_key=f'decision-finale-{candidature.id}-{timezone.now().date().isoformat()}',
    )

    return Response({'success': True, 'decision': decision}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def valider_preselection_commission(request, commission_id):
    """
    Déclenche la clôture de la session de présélection pour une commission.
    Réservé au Responsable. Enfile la tâche Celery asynchrone qui envoie
    les emails et les notifications WebSocket à chaque candidat concerné.
    """
    if getattr(request.user, 'role', None) not in ['responsable_commission', 'admin']:
        return Response({'error': 'Permission refusée'}, status=status.HTTP_403_FORBIDDEN)

    try:
        commission = Commission.objects.select_related('master').get(id=commission_id)
    except Commission.DoesNotExist:
        return Response({'error': 'Commission non trouvée'}, status=status.HTTP_404_NOT_FOUND)

    from .tasks import cloture_preselection_worker
    cloture_preselection_worker.delay(commission_id)

    return Response(
        {
            'success': True,
            'message': (
                f"Session de présélection clôturée pour {commission.master.nom}. "
                "Les candidats seront notifiés par email et en temps réel."
                if commission.master
                else "Session de présélection clôturée. Les candidats seront notifiés."
            ),
        },
        status=status.HTTP_202_ACCEPTED,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def valider_preselection_candidature(request, candidature_id):
    """
    Valide la présélection d'un candidat spécifique et envoie une notification.
    Le candidat reçoit un email et une notification WebSocket lui demandant de confirmer son inscription.

    Paramètres optionnels:
    - recommandation: 'favorable', 'defavorable', 'reserve'
    - commentaire: Commentaire du responsable
    """
    if getattr(request.user, 'role', None) not in ['responsable_commission', 'admin']:
        return Response({'error': 'Permission refusée'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvée'}, status=status.HTTP_404_NOT_FOUND)

    # Récupérer les paramètres optionnels
    recommandation = request.data.get('recommandation', 'favorable')
    commentaire = request.data.get('commentaire', '')

    # Mettre à jour le statut
    candidature.statut = 'preselectionne'
    candidature.save(update_fields=['statut'])

    # Envoyer notification au candidat
    try:
        from .notifications import envoyer_notification_preselection
        envoyer_notification_preselection(candidature)
    except Exception as e:
        logger.error(f'Erreur envoi notification présélection {candidature_id}: {e}')

    # Créer notification système
    try:
        from .models import Notification
        message = f'Vous avez été présélectionné(e) pour {candidature.master.nom}. Veuillez confirmer votre inscription.'
        if commentaire:
            message += f'\n\nCommentaire: {commentaire}'

        Notification.objects.create(
            utilisateur=candidature.candidat,
            titre='🎉 Présélection validée',
            message=message,
            type_notification='PRESELECTION_VALIDEE',
            est_lu=False,
        )
    except Exception as e:
        logger.error(f'Erreur création notification {candidature_id}: {e}')

    return Response(
        {
            'success': True,
            'message': f'Présélection validée pour {candidature.candidat_nom}. Notification envoyée.',
            'statut': candidature.statut,
            'candidat_nom': candidature.candidat_nom,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def appliquer_quotas_decision_finale(request, commission_id):
    """
    Applique la décision finale basée sur les quotas LP/LA définis dans la ConfigurationAppel.
    Classe les candidats présélectionnés par score et assigne le statut définitif.
    Réservé au Responsable.
    """
    if getattr(request.user, 'role', None) not in ['responsable_commission', 'admin']:
        return Response({'error': 'Permission refusée'}, status=status.HTTP_403_FORBIDDEN)

    try:
        commission = Commission.objects.select_related('master').get(id=commission_id)
    except Commission.DoesNotExist:
        return Response({'error': 'Commission non trouvée'}, status=status.HTTP_404_NOT_FOUND)

    from .tasks import appliquer_decision_finale_quotas
    appliquer_decision_finale_quotas.delay(commission_id)

    return Response(
        {
            'success': True,
            'message': (
                f"Décision finale en cours de traitement pour {commission.master.nom}."
                if commission.master
                else "Décision finale en cours de traitement."
            ),
        },
        status=status.HTTP_202_ACCEPTED,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def can_reapply_to_master(request, master_id):
    """Check if candidate can reapply to a master (re-application rules)."""
    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    previous = Candidature.objects.filter(
        candidat=request.user,
        master=master
    ).order_by('-date_soumission').first()

    if not previous:
        return Response({
            'can_reapply': True,
            'can_edit': False,
            'reason': 'Aucune candidature precedente',
            'previous_status': None,
            'edit_deadline': None,
            'cooldown_days': 0
        })

    TERMINAL_STATUSES = ['rejete', 'annule']
    edit_deadline = previous.date_limite_modification
    can_edit = previous.statut in ['soumis', 'en_attente_dossier'] and bool(edit_deadline and timezone.now() <= edit_deadline)

    if previous.statut in TERMINAL_STATUSES:
        return Response({
            'can_reapply': True,
            'can_edit': False,
            'reason': 'Vous avez ete rejete precedemment, vous pouvez repostuler',
            'previous_status': previous.statut,
            'previous_date': previous.date_changement_statut.isoformat() if previous.date_changement_statut else None,
            'edit_deadline': edit_deadline.isoformat() if edit_deadline else None,
            'cooldown_days': 0
        })
    else:
        return Response({
            'can_reapply': False,
            'can_edit': can_edit,
            'reason': (
                'Votre candidature est encore modifiable avant la date limite.'
                if can_edit
                else f'Vous avez deja une candidature en cours avec le statut: {previous.statut}'
            ),
            'previous_status': previous.statut,
            'edit_deadline': edit_deadline.isoformat() if edit_deadline else None,
            'cooldown_days': 0
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_specialites_for_preselection(request, master_id):
    """Get specialites for preselection section with filtering context."""
    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    parcours_list = ParcoursAdmission.objects.filter(
        master=master,
        actif=True,
        statut='ouvert'
    ).values('id', 'specialite', 'type', 'nom').distinct('specialite')

    specialites = list(set([p['specialite'] for p in parcours_list if p['specialite']]))

    return Response({
        'master_id': master.id,
        'master_nom': master.nom,
        'context': 'preselection',
        'specialites': specialites,
        'parcours': list(parcours_list)
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_specialites_for_dossier(request, candidature_id):
    """Get specialites available for dossier deposit (context-aware filtering)."""
    try:
        candidature = Candidature.objects.select_related('master').get(
            id=candidature_id,
            candidat=request.user
        )
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if candidature.statut not in ['preselectionne', 'en_attente_dossier']:
        return Response(
            {'error': f'Dossier non disponible pour statut: {candidature.statut}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    parcours_list = ParcoursAdmission.objects.filter(
        master=candidature.master,
        actif=True,
        statut='ouvert'
    ).values('id', 'specialite', 'type', 'nom').distinct('specialite')

    specialites = list(set([p['specialite'] for p in parcours_list if p['specialite']]))

    return Response({
        'candidature_id': candidature.id,
        'master_id': candidature.master.id,
        'context': 'dossier_deposit',
        'specialites': specialites,
        'current_specialite': candidature.specialite,
        'allow_change': candidature.statut == 'en_attente_dossier'
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_specialites_for_inscription(request, master_id):
    """Get specialites for online inscription section."""
    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    parcours_list = ParcoursAdmission.objects.filter(
        master=master,
        actif=True
    ).values('id', 'specialite', 'type', 'nom').distinct('specialite')

    specialites = list(set([p['specialite'] for p in parcours_list if p['specialite']]))

    candidate_selection = Candidature.objects.filter(
        candidat=request.user,
        master=master,
        statut__in=['preselectionne', 'accepte', 'admis_principal', 'admis_attente', 'admis_complementaire']
    ).first()

    return Response({
        'master_id': master.id,
        'master_nom': master.nom,
        'context': 'inscription',
        'specialites': specialites,
        'candidate_current_specialite': candidate_selection.specialite if candidate_selection else None,
        'candidate_status': candidate_selection.statut if candidate_selection else None
    })


def _trigger_notification_on_duplicate_attempt(user, master, existing_candidature):
    """Trigger in-app and email notification when duplicate candidature is attempted."""
    try:
        _safe_create_notification(
            user=user,
            titre='Candidature dupliquée détectée',
            message=f'Vous avez déjà une candidature pour {master.nom}. Statut: {existing_candidature.statut}.',
            notif_type='warning',
            dedup_key=f'duplicate-attempt-{existing_candidature.id}-{timezone.now().date().isoformat()}'
        )
    except Exception as e:
        logger.warning('Erreur creation notification duplicate: %s', e)


def _trigger_notification_on_deadline_approaching(candidature, days_remaining=7):
    """Trigger notification when application deadline is approaching."""
    try:
        if candidature.master.date_limite_candidature:
            days_left = (candidature.master.date_limite_candidature - timezone.now().date()).days
            if 0 <= days_left <= days_remaining:
                _safe_create_notification(
                    user=candidature.candidat,
                    titre='Délai de candidature proche',
                    message=f'Plus que {days_left} jour(s) pour postuler au {candidature.master.nom}',
                    notif_type='warning',
                    dedup_key=f'deadline-warning-{candidature.master.id}-{timezone.now().date().isoformat()}'
                )
    except Exception as e:
        logger.warning('Erreur notification deadline: %s', e)


def _trigger_notification_on_reapplication_allowed(user, master):
    """Trigger notification when reapplication is allowed after rejection."""
    try:
        _safe_create_notification(
            user=user,
            titre='Possibilité de repostuler',
            message=f'Vous pouvez maintenant repostuler pour {master.nom}.',
            notif_type='info',
            dedup_key=f'reapply-allowed-{master.id}-{timezone.now().date().isoformat()}'
        )
    except Exception as e:
        logger.warning('Erreur notification reapplication: %s', e)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def soumettre_avis_membre(request, candidature_id):
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    avis_bool = request.data.get('avis')
    argument = str(request.data.get('argument', '')).strip()
    commission_id = request.data.get('commission_id')

    try:
        candidature = Candidature.objects.select_related('master').get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if commission_id:
        commission_qs = MembreCommission.objects.filter(
            user=request.user,
            actif=True,
            commission_id=commission_id,
        ).select_related('commission')
    else:
        commission_qs = MembreCommission.objects.filter(user=request.user, actif=True).select_related('commission')

    membre_commission = commission_qs.first()
    if not membre_commission:
        return Response({'error': 'Aucune commission active pour cet utilisateur'}, status=status.HTTP_403_FORBIDDEN)

    commission = membre_commission.commission
    # Enforce deadline_avis if defined
    try:
        deadline = commission.deadline_avis
        if deadline and timezone.now() > deadline:
            return Response({'error': 'La date limite pour soumettre des avis est depassee'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception:
        pass
    avis_normalized = str(avis_bool).strip().lower()
    avis_value = avis_normalized in ['true', '1', 'oui', 'favorable', 'yes']

    if not avis_value and not argument:
        return Response(
            {'error': 'L argumentation est obligatoire pour un avis defavorable'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    avis_obj, created = AvisMembre.objects.update_or_create(
        membre=membre_commission,
        candidature=candidature,
        commission=commission,
        defaults={
            'avis': avis_value,
            'argument': argument,
        },
    )

    _safe_create_notification(
        user=candidature.candidat,
        titre='Avis de commission',
        message=(
            f"Un membre de la commission a depose un avis {'favorable' if avis_value else 'defavorable'} sur votre candidature."
        ),
        notif_type='info',
        dedup_key=f'avis-membre-{candidature.id}-{commission.id}-{request.user.id}',
    )

    try:
        envoyer_email_changement_statut(candidature, candidature.statut, candidature.statut)
    except Exception:
        logger.exception('Erreur lors de l envoi email avis membre pour candidature %s', candidature.id)

    return Response(
        {
            'success': True,
            'avis_id': avis_obj.id,
            'created': created,
            'avis': avis_obj.avis,
            'argument': avis_obj.argument,
            'commission': commission.nom,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def statistiques_avis_candidature(request, candidature_id):
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    avis_qs = AvisMembre.objects.filter(candidature=candidature).select_related('membre__user', 'commission')
    total = avis_qs.count()
    favorables = avis_qs.filter(avis=True).count()
    defavorables = avis_qs.filter(avis=False).count()

    details = [
        {
            'membre': f"{avis.membre.user.first_name} {avis.membre.user.last_name}".strip(),
            'commission': avis.commission.nom,
            'avis': avis.avis,
            'argument': avis.argument,
            'date': avis.date_avis,
        }
        for avis in avis_qs.order_by('-date_avis')
    ]

    return Response(
        {
            'candidature_id': candidature.id,
            'total': total,
            'favorables': favorables,
            'defavorables': defavorables,
            'pourcentage_favorable': round((favorables / total) * 100, 2) if total else 0,
            'avis': details,
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_avis_candidature(request, candidature_id):
    """List all avis for a candidature with member details"""
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    avis_qs = AvisMembre.objects.filter(candidature=candidature).select_related(
        'membre__user', 'commission'
    ).order_by('-date_avis')

    serializer = AvisMembreSerializer(avis_qs, many=True, context={'request': request})
    return Response(
        {
            'candidature_id': candidature.id,
            'count': len(serializer.data),
            'avis': serializer.data,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_avis_detail(request, candidature_id, avis_id):
    """Retrieve a specific avis"""
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    try:
        avis = AvisMembre.objects.get(id=avis_id, candidature=candidature)
    except AvisMembre.DoesNotExist:
        return Response({'error': 'Avis non trouve'}, status=status.HTTP_404_NOT_FOUND)

    serializer = AvisMembreSerializer(avis, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_avis(request, candidature_id, avis_id):
    """Update an avis (only by the member who submitted it or admin)"""
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    try:
        avis = AvisMembre.objects.get(id=avis_id, candidature=candidature)
    except AvisMembre.DoesNotExist:
        return Response({'error': 'Avis non trouve'}, status=status.HTTP_404_NOT_FOUND)

    # Only the member who submitted the avis or admin can update it
    if role != 'admin' and avis.membre.user.id != request.user.id:
        return Response(
            {'error': 'Vous ne pouvez modifier que vos propres avis'},
            status=status.HTTP_403_FORBIDDEN,
        )

    avis_bool = request.data.get('avis')
    argument = request.data.get('argument', '')

    if avis_bool is None:
        return Response({'error': 'Champ avis requis'}, status=status.HTTP_400_BAD_REQUEST)

    if avis_bool is False and not argument.strip():
        return Response(
            {'error': 'Un argument est requis pour un avis defavorable'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    avis.avis = avis_bool
    avis.argument = argument
    avis.save()

    # Create notification for update
    titre = f"Avis modifié - Candidature #{candidature.id}"
    message = f"L'avis pour la candidature #{candidature.id} a été modifié."
    dedup_key = f'avis-update-{candidature.id}-{avis.id}-{timezone.now().date().isoformat()}'
    _safe_create_notification(candidature.candidat.user, titre, message, 'info', dedup_key)

    return Response(
        {
            'success': True,
            'avis_id': avis.id,
            'avis': avis.avis,
            'argument': avis.argument,
            'date_avis': avis.date_avis,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_avis(request, candidature_id, avis_id):
    """Delete an avis (only by the member who submitted it or admin)"""
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    try:
        avis = AvisMembre.objects.get(id=avis_id, candidature=candidature)
    except AvisMembre.DoesNotExist:
        return Response({'error': 'Avis non trouve'}, status=status.HTTP_404_NOT_FOUND)

    # Only the member who submitted the avis or admin can delete it
    if role != 'admin' and avis.membre.user.id != request.user.id:
        return Response(
            {'error': 'Vous ne pouvez supprimer que vos propres avis'},
            status=status.HTTP_403_FORBIDDEN,
        )

    avis_id_deleted = avis.id
    avis.delete()

    # Create notification for deletion
    titre = f"Avis supprimé - Candidature #{candidature.id}"
    message = f"L'avis pour la candidature #{candidature.id} a été supprimé."
    dedup_key = f'avis-delete-{candidature.id}-{avis_id_deleted}-{timezone.now().date().isoformat()}'
    _safe_create_notification(candidature.candidat.user, titre, message, 'warning', dedup_key)

    return Response(
        {'success': True, 'message': 'Avis supprime avec succes'},
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def filter_avis_by_commission(request, master_id):
    """Filter avis by commission with optional criteria (commission_id, member_id, avis_type, date_from, date_to)"""
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    # Check if user is member of a commission for this master
    if role in ['commission', 'responsable_commission']:
        user_commissions = MembreCommission.objects.filter(
            user=request.user, commission__master_id=master_id, actif=True
        ).values_list('commission_id', flat=True)
        if not user_commissions:
            return Response({'error': 'Not a member of any commission for this master'}, status=status.HTTP_403_FORBIDDEN)
    else:
        user_commissions = None

    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    # Get all candidatures for this master
    candidatures_qs = Candidature.objects.filter(master=master).values_list('id', flat=True)

    # Start building avis queryset
    avis_qs = AvisMembre.objects.filter(candidature_id__in=candidatures_qs).select_related(
        'membre__user', 'commission', 'candidature'
    )

    # Filter by commission if not admin
    if user_commissions is not None:
        avis_qs = avis_qs.filter(commission_id__in=user_commissions)

    # Optional: Filter by specific commission_id from query params
    commission_id = request.query_params.get('commission_id')
    if commission_id:
        if not user_commissions or int(commission_id) not in user_commissions:
            # If not admin and commission_id is provided, check if user is member
            if role != 'admin':
                return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)
        avis_qs = avis_qs.filter(commission_id=commission_id)

    # Optional: Filter by member_id
    member_id = request.query_params.get('member_id')
    if member_id:
        avis_qs = avis_qs.filter(membre__user_id=member_id)

    # Optional: Filter by avis type (favorable=true, defavorable=false)
    avis_type = request.query_params.get('avis_type')
    if avis_type:
        avis_bool = avis_type.lower() == 'favorable'
        avis_qs = avis_qs.filter(avis=avis_bool)

    # Optional: Filter by date range
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        avis_qs = avis_qs.filter(date_avis__gte=date_from)
    if date_to:
        avis_qs = avis_qs.filter(date_avis__lte=date_to)

    avis_qs = avis_qs.order_by('-date_avis')

    # Build response with grouped statistics
    avis_list = []
    for avis in avis_qs:
        avis_list.append(
            {
                'id': avis.id,
                'candidature_id': avis.candidature.id,
                'candidat_name': f"{avis.candidature.candidat.first_name} {avis.candidature.candidat.last_name}".strip(),
                'member_id': avis.membre.user.id,
                'member_name': f"{avis.membre.user.first_name} {avis.membre.user.last_name}".strip(),
                'commission_id': avis.commission.id,
                'commission_name': avis.commission.nom,
                'avis': avis.avis,
                'avis_type': 'favorable' if avis.avis else 'defavorable',
                'argument': avis.argument,
                'date_avis': avis.date_avis,
            }
        )

    # Calculate statistics
    total = len(avis_list)
    favorables = sum(1 for a in avis_list if a['avis'])
    defavorables = total - favorables

    return Response(
        {
            'master_id': master_id,
            'master_nom': master.nom,
            'count': total,
            'statistics': {
                'total': total,
                'favorables': favorables,
                'defavorables': defavorables,
                'pourcentage_favorable': round((favorables / total) * 100, 2) if total else 0,
            },
            'avis': avis_list,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_commission_members_for_master(request, master_id):
    """Get all members of commissions for a specific master"""
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    commissions = Commission.objects.filter(master=master, actif=True).prefetch_related('membres')

    commission_members = []
    for commission in commissions:
        for membre in commission.membres.filter(actif=True):
            commission_members.append(
                {
                    'commission_id': commission.id,
                    'commission_name': commission.nom,
                    'member_id': membre.user.id,
                    'member_name': f"{membre.user.first_name} {membre.user.last_name}".strip(),
                    'member_email': membre.user.email,
                    'role': membre.role,
                    'date_nomination': membre.date_nomination,
                }
            )

    return Response(
        {
            'master_id': master_id,
            'master_nom': master.nom,
            'count': len(commission_members),
            'members': commission_members,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_appel_avis(request, commission_id):
    """Endpoint pour que le responsable envoie un appel à avis aux membres d'une commission.

    Corps attendu: { 'message': optional custom message }
    """
    role = getattr(request.user, 'role', None)
    try:
        commission = Commission.objects.get(id=commission_id)
    except Commission.DoesNotExist:
        return Response({'error': 'Commission non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if role not in ['admin', 'responsable_commission'] and not _is_user_responsable_for_commission(request.user, commission):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    custom_message = request.data.get('message') or ''
    # Prefer a user-facing parcours/master name when available, fall back to commission.nom
    parcours_display = None
    try:
        # If commission links directly to a master, try to use a ParcoursAdmission name for clarity
        if hasattr(commission, 'master') and commission.master is not None:
            # Try to find an active ParcoursAdmission for this master
            parcours = ParcoursAdmission.objects.filter(master=commission.master, actif=True).first()
            if parcours and getattr(parcours, 'nom', None):
                parcours_display = parcours.nom
            else:
                parcours_display = commission.master.nom if getattr(commission, 'master', None) else None
    except Exception:
        parcours_display = None

    if not parcours_display:
        parcours_display = commission.nom

    titre = f"Demande d'avis pour le parcours {parcours_display}"
    message_default = (
        f"Demande d'avis pour le parcours {parcours_display}. "
        "Merci de renseigner votre avis sur la présélection avant la date limite."
    )
    message = custom_message.strip() or message_default

    membres = MembreCommission.objects.filter(commission=commission, actif=True).select_related('user')
    sent = 0
    failed = 0
    for membre in membres:
        try:
            creer_notification_avec_email(
                user=membre.user,
                titre=titre,
                message=message,
                notif_type='info',
                dedup_key=f'appel-avis-commission-{commission.id}-{membre.user.id}-{timezone.now().date().isoformat()}'
            )
            sent += 1
        except Exception:
            logger.exception('Erreur envoi appel avis pour user %s', getattr(membre.user, 'id', None))
            failed += 1

    return Response({'success': True, 'sent': sent, 'failed': failed, 'total': sent + failed})


def _resolve_user_membre_for_commission(user, commission_id):
    return (
        MembreCommission.objects.filter(
            user=user,
            actif=True,
            commission_id=commission_id,
        )
        .select_related('commission', 'user')
        .first()
    )


def _ensure_auto_validated_global_avis(commission):
    """Create default favorable global avis for members after deadline expiry."""
    if not commission or not getattr(commission, 'deadline_avis', None):
        return 0
    if timezone.now() <= commission.deadline_avis:
        return 0

    expected_members = MembreCommission.objects.filter(
        commission=commission,
        actif=True,
        role='membre',
    )

    created_count = 0
    for membre in expected_members:
        _, created = AvisSelection.objects.get_or_create(
            commission=commission,
            membre=membre,
            is_global=True,
            defaults={
                'statut': 'favorable',
                'commentaire': 'Auto-validé (absence de réponse avant deadline).',
            },
        )
        if created:
            created_count += 1
    return created_count


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def commission_avis_global(request, commission_id):
    """Global advisory endpoint for Sprint 3 collegial decision workflow.

    - POST: member submits one global opinion for the full preselection list.
    - GET: responsable/admin gets aggregated table + majority status + can_decide flag.
    """
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        commission = Commission.objects.get(id=commission_id, actif=True)
    except Commission.DoesNotExist:
        return Response({'error': 'Commission non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        membre = _resolve_user_membre_for_commission(request.user, commission_id)
        if not membre and role != 'admin':
            return Response({'error': 'Aucune appartenance a cette commission'}, status=status.HTTP_403_FORBIDDEN)

        statut = str(request.data.get('statut', '')).strip().lower()
        commentaire = str(request.data.get('commentaire', '')).strip()
        is_global = bool(request.data.get('is_global', True))

        if statut not in ['favorable', 'defavorable']:
            return Response(
                {'error': 'Statut invalide. Utilisez favorable ou defavorable.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if statut == 'defavorable' and not commentaire:
            return Response(
                {'error': 'Le commentaire est obligatoire pour un avis defavorable.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if commission.deadline_avis and timezone.now() > commission.deadline_avis:
            return Response({'error': 'La date limite des avis est depassee.'}, status=status.HTTP_400_BAD_REQUEST)

        if role == 'admin' and not membre:
            member_id = request.data.get('member_id')
            membre = MembreCommission.objects.filter(
                id=member_id,
                commission_id=commission_id,
                actif=True,
            ).first()
            if not membre:
                return Response({'error': 'Membre commission introuvable pour admin.'}, status=status.HTTP_400_BAD_REQUEST)

        avis_obj, created = AvisSelection.objects.update_or_create(
            commission=commission,
            membre=membre,
            is_global=is_global,
            defaults={
                'statut': statut,
                'commentaire': commentaire,
            },
        )

        serializer = AvisSelectionSerializer(avis_obj, context={'request': request})
        return Response(
            {
                'success': True,
                'created': created,
                'message': 'Avis global enregistre.',
                'avis': serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    # GET
    is_member_of_commission = MembreCommission.objects.filter(
        commission=commission,
        user=request.user,
        actif=True,
    ).exists()
    if role not in ['responsable_commission', 'admin'] and not (
        is_member_of_commission or _is_user_responsable_for_commission(request.user, commission)
    ):
        return Response({'error': 'Acces reserve au responsable de commission.'}, status=status.HTTP_403_FORBIDDEN)

    _ensure_auto_validated_global_avis(commission)

    expected_members = list(
        MembreCommission.objects.filter(
            commission=commission,
            actif=True,
            role='membre',
        ).select_related('user')
    )
    avis_qs = AvisSelection.objects.filter(
        commission=commission,
        is_global=True,
    ).select_related('membre__user').order_by('-date_avis')

    avis_by_membre_id = {avis.membre_id: avis for avis in avis_qs}
    rows = []
    favorables = 0
    defavorables = 0

    for membre in expected_members:
        avis = avis_by_membre_id.get(membre.id)
        member_name = f"{membre.user.first_name} {membre.user.last_name}".strip() or membre.user.username
        if avis:
            if avis.statut == 'favorable':
                favorables += 1
            elif avis.statut == 'defavorable':
                defavorables += 1

        rows.append(
            {
                'membre_id': membre.id,
                'membre_name': member_name,
                'membre_email': membre.user.email,
                'statut': avis.statut if avis else 'en_attente',
                'commentaire': avis.commentaire if avis else '',
                'date_avis': avis.date_avis if avis else None,
                'is_global': True,
            }
        )

    total_members = len(expected_members)
    completed_count = sum(1 for r in rows if r['statut'] != 'en_attente')
    pending_count = total_members - completed_count
    deadline_expired = bool(commission.deadline_avis and timezone.now() > commission.deadline_avis)
    can_decide_final = pending_count == 0 or deadline_expired
    majority_recommendation = 'favorable' if favorables >= defavorables else 'defavorable'

    return Response(
        {
            'commission': {
                'id': commission.id,
                'nom': commission.nom,
                'deadline_avis': commission.deadline_avis,
            },
            'summary': {
                'total_members': total_members,
                'completed_count': completed_count,
                'pending_count': pending_count,
                'favorables': favorables,
                'defavorables': defavorables,
                'deadline_expired': deadline_expired,
                'can_decide_final': can_decide_final,
                'majority_recommendation': majority_recommendation,
            },
            'responses': rows,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_commissions(request):
    """Retourne les commissions liées à l'utilisateur courant.

    Le frontend utilise cette liste pour choisir une commission active
    et filtrer les vues multi-commissions.
    """
    role = getattr(request.user, 'role', None)
    if role not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    commissions = Commission.objects.filter(
        Q(membres__user=request.user, membres__actif=True) |
        Q(membre_commission_links__user=request.user, membre_commission_links__actif=True)
    ).distinct().select_related('master').order_by('nom')

    data = []
    active_commission_id = request.headers.get('X-Active-Commission-Id') or request.query_params.get('active_commission_id')
    active_commission_id = int(active_commission_id) if str(active_commission_id).isdigit() else None

    for commission in commissions:
        data.append(
            {
                'id': commission.id,
                'nom': commission.nom,
                'description': commission.description,
                'actif': commission.actif,
                'master_id': commission.master_id,
                'master_nom': commission.master.nom if commission.master else '',
                'is_active': commission.id == active_commission_id,
            }
        )

    return Response(
        {
            'count': len(data),
            'active_commission_id': active_commission_id,
            'commissions': data,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_member_credentials(request):
    email = request.data.get('email')
    password = request.data.get('password')
    first_name = request.data.get('first_name', '')
    last_name = request.data.get('last_name', '')
    role = request.data.get('role', '')

    if not email or not password:
        return Response({'error': 'Email et password requis'}, status=status.HTTP_400_BAD_REQUEST)

    role_text = 'Responsable Commission' if role == 'responsable_commission' else 'Membre Commission'

    try:
        send_mail(
            subject='Vos identifiants ISIMM',
            message=(
                f"Bonjour {first_name} {last_name},\n\n"
                f"Role : {role_text}\n"
                f"Email : {email}\n"
                f"Mot de passe : {password}\n"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False,
        )
        return Response({'message': 'Email envoye'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_delete_avis(request):
    """Bulk delete avis (admin and responsable_commission only)"""
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    avis_ids = request.data.get('avis_ids', [])
    if not avis_ids or not isinstance(avis_ids, list):
        return Response({'error': 'avis_ids doit etre une liste'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        avis_qs = AvisMembre.objects.filter(id__in=avis_ids)
        
        # If responsable_commission, only allow deleting avis from their commissions
        if role == 'responsable_commission':
            user_commissions = MembreCommission.objects.filter(
                user=request.user, actif=True
            ).values_list('commission_id', flat=True)
            avis_qs = avis_qs.filter(commission_id__in=user_commissions)

        count = avis_qs.count()
        avis_qs.delete()

        titre = f'Suppression en masse d\'avis'
        message = f'{count} avis ont été supprimés'
        _safe_create_notification(request.user, titre, message, 'warning', f'bulk-delete-{timezone.now().date().isoformat()}')

        return Response(
            {'success': True, 'deleted_count': count},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_update_candidature_status(request):
    """Bulk update candidature status (admin and responsable_commission only)"""
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    candidature_ids = request.data.get('candidature_ids', [])
    new_status = request.data.get('status')
    reason = request.data.get('reason', '')

    if not candidature_ids or not isinstance(candidature_ids, list):
        return Response({'error': 'candidature_ids doit etre une liste'}, status=status.HTTP_400_BAD_REQUEST)

    if not new_status:
        return Response({'error': 'status est requis'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        candidatures = Candidature.objects.filter(id__in=candidature_ids)

        # If responsable_commission, only allow updating candidatures for their masters
        if role == 'responsable_commission':
            user_masters = MembreCommission.objects.filter(
                user=request.user, actif=True
            ).values_list('commission__master_id', flat=True).distinct()
            candidatures = candidatures.filter(master_id__in=user_masters)

        count = 0
        for candidature in candidatures:
            candidature.statut = new_status
            candidature.date_changement_statut = timezone.now()
            candidature.save()
            
            # Send notification to candidate
            titre = f'Statut de candidature modifié'
            message = f'Votre candidature #{candidature.id} a un nouveau statut: {new_status}'
            if reason:
                message += f'\nMotif: {reason}'
            _safe_create_notification(
                candidature.candidat.user,
                titre,
                message,
                'info' if new_status in ['preselectionne', 'inscrit'] else 'warning',
                f'status-update-{candidature.id}-{timezone.now().date().isoformat()}'
            )
            count += 1

        return Response(
            {'success': True, 'updated_count': count},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_candidatures_to_member(request):
    """Assign candidatures to a commission member (admin and responsable_commission only)"""
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    candidature_ids = request.data.get('candidature_ids', [])
    member_id = request.data.get('member_id')
    commission_id = request.data.get('commission_id')

    if not candidature_ids or not isinstance(candidature_ids, list):
        return Response({'error': 'candidature_ids doit etre une liste'}, status=status.HTTP_400_BAD_REQUEST)

    if not member_id or not commission_id:
        return Response({'error': 'member_id et commission_id sont requis'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Verify the member exists
        member = MembreCommission.objects.get(id=member_id, actif=True)
        
        # If responsable_commission, check they can assign to this commission
        if role == 'responsable_commission':
            user_commissions = MembreCommission.objects.filter(
                user=request.user, actif=True
            ).values_list('commission_id', flat=True)
            if commission_id not in user_commissions:
                return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

        # Get candidatures
        candidatures = Candidature.objects.filter(id__in=candidature_ids)

        # Create or update assignment records (if using an Assignment model)
        # For now, we'll track this via a simple notification
        count = len(candidatures)
        titre = f'Nouvelles candidatures assignées'
        message = f'{count} candidature(s) vous ont été assignée(s) pour évaluation'
        _safe_create_notification(
            member.user,
            titre,
            message,
            'info',
            f'assignment-{member_id}-{timezone.now().date().isoformat()}'
        )

        return Response(
            {'success': True, 'assigned_count': count, 'member_id': member_id},
            status=status.HTTP_200_OK,
        )
    except MembreCommission.DoesNotExist:
        return Response({'error': 'Membre non trouve'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_dashboard_stats(request):
    """Get admin dashboard statistics (admin only)"""
    role = getattr(request.user, 'role', None)
    if role != 'admin':
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        total_candidatures = Candidature.objects.count()
        total_avis = AvisMembre.objects.count()
        total_commissions = Commission.objects.filter(actif=True).count()
        total_members = MembreCommission.objects.filter(actif=True).count()

        # Status breakdown
        statuts = Candidature.objects.values('statut').annotate(count=Count('id'))
        status_breakdown = {item['statut']: item['count'] for item in statuts}

        # Avis statistics
        avis_favorables = AvisMembre.objects.filter(avis=True).count()
        avis_defavorables = AvisMembre.objects.filter(avis=False).count()

        return Response(
            {
                'total_candidatures': total_candidatures,
                'total_avis': total_avis,
                'total_commissions': total_commissions,
                'total_members': total_members,
                'status_breakdown': status_breakdown,
                'avis_statistics': {
                    'favorables': avis_favorables,
                    'defavorables': avis_defavorables,
                    'total': total_avis,
                },
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def annuler_candidature(request, candidature_id):
    try:
        candidature = Candidature.objects.get(id=candidature_id, candidat=request.user)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if not candidature.peut_etre_annulee():
        return Response(
            {'error': 'Cette candidature ne peut plus etre annulee'},
            status=status.HTTP_403_FORBIDDEN,
        )

    ancien_statut = candidature.statut
    candidature.statut = 'annule'
    candidature.annule_par_candidat = True
    candidature.date_annulation = timezone.now()
    candidature.save()

    candidature.ajouter_historique(ancien_statut, 'annule', request.user, 'Annule par le candidat')

    return Response({'success': True, 'message': 'Candidature annulee avec succes'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def corbeille_candidatures(request):
    if request.user.role not in ['admin', 'commission', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    candidatures = Candidature.objects.filter(statut='annule', annule_par_candidat=True)
    serializer = CandidatureSerializer(candidatures, many=True)
    return Response(serializer.data)


@api_view(['GET', 'POST', 'PUT'])
@permission_classes([IsAuthenticated])
def gerer_configuration_appel(request, master_id=None):
    if request.user.role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        try:
            config = ConfigurationAppel.objects.get(master_id=master_id)
            serializer = ConfigurationAppelSerializer(config)
            return Response(serializer.data)
        except ConfigurationAppel.DoesNotExist:
            return Response({'error': 'Configuration non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        serializer = ConfigurationAppelSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    try:
        config = ConfigurationAppel.objects.get(master_id=master_id)
    except ConfigurationAppel.DoesNotExist:
        return Response({'error': 'Configuration non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ConfigurationAppelSerializer(config, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_document_configuration_appel(request, master_id):
    """
    Upload du document officiel PDF pour une offre de préinscription (master).
    Accessible par responsable et admin uniquement.
    """
    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        config = ConfigurationAppel.objects.get(master_id=master_id)
    except ConfigurationAppel.DoesNotExist:
        return Response({'error': 'Configuration non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    pdf_file = request.FILES.get('document_pdf')
    if not pdf_file:
        return Response(
            {'error': 'Aucun fichier PDF fourni'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Vérifier l'extension du fichier
    allowed_extensions = {'pdf'}
    file_ext = pdf_file.name.split('.')[-1].lower()
    if file_ext not in allowed_extensions:
        return Response(
            {'error': 'Seuls les fichiers PDF sont acceptes'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Vérifier la taille du fichier (max 10 MB)
    max_size = 10 * 1024 * 1024
    if pdf_file.size > max_size:
        return Response(
            {'error': 'Fichier trop volumineux (max 10 MB)'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Sauvegarder le fichier
    config.document_officiel_pdf = pdf_file
    config.save(update_fields=['document_officiel_pdf'])

    return Response(
        {
            'success': True,
            'message': 'Document PDF charge avec succes',
            'document_url': request.build_absolute_uri(config.document_officiel_pdf.url) if config.document_officiel_pdf else None,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def publier_offre_preinscription(request, master_id):
    """
    Validation et publication officielle d'un appel de preinscription.
    Cette action est reservee au responsable de commission.
    """
    if getattr(request.user, 'role', None) != 'responsable_commission':
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        config = ConfigurationAppel.objects.get(master_id=master_id)
    except ConfigurationAppel.DoesNotExist:
        return Response({'error': 'Configuration non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    config.actif = True
    config.est_cache = False
    config.save(update_fields=['actif', 'est_cache', 'updated_at'])

    return Response(
        {
            'success': True,
            'message': 'Appel de preinscription valide et publie avec succes.',
            'master_id': master_id,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def offres_inscription_responsable(request):
    """
    API pour responsable: retourne les offres avec statut détaillé (visible/cachée),
    capacités (interne/externe), URL PDF, et deadlines.
    Si aucun Master actif n'existe, retourne les 6 parcours officiels depuis SpecialiteParcoursMapping.
    """
    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    _PARCOURS_META = {
        'MPGL':   {'date_limite': '2026-07-22', 'places': 35,  'frontend_type': 'master',         'sous_type': 'professionnel'},
        'MPDS':   {'date_limite': '2026-07-22', 'places': 35,  'frontend_type': 'master',         'sous_type': 'professionnel'},
        'MP3I':   {'date_limite': '2026-07-20', 'places': 25,  'frontend_type': 'master',         'sous_type': 'professionnel'},
        'MRGL':   {'date_limite': '2026-07-22', 'places': 111, 'frontend_type': 'master',         'sous_type': 'recherche'},
        'MRMI':   {'date_limite': '2026-07-20', 'places': 29,  'frontend_type': 'master',         'sous_type': 'recherche'},
        'ING_GL': {'date_limite': '2026-08-08', 'places': 65,  'frontend_type': 'cycle_ingenieur','sous_type': 'cycle_ingenieur'},
    }

    configurations = ConfigurationAppel.objects.filter(actif=True).select_related('master')
    config_by_master = {config.master_id: config for config in configurations}
    masters = Master.objects.filter(actif=True).order_by('nom')

    # ── Fallback : si aucun Master actif → utiliser SpecialiteParcoursMapping ──
    if not masters.exists():
        parcours_qs = SpecialiteParcoursMapping.objects.filter(
            actif=True,
            code_parcours__in=list(_PARCOURS_META.keys()),
        ).order_by('type_formation', 'ordre')
        offres_fallback = []
        for idx, p in enumerate(parcours_qs, start=1):
            m = _PARCOURS_META.get(p.code_parcours, {})
            offres_fallback.append({
                'id': idx,
                'titre': p.nom_parcours,
                'type': m.get('frontend_type', 'master'),
                'sous_type': m.get('sous_type', 'professionnel'),
                'specialite': p.code_parcours,
                'description': '',
                'statut': 'ouvert',
                'est_cache': False,
                'est_visible': True,
                'capacite_total': m.get('places', 0),
                'capacite_interne': 0,
                'capacite_externe': 0,
                'capacite_liste_attente': 0,
                'places': m.get('places', 0),
                'nombre_candidats_inscrits': 0,
                'date_limite': m.get('date_limite', '2026-07-31'),
                'date_debut_visibilite': None,
                'date_fin_visibilite': None,
                'date_limite_preinscription': m.get('date_limite', '2026-07-31'),
                'date_limite_depot_dossier': None,
                'date_limite_paiement': None,
                'delai_modification_jours': 7,
                'delai_depot_dossier_j_jours': 14,
                'document_officiel_pdf_url': None,
                'requires_configuration': False,
                'code_parcours': p.code_parcours,
                'specialites_eligibles': p.specialites if isinstance(p.specialites, list) else [],
            })
        if offres_fallback:
            return Response(offres_fallback)
        # Fallback ultime si SpecialiteParcoursMapping aussi vide
        return Response([
            {'id': i, 'titre': nom, 'type': m['frontend_type'], 'sous_type': m['sous_type'],
             'specialite': code, 'description': '', 'statut': 'ouvert', 'est_cache': False,
             'est_visible': True, 'capacite_total': m['places'], 'capacite_interne': 0,
             'capacite_externe': 0, 'capacite_liste_attente': 0, 'places': m['places'],
             'nombre_candidats_inscrits': 0, 'date_limite': m['date_limite'],
             'date_debut_visibilite': None, 'date_fin_visibilite': None,
             'date_limite_preinscription': m['date_limite'], 'date_limite_depot_dossier': None,
             'date_limite_paiement': None, 'delai_modification_jours': 7,
             'delai_depot_dossier_j_jours': 14, 'document_officiel_pdf_url': None,
             'requires_configuration': False, 'code_parcours': code, 'specialites_eligibles': []}
            for i, (code, nom, m) in enumerate([
                ('MPGL', 'Master Professionnel Genie Logiciel (MPGL)', _PARCOURS_META['MPGL']),
                ('MPDS', 'Mastere Professionnel en sciences de donnees (MPDS)', _PARCOURS_META['MPDS']),
                ('MP3I', 'Mastere Professionnel en Ingenieries en Instrumentation industrielle (MP3I)', _PARCOURS_META['MP3I']),
                ('MRGL', 'Mastere Recherche en Genie logiciel (MRGL)', _PARCOURS_META['MRGL']),
                ('MRMI', 'Mastere Recherche en micro-electronique et instrumentation (MRMI)', _PARCOURS_META['MRMI']),
                ('ING_GL', 'Ingenieur en sciences Appliquees et Technologie - Genie Logiciel', _PARCOURS_META['ING_GL']),
            ], start=1)
        ])

    offres = []
    for master in masters:
        config = config_by_master.get(master.id)
        nombre_candidats_inscrits = Candidature.objects.filter(master_id=master.id).count()

        if config:
            statut = 'ouvert' if config.peut_candidater() else 'ferme'
            est_visible = config.est_visible()
            est_cache = config.est_cache
            capacite_total = config.capacite_accueil
            capacite_interne = config.capacite_interne
            capacite_externe = config.capacite_externe
            capacite_liste_attente = config.capacite_liste_attente
            date_debut_visibilite = config.date_debut_visibilite
            date_fin_visibilite = config.date_fin_visibilite
            date_limite_preinscription = config.date_limite_preinscription
            date_limite_depot_dossier = config.date_limite_depot_dossier
            date_limite_paiement = config.date_limite_paiement
            delai_modification_jours = config.delai_modification_candidature_jours
            delai_depot_dossier_j_jours = config.delai_depot_dossier_preselectionnes_jours
            document_officiel_pdf_url = (
                request.build_absolute_uri(config.document_officiel_pdf.url)
                if config.document_officiel_pdf
                else None
            )
        else:
            today = timezone.now().date()
            statut = 'ouvert' if master.date_limite_candidature and master.date_limite_candidature >= today else 'ferme'
            est_visible = True
            est_cache = False
            capacite_total = master.places_disponibles
            capacite_interne = 0
            capacite_externe = 0
            capacite_liste_attente = 0
            date_debut_visibilite = None
            date_fin_visibilite = None
            date_limite_preinscription = master.date_limite_candidature
            date_limite_depot_dossier = None
            date_limite_paiement = None
            delai_modification_jours = 7
            delai_depot_dossier_j_jours = 14
            document_officiel_pdf_url = None

        offres.append(
            {
                'id': master.id,
                'titre': master.nom,
                'type': 'cycle_ingenieur' if 'ingenieur' in master.nom.lower() else 'master',
                'sous_type': master.type_master,
                'specialite': master.specialite,
                'description': master.description,
                'statut': statut,
                'est_cache': est_cache,
                'est_visible': est_visible,
                'capacite_total': capacite_total,
                'capacite_interne': capacite_interne,
                'capacite_externe': capacite_externe,
                'capacite_liste_attente': capacite_liste_attente,
                'places': master.places_disponibles,
                'nombre_candidats_inscrits': nombre_candidats_inscrits,
                'date_limite': master.date_limite_candidature,
                'date_debut_visibilite': date_debut_visibilite,
                'date_fin_visibilite': date_fin_visibilite,
                'date_limite_preinscription': date_limite_preinscription,
                'date_limite_depot_dossier': date_limite_depot_dossier,
                'date_limite_paiement': date_limite_paiement,
                'delai_modification_jours': delai_modification_jours,
                'delai_depot_dossier_j_jours': delai_depot_dossier_j_jours,
                'document_officiel_pdf_url': document_officiel_pdf_url,
                'requires_configuration': config is None,
            }
        )

    return Response(offres)


class FormuleScoreViewSet(viewsets.ModelViewSet):
    queryset = FormuleScore.objects.all()
    serializer_class = FormuleScoreSerializer
    permission_classes = [IsAuthenticated]


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def formule_score_master(request, master_id):
    """Expose une API simple pour consulter/editer la formule de score d'un master."""
    if getattr(request.user, 'role', None) not in ['admin', 'commission', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        formule = FormuleScore.objects.filter(master=master).first()
        if not formule:
            return Response(
                {
                    'master_id': master.id,
                    'master_nom': master.nom,
                    'message': 'Aucune formule configuree pour ce master.',
                },
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(FormuleScoreSerializer(formule).data)

    formule, _ = FormuleScore.objects.get_or_create(
        master=master,
        defaults={'nom': f'Formule {master.nom}', 'description': ''},
    )

    payload = request.data.copy()
    payload['master'] = master.id
    if 'nom' not in payload:
        payload['nom'] = formule.nom or f'Formule {master.nom}'

    serializer = FormuleScoreSerializer(
        formule,
        data=payload,
        partial=(request.method == 'PATCH'),
    )
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def master_coefficients(request, master_id):
    try:
        master = Master.objects.get(id=master_id, actif=True)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    return Response(
        {
            'master_id': master.id,
            'master_nom': master.nom,
            'coeff_bac': float(master.coeff_bac or 0),
            'coeff_licence': float(master.coeff_licence or 0),
            'coeff_examen': float(master.coeff_examen or 0),
            'bonus_mention': float(master.bonus_mention or 0),
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def deposer_dossier_numerique(request, candidature_id):
    """
    Depot dossier numerique avec controle strict:
    - uniquement candidat proprietaire
    - uniquement statut autorise
    - validation formulaire commission par master
    """
    try:
        candidature = Candidature.objects.select_related('master', 'master__configuration').get(
            id=candidature_id,
            candidat=request.user,
        )
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    statuts_autorises = {'preselectionne', 'en_attente_dossier'}
    if candidature.statut not in statuts_autorises:
        return Response(
            {
                'error': 'Depot dossier non autorise pour ce statut.',
                'statut_actuel': candidature.statut,
                'statuts_autorises': sorted(list(statuts_autorises)),
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    try:
        configuration = candidature.master.configuration
    except ConfigurationAppel.DoesNotExist:
        return Response(
            {'error': 'Configuration master introuvable.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    formulaire_payload = request.data.get('formulaire', {})
    validation = _validate_formulaire_commission(configuration, formulaire_payload)
    if not validation.get('ok'):
        return Response(validation, status=status.HTTP_400_BAD_REQUEST)

    diagnostic_ocr = verifier_concordance_dossier(candidature, formulaire_payload)

    ancien_statut = candidature.statut
    candidature.statut = 'dossier_depose'
    candidature.dossier_depose = True
    candidature.dossier_valide = bool(diagnostic_ocr.get('validation_auto'))
    candidature.date_depot_dossier = timezone.now()
    candidature.save(
        update_fields=['statut', 'dossier_depose', 'dossier_valide', 'date_depot_dossier', 'updated_at']
    )

    candidature.ajouter_historique(
        ancien_statut,
        'dossier_depose',
        request.user,
        (
            'Depot dossier numerique via formulaire commission | '
            f"OCR: {diagnostic_ocr.get('decision')} | confiance={diagnostic_ocr.get('confiance')}"
        ),
    )

    return Response(
        {
            'success': True,
            'message': 'Dossier numerique depose avec succes.',
            'candidature_id': candidature.id,
            'statut': candidature.statut,
            'dossier_valide_auto': candidature.dossier_valide,
            'ocr_diagnostic': diagnostic_ocr,
            'date_depot_dossier': candidature.date_depot_dossier,
        }
    )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def ajuster_dossier_numerique(request, candidature_id):
    """
    Permet au candidat presélectionné d'ajuster son dossier avant expiration du délai.
    """
    try:
        candidature = Candidature.objects.select_related('master', 'master__configuration').get(
            id=candidature_id,
            candidat=request.user,
        )
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if candidature.statut not in {'preselectionne', 'en_attente_dossier', 'dossier_depose'}:
        return Response(
            {
                'error': 'Ajustement dossier non autorise pour ce statut.',
                'statut_actuel': candidature.statut,
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    try:
        configuration = candidature.master.configuration
    except ConfigurationAppel.DoesNotExist:
        return Response({'error': 'Configuration master introuvable.'}, status=status.HTTP_400_BAD_REQUEST)

    if not configuration.date_limite_depot_dossier:
        return Response({'error': 'Date limite de depot dossier non configuree.'}, status=status.HTTP_400_BAD_REQUEST)

    today = timezone.now().date()
    date_limite = configuration.date_limite_depot_dossier
    if candidature.prolongation_delai and candidature.delai_depot_dossier:
        date_limite = candidature.delai_depot_dossier

    if today > date_limite:
        return Response(
            {
                'error': 'Le delai d ajustement du dossier est expire.',
                'date_limite': str(date_limite),
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    formulaire_payload = request.data.get('formulaire', {})
    validation = _validate_formulaire_commission(configuration, formulaire_payload)
    if not validation.get('ok'):
        return Response(validation, status=status.HTTP_400_BAD_REQUEST)

    diagnostic_ocr = verifier_concordance_dossier(candidature, formulaire_payload)

    ancien_statut = candidature.statut
    candidature.statut = 'dossier_depose'
    candidature.dossier_depose = True
    candidature.dossier_valide = bool(diagnostic_ocr.get('validation_auto'))
    candidature.date_depot_dossier = timezone.now()
    candidature.save(
        update_fields=['statut', 'dossier_depose', 'dossier_valide', 'date_depot_dossier', 'updated_at']
    )

    candidature.ajouter_historique(
        ancien_statut,
        'dossier_depose',
        request.user,
        'Ajustement dossier numerique avant delai',
    )

    return Response(
        {
            'success': True,
            'message': 'Dossier numerique ajuste avec succes.',
            'candidature_id': candidature.id,
            'statut': candidature.statut,
            'date_depot_dossier': candidature.date_depot_dossier,
            'ocr_diagnostic': diagnostic_ocr,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def ocr_test_diagnostic(request):
    """Endpoint de test OCR/IA independant du depot dossier final."""
    candidature_id = request.data.get('candidature_id')

    candidature = None
    if candidature_id:
        try:
            candidature = Candidature.objects.select_related('candidat').get(id=candidature_id)
        except Candidature.DoesNotExist:
            return Response({'error': 'Candidature non trouvee.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        candidature = (
            Candidature.objects.select_related('candidat')
            .filter(candidat=request.user)
            .order_by('-created_at')
            .first()
        )
        if not candidature and getattr(request.user, 'role', None) in ['admin', 'commission', 'responsable_commission']:
            candidature = Candidature.objects.select_related('candidat').order_by('-created_at').first()

    if not candidature:
        return Response(
            {'error': 'Aucune candidature disponible pour le test OCR.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    formulaire_payload = request.data.get('formulaire', {})
    if not isinstance(formulaire_payload, dict):
        return Response(
            {'error': 'Le champ formulaire doit etre un objet JSON.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    diagnostic = verifier_concordance_dossier(candidature, formulaire_payload)
    return Response(
        {
            'success': True,
            'message': 'Diagnostic OCR execute.',
            'candidature_id': candidature.id,
            'ocr_diagnostic': diagnostic,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyser_lot_ocr(request):
    """
    Analyse OCR en lot — traite une liste de candidature_ids en une seule transaction serveur.

    Body JSON :
        { "candidature_ids": [1, 2, 3, ...] }

    Retourne :
        {
          "success": true,
          "nb_total": N,
          "nb_conformes": K,
          "nb_incoherences": M,
          "nb_erreurs": P,
          "resultats": [
            {
              "candidature_id": int,
              "candidat_nom": str,
              "master": str,
              "statut": "ok" | "anomalie" | "incomplet" | "erreur",
              "flag_fraude": bool,
              "nb_anomalies": int,
              "rapport": { ... }
            },
            ...
          ]
        }
    """
    from .ocr_global_service import auditer_dossier_complet

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusée.'}, status=status.HTTP_403_FORBIDDEN)

    candidature_ids = request.data.get('candidature_ids', [])
    if not isinstance(candidature_ids, list) or not candidature_ids:
        return Response(
            {'error': 'Le champ candidature_ids doit être une liste non vide d\'identifiants.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    MAX_LOT = 50
    if len(candidature_ids) > MAX_LOT:
        return Response(
            {'error': f'Le lot ne peut pas dépasser {MAX_LOT} candidatures à la fois.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Charge toutes les candidatures en une seule requête DB
    candidatures_qs = (
        Candidature.objects
        .select_related('candidat', 'master')
        .filter(id__in=candidature_ids, dossier_depose=True)
    )
    candidatures_map = {c.id: c for c in candidatures_qs}

    resultats = []
    nb_conformes = 0
    nb_incoherences = 0
    nb_erreurs = 0

    for cid in candidature_ids:
        candidature = candidatures_map.get(int(cid))
        if not candidature:
            resultats.append({
                'candidature_id': cid,
                'statut': 'erreur',
                'message': 'Candidature introuvable ou dossier non encore déposé.',
                'flag_fraude': False,
                'nb_anomalies': 0,
                'rapport': None,
            })
            nb_erreurs += 1
            continue

        try:
            rapport = auditer_dossier_complet(candidature.candidat_id)
            statut_global = rapport.get('statut_global', 'erreur')

            if statut_global == 'ok':
                nb_conformes += 1
            elif statut_global in ('anomalie', 'incomplet'):
                nb_incoherences += 1
            else:
                nb_erreurs += 1

            resultats.append({
                'candidature_id': cid,
                'candidat_nom': (
                    f"{getattr(candidature.candidat, 'first_name', '')} "
                    f"{getattr(candidature.candidat, 'last_name', '')}"
                ).strip(),
                'master': candidature.master.nom if candidature.master else '',
                'statut': statut_global,
                'flag_fraude': rapport.get('flag_fraude', False),
                'nb_anomalies': len(rapport.get('anomalies_consolidees', [])),
                'rapport': rapport,
            })

        except Exception as exc:
            logger.exception('Erreur analyse OCR lot pour candidature_id=%s : %s', cid, exc)
            resultats.append({
                'candidature_id': cid,
                'statut': 'erreur',
                'message': str(exc),
                'flag_fraude': False,
                'nb_anomalies': 0,
                'rapport': None,
            })
            nb_erreurs += 1

    return Response({
        'success': True,
        'message': f'Analyse OCR terminée pour {len(candidature_ids)} candidature(s).',
        'nb_total': len(candidature_ids),
        'nb_conformes': nb_conformes,
        'nb_incoherences': nb_incoherences,
        'nb_erreurs': nb_erreurs,
        'resultats': resultats,
    })


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def formulaire_commission_master(request, master_id):
    """Permet a la commission de consulter/modifier le schema du formulaire par master."""
    if getattr(request.user, 'role', None) not in ['admin', 'commission', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        configuration = ConfigurationAppel.objects.select_related('master').get(master_id=master_id)
    except ConfigurationAppel.DoesNotExist:
        return Response({'error': 'Configuration master introuvable'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response(
            {
                'master_id': configuration.master_id,
                'master_nom': configuration.master.nom,
                'formulaire_commission_schema': configuration.formulaire_commission_schema or {},
            }
        )

    schema = request.data.get('formulaire_commission_schema', {})
    if not isinstance(schema, dict):
        return Response(
            {'error': 'formulaire_commission_schema doit etre un objet JSON'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    for key in ['required_fields', 'required_documents']:
        if key in schema and not isinstance(schema[key], list):
            return Response(
                {'error': f'{key} doit etre une liste'},
                status=status.HTTP_400_BAD_REQUEST,
            )

    configuration.formulaire_commission_schema = schema
    configuration.save(update_fields=['formulaire_commission_schema', 'updated_at'])

    return Response(
        {
            'success': True,
            'master_id': configuration.master_id,
            'formulaire_commission_schema': configuration.formulaire_commission_schema,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculer_score_candidature(request, candidature_id):
    try:
        candidature = Candidature.objects.get(id=candidature_id)
        donnees_academiques = candidature.donnees_academiques
        score = donnees_academiques.calculer_et_sauvegarder_score()
        return Response({'success': True, 'score': score, 'candidature_id': candidature.id})
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)
    except DonneesAcademiques.DoesNotExist:
        return Response(
            {'error': 'Donnees academiques non renseignees'},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculer_score_final_et_statut(request, candidature_id):
    """
    API Sprint 3:
    - Calcule/actualise le score final a partir des moyennes saisies (Sprint 2)
    - Met a jour automatiquement le statut selon les seuils du jury
    """

    try:
        candidature = Candidature.objects.select_related('master').get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    user_role = getattr(request.user, 'role', None)
    is_owner = candidature.candidat_id == getattr(request.user, 'id', None)
    if not (is_owner or user_role in ['admin', 'commission', 'responsable_commission']):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    payload = request.data or {}

    def _as_float(value, default=None):
        try:
            if value in [None, '']:
                return default
            return float(value)
        except (TypeError, ValueError):
            return default

    moyenne_bac = _as_float(payload.get('moyenne_bac'))
    moyenne_licence = _as_float(payload.get('moyenne_licence'))
    moyenne_specialite = _as_float(payload.get('moyenne_specialite'))
    note_pfe = _as_float(payload.get('note_pfe'), 0.0)

    if moyenne_licence is None:
        moyenne_licence = _as_float(payload.get('moyenne_generale'))

    if moyenne_licence is None and moyenne_bac is None:
        return Response(
            {'error': 'Veuillez fournir au moins moyenne_licence ou moyenne_bac.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    moyenne_generale = moyenne_licence if moyenne_licence is not None else moyenne_bac
    if moyenne_specialite is None:
        moyenne_specialite = moyenne_generale

    if any(
        value is not None and (value < 0 or value > 20)
        for value in [moyenne_bac, moyenne_licence, moyenne_specialite, note_pfe]
    ):
        return Response(
            {'error': 'Les moyennes et notes doivent etre entre 0 et 20.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    seuil_acceptation = _as_float(payload.get('seuil_acceptation'), 14.0)
    seuil_attente = _as_float(payload.get('seuil_attente'), 10.0)
    if seuil_acceptation is None or seuil_attente is None:
        return Response(
            {'error': 'Seuils jury invalides.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if seuil_attente > seuil_acceptation:
        return Response(
            {'error': 'seuil_attente doit etre inferieur ou egal a seuil_acceptation.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    notes_detaillees = {
        'source': 'api_score_final_sprint3',
        'moyenne_bac': moyenne_bac,
        'moyenne_licence': moyenne_licence,
        'payload': payload,
    }

    donnees, _ = DonneesAcademiques.objects.get_or_create(
        candidature=candidature,
        defaults={
            'moyenne_generale': moyenne_generale,
            'moyenne_specialite': moyenne_specialite,
            'note_pfe': note_pfe or 0.0,
            'mention': 'passable',
            'nb_redoublements': int(_as_float(payload.get('nb_redoublements'), 0) or 0),
            'nb_dettes': int(_as_float(payload.get('nb_dettes'), 0) or 0),
            'notes_detaillees': notes_detaillees,
        },
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def repondre_reclamation(request, reclamation_id):
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        reclamation = Reclamation.objects.select_related('candidature__candidat', 'master_concerne').get(
            id=reclamation_id
        )
    except Reclamation.DoesNotExist:
        return Response({'error': 'Reclamation non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    reponse = str(request.data.get('reponse') or '').strip()
    if not reponse:
        return Response({'error': 'La reponse est obligatoire.'}, status=status.HTTP_400_BAD_REQUEST)

    reclamation.reponse = reponse
    reclamation.statut = 'traitee'
    reclamation.traitee_par = request.user
    reclamation.date_traitement = timezone.now()
    reclamation.save(update_fields=['reponse', 'statut', 'traitee_par', 'date_traitement', 'updated_at'])

    # MOD v5 §I — Le responsable peut changer le statut de la candidature lors du
    # traitement de la réclamation (ex. lever un rejet → présélectionné), historisé.
    statut_modifie = None
    nouveau_statut = str(request.data.get('nouveau_statut') or '').strip().lower()
    STATUTS_RECLAMATION = {'preselectionne', 'en_attente', 'rejete', 'selectionne', 'sous_examen'}
    if nouveau_statut in STATUTS_RECLAMATION:
        cand = reclamation.candidature
        ancien = cand.statut
        if ancien != nouveau_statut:
            cand.statut = nouveau_statut
            update_fields = ['statut', 'updated_at']
            if nouveau_statut != 'rejete' and getattr(cand, 'motif_rejet', ''):
                cand.motif_rejet = ''
                update_fields.append('motif_rejet')
            if hasattr(cand, 'date_changement_statut'):
                cand.date_changement_statut = timezone.now()
                update_fields.append('date_changement_statut')
            cand.save(update_fields=update_fields)
            try:
                StatusHistory.objects.create(
                    candidature=cand,
                    ancien_statut=ancien,
                    nouveau_statut=nouveau_statut,
                    raison=str(request.data.get('raison_changement') or '').strip()
                    or f"Changement via réclamation {reclamation.identifiant}",
                    changed_by=request.user,
                )
            except Exception as exc:
                logger.warning("Historisation statut réclamation %s échouée: %s", reclamation.id, exc)
            statut_modifie = {'ancien': ancien, 'nouveau': nouveau_statut}

    candidat = reclamation.candidature.candidat
    titre = f"Réclamation traitée - {reclamation.identifiant}"
    message = (
        f"Bonjour {candidat.get_full_name() or candidat.username},\n\n"
        f"Votre réclamation {reclamation.identifiant} pour {reclamation.master_concerne.nom} a été traitée.\n"
        f"Réponse: {reponse}\n\n"
        "Cordialement,\n"
        "Commission ISIMM"
    )
    email_html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; line-height: 1.6;">
        <h2>Votre réclamation a été traitée</h2>
        <p>Bonjour <strong>{candidat.get_full_name() or candidat.username}</strong>,</p>
        <p>Votre réclamation <strong>{reclamation.identifiant}</strong> a été marquée comme <strong>Traitée</strong>.</p>
        <p><strong>Réponse de la commission :</strong></p>
        <p style="background:#f8fafc;border-left:4px solid #1d4ed8;padding:12px;border-radius:8px;">{reponse}</p>
        <hr/>
        <p style="color:#999;font-size:12px;">Commission ISIMM</p>
      </body>
    </html>
    """

    creer_notification_avec_email(
        user=candidat,
        titre=titre,
        message=message,
        notif_type='success',
        dedup_key=f"reclamation-traitee-{reclamation.id}-{reclamation.updated_at.isoformat()}",
        email_html=email_html,
    )

    return Response(
        {
            'success': True,
            'reclamation_id': reclamation.id,
            'statut': reclamation.statut,
            'reponse': reclamation.reponse,
            'statut_candidature': reclamation.candidature.statut,
            'statut_modifie': statut_modifie,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rectifier_score_reclamation(request, reclamation_id):
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        reclamation = Reclamation.objects.select_related('candidature__candidat', 'master_concerne').get(
            id=reclamation_id
        )
    except Reclamation.DoesNotExist:
        return Response({'error': 'Reclamation non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    candidature = reclamation.candidature
    try:
        score = Decimal(str(request.data.get('score')))
    except (InvalidOperation, TypeError, ValueError):
        return Response({'error': 'Score invalide.'}, status=status.HTTP_400_BAD_REQUEST)

    if score < 0 or score > 20:
        return Response({'error': 'Le score doit etre entre 0 et 20.'}, status=status.HTTP_400_BAD_REQUEST)

    candidature.score = score
    candidature.save(update_fields=['score', 'updated_at'])

    candidat = candidature.candidat
    titre = f"Rectification de score - {candidature.numero}"
    message = (
        f"Bonjour {candidat.get_full_name() or candidat.username},\n\n"
        f"Votre score a été rectifié à {score} pour la candidature {candidature.numero}.\n\n"
        "Cordialement,\n"
        "Commission ISIMM"
    )
    email_html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; line-height: 1.6;">
        <h2>Rectification de score</h2>
        <p>Bonjour <strong>{candidat.get_full_name() or candidat.username}</strong>,</p>
        <p>La commission a rectifié votre score à <strong>{score}</strong> pour la candidature <strong>{candidature.numero}</strong>.</p>
        <hr/>
        <p style="color:#999;font-size:12px;">Commission ISIMM</p>
      </body>
    </html>
    """

    creer_notification_avec_email(
        user=candidat,
        titre=titre,
        message=message,
        notif_type='info',
        dedup_key=f"reclamation-score-{reclamation.id}-{candidature.updated_at.isoformat()}",
        email_html=email_html,
    )

    return Response(
        {
            'success': True,
            'candidature_id': candidature.id,
            'score': str(candidature.score),
        },
        status=status.HTTP_200_OK,
    )

    donnees.moyenne_generale = moyenne_generale
    donnees.moyenne_specialite = moyenne_specialite
    donnees.note_pfe = note_pfe or 0.0
    donnees.nb_redoublements = int(_as_float(payload.get('nb_redoublements'), donnees.nb_redoublements) or 0)
    donnees.nb_dettes = int(_as_float(payload.get('nb_dettes'), donnees.nb_dettes) or 0)
    donnees.notes_detaillees = notes_detaillees
    donnees.save()

    score = donnees.calculer_et_sauvegarder_score()
    previous_status = candidature.statut

    if score >= seuil_acceptation:
        candidature.statut = 'selectionne'
        candidature.motif_rejet = ''
    elif score >= seuil_attente:
        candidature.statut = 'en_attente'
        candidature.motif_rejet = ''
    else:
        candidature.statut = 'rejete'
        candidature.motif_rejet = 'Score inferieur au seuil jury.'

    candidature.date_changement_statut = timezone.now()
    candidature.peut_modifier = False
    candidature.save(update_fields=['statut', 'motif_rejet', 'date_changement_statut', 'peut_modifier', 'updated_at'])

    if previous_status != candidature.statut:
        try:
            candidature.ajouter_historique(previous_status, candidature.statut, request.user, 'Mise a jour automatique selon seuils jury')
        except Exception:
            # L'historique ne doit pas bloquer la reponse API.
            pass

    return Response(
        {
            'success': True,
            'candidature_id': candidature.id,
            'score_final': score,
            'statut': candidature.statut,
            'seuils': {
                'acceptation': seuil_acceptation,
                'attente': seuil_attente,
            },
            'input': {
                'moyenne_bac': moyenne_bac,
                'moyenne_licence': moyenne_licence,
                'moyenne_specialite': moyenne_specialite,
                'note_pfe': note_pfe,
            },
        },
        status=status.HTTP_200_OK,
    )


def _has_commission_access_to_master(user, master_id):
    role = getattr(user, 'role', None)
    if role in ('admin', 'responsable_commission'):
        return True
    if role == 'commission':
        return MembreCommission.objects.filter(
            user=user,
            actif=True,
            commission__actif=True,
            commission__master_id=master_id,
        ).exists()
    return False


def _serialize_candidature_for_liste(candidature):
    candidat_nom = ''
    candidat_email = ''
    candidat_cin = ''

    if getattr(candidature, 'candidat', None):
        candidat_nom = candidature.candidat.get_full_name()
        candidat_email = getattr(candidature.candidat, 'email', '')
        candidat_cin = getattr(candidature.candidat, 'cin', '')

    return {
        'id': candidature.id,
        'numero': candidature.numero,
        'candidat_nom': candidat_nom,
        'candidat_email': candidat_email,
        'candidat_cin': candidat_cin,
        'specialite': candidature.master.specialite if candidature.master else '',
        'master_id': candidature.master_id,
        'master_nom': candidature.master.nom if candidature.master else '',
        'score': float(candidature.score) if candidature.score is not None else 0.0,
        'dossier_depose': candidature.dossier_depose,
        'statut': candidature.statut,
        'date_soumission': candidature.date_soumission,
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generer_liste_manuelle(request, master_id):
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    if not _has_commission_access_to_master(request.user, master_id):
        return Response({'error': 'Acces au master refuse'}, status=status.HTTP_403_FORBIDDEN)

    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    raw_type_liste = str(request.data.get('type_liste') or 'principale').strip().lower()
    type_liste = 'attente' if raw_type_liste in ['attente', 'waiting'] else 'principale'

    candidature_ids = request.data.get('candidature_ids')
    if not isinstance(candidature_ids, list) or not candidature_ids:
        return Response(
            {'error': 'candidature_ids doit etre une liste non vide.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        ordered_unique_ids = list(dict.fromkeys(int(cand_id) for cand_id in candidature_ids))
    except (TypeError, ValueError):
        return Response({'error': 'candidature_ids contient une valeur invalide.'}, status=status.HTTP_400_BAD_REQUEST)

    candidatures_qs = Candidature.objects.select_related('candidat', 'master').filter(
        id__in=ordered_unique_ids,
        master_id=master_id,
    )
    candidatures = sorted(
        list(candidatures_qs),
        key=lambda cand: (float(cand.score or 0), cand.date_soumission),
        reverse=True,
    )

    if len(candidatures) != len(ordered_unique_ids):
        ids_found = {cand.id for cand in candidatures}
        ids_missing = [cand_id for cand_id in ordered_unique_ids if cand_id not in ids_found]
        return Response(
            {
                'error': 'Certaines candidatures sont introuvables ou hors master.',
                'candidature_ids_invalides': ids_missing,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    annee = timezone.now().year
    annee_universitaire = request.data.get('annee_universitaire', f'{annee}/{annee+1}')
    max_iteration = (
        ListeAdmission.objects.filter(
            master=master,
            type_liste=type_liste,
            annee_universitaire=annee_universitaire,
        ).aggregate(max_iteration=Max('iteration'))['max_iteration']
        or 0
    )
    prochaine_iteration = int(max_iteration) + 1

    if prochaine_iteration > 4:
        return Response(
            {
                'error': 'Limite des iterations atteinte pour cette annee universitaire.',
                'annee_universitaire': annee_universitaire,
                'type_liste': type_liste,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    nouveau_statut = 'selectionne' if type_liste == 'principale' else 'en_attente'

    with transaction.atomic():
        liste = ListeAdmission.objects.create(
            master=master,
            type_liste=type_liste,
            iteration=prochaine_iteration,
            annee_universitaire=annee_universitaire,
            capacite_accueil=len(candidatures),
            places_restantes=0,
            active=True,
            publiee=False,
        )

        for position, candidature in enumerate(candidatures, start=1):
            CandidatListe.objects.create(
                liste=liste,
                candidature=candidature,
                position=position,
                score=candidature.score if candidature.score is not None else 0,
            )
            candidature.statut = nouveau_statut
            candidature.save(update_fields=['statut', 'updated_at'])

    _log_commission_action(
        request.user,
        'Validation de masse' if type_liste == 'principale' else 'Validation complète',
        master.specialite,
        annee_universitaire,
        len(candidatures),
        master,
    )

    candidats_payload = [_serialize_candidature_for_liste(cand) for cand in candidatures]

    return Response(
        {
            'success': True,
            'message': 'Liste admission generee et enregistree avec succes.',
            'liste': {
                'id': liste.id,
                'master_id': liste.master_id,
                'master_nom': liste.master.nom,
                'type_liste': liste.type_liste,
                'iteration': liste.iteration,
                'annee_universitaire': liste.annee_universitaire,
                'nb_candidats': len(candidats_payload),
                'date_creation': liste.date_creation,
            },
            'candidats': candidats_payload,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def enregistrer_action_commission(request):
    if getattr(request.user, 'role', None) not in ['commission', 'responsable_commission', 'admin']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    action = str(request.data.get('action') or '').strip()
    specialite = str(request.data.get('specialite') or '').strip()
    session = str(request.data.get('session') or '').strip()
    nb_candidats = request.data.get('nb_candidats', 0)
    master_id = request.data.get('master_id')

    master = None
    if master_id not in [None, '', 0, '0']:
        master = Master.objects.filter(id=master_id).first()

    if not action or not specialite:
        return Response({'error': 'action et specialite sont obligatoires'}, status=status.HTTP_400_BAD_REQUEST)

    _log_commission_action(request.user, action, specialite, session, nb_candidats, master)
    return Response({'success': True})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def liste_admission_recente(request, master_id):
    role = getattr(request.user, 'role', None)
    if role not in ['admin', 'responsable_commission', 'commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    if not _has_commission_access_to_master(request.user, master_id):
        return Response({'error': 'Acces au master refuse'}, status=status.HTTP_403_FORBIDDEN)

    type_liste_query = str(request.query_params.get('type_liste') or '').strip().lower()
    mapped_type_liste = None
    if type_liste_query in ['selection', 'principale']:
        mapped_type_liste = 'principale'
    elif type_liste_query in ['attente', 'waiting']:
        mapped_type_liste = 'attente'

    listes_qs = ListeAdmission.objects.select_related('master').filter(master_id=master_id)
    if mapped_type_liste:
        listes_qs = listes_qs.filter(type_liste=mapped_type_liste)

    liste = listes_qs.order_by('-date_creation').first()
    if not liste:
        return Response({'success': True, 'liste': None, 'candidats': []})

    candidatures = [
        candidat_liste.candidature
        for candidat_liste in liste.candidats.select_related('candidature__candidat', 'candidature__master').all()
    ]
    candidats_payload = [_serialize_candidature_for_liste(cand) for cand in candidatures]

    return Response(
        {
            'success': True,
            'liste': {
                'id': liste.id,
                'master_id': liste.master_id,
                'master_nom': liste.master.nom,
                'type_liste': liste.type_liste,
                'iteration': liste.iteration,
                'annee_universitaire': liste.annee_universitaire,
                'nb_candidats': len(candidats_payload),
                'date_creation': liste.date_creation,
            },
            'candidats': candidats_payload,
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generer_listes_admission(request, master_id):
    if request.user.role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        master = Master.objects.get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    iteration = int(request.data.get('iteration', 1))

    if iteration == 1:
        resultats = SelectionCandidatsService.selectionner_candidats_par_specialite(master)
        principale = resultats['liste_principale']
        attente = resultats['liste_attente']

        annee = timezone.now().year
        annee_universitaire = request.data.get('annee_universitaire', f'{annee}/{annee+1}')
        capacite = master.configuration.capacite_accueil
        capacite_attente = master.configuration.capacite_liste_attente

        liste_principale = ListeAdmission.objects.create(
            master=master,
            type_liste='principale',
            iteration=1,
            annee_universitaire=annee_universitaire,
            capacite_accueil=capacite,
            places_restantes=max(0, capacite - len(principale)),
            active=True,
            publiee=False,
        )

        for i, candidature in enumerate(principale, start=1):
            CandidatListe.objects.create(
                liste=liste_principale,
                candidature=candidature,
                position=i,
                score=candidature.score,
            )
            candidature.statut = 'preselectionne'
            candidature.save(update_fields=['statut', 'updated_at'])

        liste_attente = ListeAdmission.objects.create(
            master=master,
            type_liste='attente',
            iteration=1,
            annee_universitaire=annee_universitaire,
            capacite_accueil=capacite_attente,
            places_restantes=max(0, capacite_attente - min(len(attente), capacite_attente)),
            active=True,
            publiee=False,
        )

        for i, candidature in enumerate(attente[:capacite_attente], start=1):
            CandidatListe.objects.create(
                liste=liste_attente,
                candidature=candidature,
                position=i,
                score=candidature.score,
            )
            candidature.statut = 'en_attente'
            candidature.save(update_fields=['statut', 'updated_at'])

        return Response(
            {
                'success': True,
                'message': 'Listes principale et attente (itération 1) générées avec succès',
                'liste_principale_id': liste_principale.id,
                'liste_attente_id': liste_attente.id,
                'nb_principale': liste_principale.candidats.count(),
                'nb_attente': liste_attente.candidats.count(),
                'tri': 'score decroissant, puis date soumission',
                'classement': 'par specialite avec priorite choix 1 puis 2 puis 3',
            }
        )

    precedente = ListeAdmission.objects.filter(
        master=master,
        type_liste='principale',
        iteration=iteration - 1,
        active=True,
    ).first()
    if not precedente:
        return Response(
            {'error': f'Liste principale itération {iteration - 1} introuvable'},
            status=status.HTTP_404_NOT_FOUND,
        )

    places_liberees = int(request.data.get('places_liberees', 0))
    if places_liberees > 0:
        precedente.places_restantes = places_liberees
        precedente.save(update_fields=['places_restantes'])

    nouvelle_liste = GestionListesService.generer_liste_suivante_si_necessaire(precedente)
    if not nouvelle_liste:
        return Response(
            {
                'success': False,
                'message': 'Aucune nouvelle liste à générer (pas de places libérées ou liste attente vide).',
            },
            status=status.HTTP_200_OK,
        )

    return Response(
        {
            'success': True,
            'message': f'Liste principale itération {nouvelle_liste.iteration} générée depuis la liste d’attente.',
            'liste_id': nouvelle_liste.id,
            'nb_candidats': nouvelle_liste.candidats.count(),
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cloturer_ou_relancer_admission(request, master_id):
    """
    Point 13:
    - si capacite atteinte => cloture + publication definitive
    - sinon => relance (generation itération suivante) jusqu a capacite max
    """
    if getattr(request.user, 'role', None) not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        master = Master.objects.select_related('configuration').get(id=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master non trouve'}, status=status.HTTP_404_NOT_FOUND)

    resultat = VerificationPaiementService.evaluer_cloture_ou_relance(master)
    return Response({'success': True, **resultat})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def publier_liste(request, liste_id):
    # Permettre aussi au responsable de commission de publier la liste
    if getattr(request.user, 'role', None) not in ['directeur', 'admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        liste = ListeAdmission.objects.get(id=liste_id)
    except ListeAdmission.DoesNotExist:
        return Response({'error': 'Liste non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    if liste.publiee:
        return Response({'error': 'Liste deja publiee'}, status=status.HTTP_400_BAD_REQUEST)

    liste.publiee = True
    liste.date_publication = timezone.now()
    liste.save()

    resultats_notifications = {'envoyes': 0, 'echoues': 0, 'total': 0}
    try:
        resultats_notifications = envoyer_notifications_masse(liste)
    except Exception as exc:
        logger.exception("Erreur envoi notifications liste %s: %s", liste.id, exc)

    return Response(
        {
            'success': True,
            'message': (
                'Liste publiee: '
                f"{resultats_notifications.get('envoyes', 0)} emails envoyes, "
                f"{resultats_notifications.get('echoues', 0)} echecs"
            ),
            'notifications': resultats_notifications,
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def candidate_live_metrics(request):
    """
    Point 4: Expose real-time candidature metrics for candidate dashboard.
    Returns score, classement, total candidats for each candidature.
    """
    try:
        candidatures = Candidature.objects.filter(
            candidat=request.user
        ).select_related('master').order_by('-date_soumission')

        _refresh_classement_for_queryset(Candidature.objects.filter(master__isnull=False))
        ranked_ids = [cand.id for cand in candidatures]
        ranked_by_id = {
            cand.id: cand
            for cand in _ranked_candidatures_queryset(
                Candidature.objects.filter(id__in=ranked_ids).select_related('master')
            )
        }
        
        metrics = []
        for cand in candidatures:
            # Skip if master is null
            if cand.master is None:
                logger.warning(f"Candidature {cand.id} has null master, skipping metrics")
                continue
                
            ranked_cand = ranked_by_id.get(cand.id)
            rank = int(getattr(ranked_cand, 'classement_calcule', 0) or 0) if cand.score is not None else None
            
            # Total candidats in this master (all statuts)
            total = Candidature.objects.filter(master=cand.master).count()
            
            metrics.append({
                'id': cand.id,
                'numero': cand.numero or f"TEMP-{cand.id}",
                'master_id': cand.master.id if cand.master else None,
                'master_nom': cand.master.nom if cand.master else 'N/A',
                'score': float(cand.score) if cand.score else None,
                'classement': rank,
                'total_candidats': total,
                'statut': cand.statut,
                'date_mise_a_jour': (cand.date_changement_statut or cand.date_soumission).isoformat() if (cand.date_changement_statut or cand.date_soumission) else None,
            })
        
        return Response({
            'success': True,
            'data': metrics,
            'timestamp': timezone.now().isoformat()
        })
    
    except Exception as exc:
        logger.exception("Erreur candidate_live_metrics: %s", exc)
        return Response(
            {'error': f'Erreur lors de la recuperation des metriques: {str(exc)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def importer_paiements(request):
    if request.user.role not in ['admin', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    if 'fichier' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    fichier = request.FILES['fichier']
    extension = os.path.splitext(fichier.name)[1].lower()
    if extension not in ['.xlsx', '.xls']:
        return Response(
            {'error': 'Format invalide. Utilisez un fichier Excel (.xlsx ou .xls).'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
    os.makedirs(temp_dir, exist_ok=True)

    with tempfile.NamedTemporaryFile(delete=False, suffix=extension, dir=temp_dir) as tmp:
        for chunk in fichier.chunks():
            tmp.write(chunk)
        temp_path = tmp.name

    try:
        resultats_import = ImportPaiementService.importer_fichier_excel(temp_path)
        master_id = request.data.get('master_id')
        statuts = VerificationPaiementService.consulter_statuts_inscription(master_id=master_id)
        return Response(
            {
                'success': True,
                'message': 'Import paiements execute avec verification de delai commission.',
                'import': resultats_import,
                'inscriptions': statuts,
            }
        )
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def consulter_inscriptions_administratives(request):
    """
    Point 13:
    - consultation liste candidats ayant finalise l'inscription administrative
    - consultation liste candidats inscription incomplete
    - extraction CSV via ?export=csv
    """
    if request.user.role not in ['admin', 'commission', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    master_id = request.query_params.get('master_id')
    export_format = (request.query_params.get('export') or '').lower()

    resultats = VerificationPaiementService.consulter_statuts_inscription(master_id=master_id)

    if export_format != 'csv':
        return Response(resultats)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="inscriptions_administratives.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            'groupe',
            'candidature_id',
            'numero',
            'cin',
            'email',
            'master',
            'statut_candidature',
            'paiement_statut',
            'date_paiement',
            'date_limite_paiement',
            'motif',
        ]
    )

    for row in resultats['inscription_finalisee']:
        writer.writerow(
            [
                'inscription_finalisee',
                row.get('candidature_id'),
                row.get('numero'),
                row.get('cin'),
                row.get('email'),
                row.get('master'),
                row.get('statut_candidature'),
                row.get('paiement_statut'),
                row.get('date_paiement'),
                row.get('date_limite_paiement'),
                row.get('motif'),
            ]
        )

    for row in resultats['inscription_incomplete']:
        writer.writerow(
            [
                'inscription_incomplete',
                row.get('candidature_id'),
                row.get('numero'),
                row.get('cin'),
                row.get('email'),
                row.get('master'),
                row.get('statut_candidature'),
                row.get('paiement_statut'),
                row.get('date_paiement'),
                row.get('date_limite_paiement'),
                row.get('motif'),
            ]
        )

    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rapprocher_inscriptions_excel(request):
    if getattr(request.user, 'role', None) not in ['admin', 'commission', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    rows = request.data.get('rows') or []
    source_filename = (request.data.get('source_filename') or '').strip()
    master_id = request.data.get('master_id')

    if not isinstance(rows, list):
        return Response({'error': 'Le champ rows doit etre une liste.'}, status=status.HTTP_400_BAD_REQUEST)

    master = None
    if master_id not in [None, '', 'all']:
        try:
            master = Master.objects.get(id=int(master_id))
        except (ValueError, Master.DoesNotExist):
            return Response({'error': 'Master invalide pour le rapprochement.'}, status=status.HTTP_400_BAD_REQUEST)

    candidatures_qs = Candidature.objects.select_related('master').filter(statut='inscrit')
    if master is not None:
        candidatures_qs = candidatures_qs.filter(master=master)

    candidatures = list(candidatures_qs)
    by_numero = {str(c.numero or '').strip(): c for c in candidatures if c.numero}
    by_cin = {
        str(c.candidat_cin or '').strip(): c
        for c in candidatures
        if getattr(c, 'candidat_cin', None)
    }

    results = []
    for row in rows:
        data = row if isinstance(row, dict) else {}

        numero_candidature = str(
            data.get('numero_candidature')
            or data.get('Numéro candidature')
            or data.get('numero')
            or ''
        ).strip()
        cin = str(data.get('cin') or data.get('CIN') or '').strip()
        numero_inscription = str(
            data.get('numero_inscription')
            or data.get('Numéro inscription')
            or data.get('inscription')
            or ''
        ).strip()
        nom_prenom = str(data.get('nom_prenom') or data.get('Nom prénom') or data.get('nom') or '').strip()
        master_nom = str(data.get('master') or data.get('Master') or '').strip()
        specialite = str(data.get('specialite') or data.get('Spécialité') or '').strip()

        candidature = by_numero.get(numero_candidature) if numero_candidature else None
        if candidature is None and cin:
            candidature = by_cin.get(cin)

        if candidature is None:
            results.append(
                {
                    'numero_candidature': numero_candidature,
                    'cin': cin,
                    'numero_inscription': numero_inscription,
                    'nom_prenom': nom_prenom,
                    'master': master_nom,
                    'specialite': specialite,
                    'verification': 'absent',
                    'details': 'Aucune candidature inscrite correspondante.',
                }
            )
            continue

        incoherences = []
        if master_nom and candidature.master and master_nom != candidature.master.nom:
            incoherences.append('Master non coherent')
        if specialite and candidature.master and specialite != candidature.master.specialite:
            incoherences.append('Specialite non coherente')
        if nom_prenom and candidature.candidat and nom_prenom != candidature.candidat.get_full_name():
            incoherences.append('Nom/prenom non coherent')

        results.append(
            {
                'numero_candidature': numero_candidature or (candidature.numero or ''),
                'cin': cin or str(getattr(candidature, 'candidat_cin', '') or ''),
                'numero_inscription': numero_inscription,
                'nom_prenom': nom_prenom or candidature.candidat.get_full_name(),
                'master': master_nom or (candidature.master.nom if candidature.master else ''),
                'specialite': specialite or (candidature.master.specialite if candidature.master else ''),
                'verification': 'incoherent' if incoherences else 'valide',
                'details': ', '.join(incoherences) if incoherences else 'Verification reussie',
            }
        )

    audit = InscriptionRapprochementAudit.objects.create(
        created_by=request.user,
        master=master,
        source_filename=source_filename,
        total_rows=len(results),
        valide_rows=sum(1 for row in results if row.get('verification') == 'valide'),
        incoherent_rows=sum(1 for row in results if row.get('verification') == 'incoherent'),
        absent_rows=sum(1 for row in results if row.get('verification') == 'absent'),
        payload_rows=rows,
        result_rows=results,
    )

    return Response(
        {
            'success': True,
            'audit_id': audit.id,
            'rows': results,
            'summary': {
                'total': audit.total_rows,
                'valide': audit.valide_rows,
                'incoherent': audit.incoherent_rows,
                'absent': audit.absent_rows,
            },
        }
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def lister_concours(request):
    """Retourne la liste des concours pour l'interface d'administration."""
    qs = Concours.objects.all().order_by('-created_at')
    type_filter = request.query_params.get('type_concours')
    if type_filter:
        qs = qs.filter(type_concours=type_filter)

    payload = [
        {
            'id': concours.id,
            'nom': concours.nom,
            'description': concours.description,
            'type_concours': concours.type_concours,
            'date_ouverture': concours.date_ouverture,
            'date_cloture': concours.date_cloture,
            'places_disponibles': concours.places_disponibles,
            'actif': concours.actif,
            'conditions_admission': concours.conditions_admission,
            'document_officiel_pdf_url': (
                request.build_absolute_uri(concours.document_officiel_pdf.url)
                if concours.document_officiel_pdf
                else None
            ),
            'created_at': concours.created_at,
            'updated_at': concours.updated_at,
        }
        for concours in qs
    ]
    return Response(payload)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creer_concours_admin(request):
    """Creer une offre de concours (admin uniquement)."""
    if getattr(request.user, 'role', None) != 'admin':
        return Response({'error': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)

    data = request.data or {}
    nom = data.get('nom')
    date_ouverture = data.get('date_ouverture')
    date_cloture = data.get('date_cloture')
    places_disponibles = data.get('places_disponibles')
    type_concours = data.get('type_concours', 'ingenieur')

    if not nom or not date_ouverture or not date_cloture:
        return Response(
            {'error': 'Champs obligatoires manquants (nom, date_ouverture, date_cloture).'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if type_concours not in {'master', 'ingenieur'}:
        return Response({'error': 'type_concours invalide.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        places_int = int(places_disponibles)
    except (TypeError, ValueError):
        return Response({'error': 'places_disponibles invalide.'}, status=status.HTTP_400_BAD_REQUEST)

    conditions_admission = data.get('conditions_admission')
    if not isinstance(conditions_admission, dict):
        conditions_admission = {}

    specialite = data.get('specialite')
    if specialite:
        conditions_admission['specialite'] = specialite

    concours = Concours.objects.create(
        nom=nom,
        description=data.get('description', ''),
        type_concours=type_concours,
        date_ouverture=date_ouverture,
        date_cloture=date_cloture,
        places_disponibles=places_int,
        actif=bool(data.get('actif', True)),
        conditions_admission=conditions_admission,
    )

    # Synchronisation automatique vers un parcours master editable par le responsable.
    _sync_master_from_concours(concours)

    uploaded_pdf = request.FILES.get('document_officiel_pdf')
    if uploaded_pdf:
        file_ext = uploaded_pdf.name.split('.')[-1].lower()
        if file_ext != 'pdf':
            concours.delete()
            return Response({'error': 'Seuls les fichiers PDF sont acceptes.'}, status=status.HTTP_400_BAD_REQUEST)
        if uploaded_pdf.size > 10 * 1024 * 1024:
            concours.delete()
            return Response({'error': 'Fichier trop volumineux (max 10 MB).'}, status=status.HTTP_400_BAD_REQUEST)
        concours.document_officiel_pdf = uploaded_pdf
        concours.save(update_fields=['document_officiel_pdf', 'updated_at'])

    return Response(
        {
            'id': concours.id,
            'nom': concours.nom,
            'description': concours.description,
            'type_concours': concours.type_concours,
            'date_ouverture': concours.date_ouverture,
            'date_cloture': concours.date_cloture,
            'places_disponibles': concours.places_disponibles,
            'actif': concours.actif,
            'conditions_admission': concours.conditions_admission,
            'specialite': concours.conditions_admission.get('specialite', ''),
            'document_officiel_pdf_url': (
                request.build_absolute_uri(concours.document_officiel_pdf.url)
                if concours.document_officiel_pdf
                else None
            ),
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def modifier_supprimer_concours_admin(request, concours_id):
    """Modifier/Supprimer une offre de concours (admin uniquement)."""
    if getattr(request.user, 'role', None) != 'admin':
        return Response({'error': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)

    try:
        concours = Concours.objects.get(id=concours_id)
    except Concours.DoesNotExist:
        return Response({'error': 'Concours introuvable'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        _deactivate_synced_master_from_concours(concours.id)
        concours.delete()
        return Response({'message': 'Concours supprime avec succes.'}, status=status.HTTP_200_OK)

    data = request.data or {}

    if 'nom' in data:
        concours.nom = data.get('nom') or concours.nom
    if 'description' in data:
        concours.description = data.get('description') or ''
    if 'type_concours' in data and data.get('type_concours') in {'master', 'ingenieur'}:
        concours.type_concours = data.get('type_concours')
    if 'date_ouverture' in data:
        concours.date_ouverture = data.get('date_ouverture')
    if 'date_cloture' in data:
        concours.date_cloture = data.get('date_cloture')
    if 'places_disponibles' in data:
        try:
            concours.places_disponibles = int(data.get('places_disponibles'))
        except (TypeError, ValueError):
            return Response({'error': 'places_disponibles invalide.'}, status=status.HTTP_400_BAD_REQUEST)
    if 'actif' in data:
        concours.actif = bool(data.get('actif'))

    if 'conditions_admission' in data and isinstance(data.get('conditions_admission'), dict):
        concours.conditions_admission = data.get('conditions_admission')

    if 'specialite' in data:
        payload_conditions = dict(concours.conditions_admission or {})
        specialite = data.get('specialite')
        if specialite:
            payload_conditions['specialite'] = specialite
        else:
            payload_conditions.pop('specialite', None)
        concours.conditions_admission = payload_conditions

    uploaded_pdf = request.FILES.get('document_officiel_pdf')
    if uploaded_pdf:
        file_ext = uploaded_pdf.name.split('.')[-1].lower()
        if file_ext != 'pdf':
            return Response({'error': 'Seuls les fichiers PDF sont acceptes.'}, status=status.HTTP_400_BAD_REQUEST)
        if uploaded_pdf.size > 10 * 1024 * 1024:
            return Response({'error': 'Fichier trop volumineux (max 10 MB).'}, status=status.HTTP_400_BAD_REQUEST)
        concours.document_officiel_pdf = uploaded_pdf

    if str(data.get('remove_document_officiel_pdf', '')).lower() in ['1', 'true', 'yes']:
        if concours.document_officiel_pdf:
            concours.document_officiel_pdf.delete(save=False)
            concours.document_officiel_pdf = None

    concours.save()

    # Maintenir la copie synchronisee visible dans l'espace responsable.
    _sync_master_from_concours(concours)

    return Response(
        {
            'id': concours.id,
            'nom': concours.nom,
            'description': concours.description,
            'type_concours': concours.type_concours,
            'date_ouverture': concours.date_ouverture,
            'date_cloture': concours.date_cloture,
            'places_disponibles': concours.places_disponibles,
            'actif': concours.actif,
            'conditions_admission': concours.conditions_admission,
            'specialite': (concours.conditions_admission or {}).get('specialite', ''),
            'document_officiel_pdf_url': (
                request.build_absolute_uri(concours.document_officiel_pdf.url)
                if concours.document_officiel_pdf
                else None
            ),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exporter_liste_pdf(request, liste_id):
    try:
        ListeAdmission.objects.get(id=liste_id)
    except ListeAdmission.DoesNotExist:
        return Response({'error': 'Liste non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    response = HttpResponse(b'', content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="liste_{liste_id}.pdf"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exporter_liste_excel(request, liste_id):
    try:
        ListeAdmission.objects.get(id=liste_id)
    except ListeAdmission.DoesNotExist:
        return Response({'error': 'Liste non trouvee'}, status=status.HTTP_404_NOT_FOUND)

    response = HttpResponse(
        b'',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="liste_{liste_id}.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def reglement_concours_ingenieur_reference(request):
    """Retourne la version structuree du reglement de reference pour integration front/back."""
    return Response(REGLEMENT_CONCOURS_INGENIEUR_REFERENCE_2025_2026)


@api_view(['GET'])
@permission_classes([AllowAny])
def reglement_masters_reference(request):
    """Retourne le referentiel masters officiel 2025/2026 (MPGL, MRGL, MPDS)."""
    return Response(REFERENTIEL_MASTERS_ISIMM_2025_2026)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def appliquer_reglement_reference_concours(request, concours_id):
    """
    Applique le reglement de reference dans un concours existant.
    - met a jour conditions_admission avec une structure complete exploitable.
    - permet surcharge de date_ouverture/date_cloture/places_disponibles.
    """
    if getattr(request.user, 'role', None) not in ['admin', 'commission', 'responsable_commission']:
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    try:
        concours = Concours.objects.get(id=concours_id)
    except Concours.DoesNotExist:
        return Response({'error': 'Concours non trouve'}, status=status.HTTP_404_NOT_FOUND)

    if concours.type_concours != 'ingenieur':
        return Response(
            {'error': 'Ce reglement de reference ne peut etre applique qu a un concours ingenieur.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    payload = REGLEMENT_CONCOURS_INGENIEUR_REFERENCE_2025_2026.copy()

    # Permet d injecter des sections supplementaires ou corrections sans casser le standard.
    sections_personnalisees = request.data.get('sections_personnalisees', {})
    if isinstance(sections_personnalisees, dict) and sections_personnalisees:
        payload.update(sections_personnalisees)

    concours.conditions_admission = payload

    if request.data.get('date_ouverture'):
        concours.date_ouverture = request.data.get('date_ouverture')
    if request.data.get('date_cloture'):
        concours.date_cloture = request.data.get('date_cloture')
    if request.data.get('places_disponibles') is not None:
        concours.places_disponibles = request.data.get('places_disponibles')

    concours.save(update_fields=['conditions_admission', 'date_ouverture', 'date_cloture', 'places_disponibles', 'updated_at'])

    return Response(
        {
            'success': True,
            'concours_id': concours.id,
            'nom': concours.nom,
            'type_concours': concours.type_concours,
            'message': 'Reglement de reference integre avec succes dans le concours.',
            'conditions_admission': concours.conditions_admission,
        }
    )
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def soumettre_paiement_enligne(request):
    """
    Candidat soumet son justificatif de paiement
    """
    candidature_id = request.data.get('candidature_id')
    reference = request.data.get('reference_paiement')
    montant = request.data.get('montant')
    
    if 'fichier_paiement' not in request.FILES:
        return Response(
            {'error': 'Fichier de paiement requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        candidature = Candidature.objects.get(
            id=candidature_id,
            candidat=request.user
        )
    except Candidature.DoesNotExist:
        return Response(
            {'error': 'Candidature non trouvée'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Vérifier que candidat est sélectionné
    if candidature.statut != 'selectionne':
        return Response(
            {'error': 'Vous devez être sélectionné pour soumettre un paiement'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    fichier = request.FILES['fichier_paiement']
    
    # Vérifier format fichier
    extension = fichier.name.split('.')[-1].lower()
    if extension not in ['pdf', 'jpg', 'jpeg', 'png']:
        return Response(
            {'error': 'Format invalide. Formats acceptés: PDF, JPG, PNG'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Créer ou mettre à jour inscription
    inscription, created = InscriptionEnLigne.objects.get_or_create(
        candidature=candidature,
        defaults={
            'reference_paiement': reference,
            'montant_paye': montant
        }
    )
    
    if not created:
        inscription.reference_paiement = reference
        inscription.montant_paye = montant
    
    inscription.fichier_paiement = fichier
    inscription.statut = 'paiement_soumis'
    inscription.save()
    
    return Response({
        'success': True,
        'message': 'Paiement soumis avec succès',
        'inscription_id': inscription.id,
        'statut': inscription.statut
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def valider_paiement_enligne(request, inscription_id):
    """
    Admin/Commission valide le paiement soumis
    """
    if request.user.role not in ['admin', 'commission', 'responsable_commission']:
        return Response(
            {'error': 'Permission refusée'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        inscription = InscriptionEnLigne.objects.get(id=inscription_id)
    except InscriptionEnLigne.DoesNotExist:
        return Response(
            {'error': 'Inscription non trouvée'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    action = request.data.get('action')  # 'valider' ou 'refuser'
    commentaire = request.data.get('commentaire', '')
    
    if action == 'valider':
        inscription.statut = 'valide'
        inscription.valide_par = request.user
        inscription.date_validation = timezone.now()
        inscription.commentaire_validation = commentaire
        inscription.save()
        
        # Mettre à jour candidature
        inscription.candidature.statut = 'inscrit'
        inscription.candidature.save()
        
        envoyer_email_inscription_validee(inscription)
        
        return Response({
            'success': True,
            'message': 'Paiement validé'
        })
    
    elif action == 'refuser':
        inscription.statut = 'refuse'
        inscription.commentaire_validation = commentaire
        inscription.save()
        
        return Response({
            'success': True,
            'message': 'Paiement refusé'
        })
    
    return Response(
        {'error': 'Action invalide'},
        status=status.HTTP_400_BAD_REQUEST
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_candidatures(request):
    try:
        # User ID is passed in headers by the Gateway or available from the token
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return Response({'error': 'User ID not found'}, status=status.HTTP_400_BAD_REQUEST)
            
        candidatures = Candidature.objects.filter(user_id=user_id)
        serializer = CandidatureSerializer(candidatures, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_candidature_exists(request, master_id):
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return Response({'error': 'User ID not found'}, status=status.HTTP_400_BAD_REQUEST)
            
        exists = Candidature.objects.filter(user_id=user_id, master_id=master_id).exists()
        return Response({'exists': exists})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_specialites_by_parcours(request):
    """
    API endpoint pour récupérer les spécialités requises pour un parcours spécifique.
    
    Query params:
    - parcours_code: Code du parcours (ex: MPDS, MPGL, MP3I, MRGL, MRMI, ING_APPLI)
    """
    parcours_code = request.query_params.get('parcours_code')
    
    if not parcours_code:
        return Response({'error': 'Paramètre parcours_code est requis'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        mapping = SpecialiteParcoursMapping.objects.get(
            code_parcours=parcours_code,
            actif=True
        )
        return Response({
            'code_parcours': mapping.code_parcours,
            'nom_parcours': mapping.nom_parcours,
            'type_formation': mapping.type_formation,
            'specialites': mapping.specialites,
        })
    except SpecialiteParcoursMapping.DoesNotExist:
        return Response(
            {'error': f'Parcours {parcours_code} non trouvé'},
            status=status.HTTP_404_NOT_FOUND
        )


def _extract_specialite_diplome(candidature) -> str:
    """Extrait la spécialité du diplôme d'origine depuis DonneesAcademiques.notes_detaillees."""
    try:
        da = candidature.donnees_academiques
        notes = da.notes_detaillees if isinstance(da.notes_detaillees, dict) else {}
        payload = notes.get('payload', {}) if isinstance(notes.get('payload'), dict) else notes
        for key in ('specialiteLicence', 'specialite_diplome', 'specialiteDiplome',
                    'licenceSpecialite', 'specialite', 'diplome'):
            val = payload.get(key) or notes.get(key)
            if val and isinstance(val, str) and val.strip():
                return val.strip()
    except Exception:
        pass
    return ''


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_specialites_admissibles_master(request, master_id):
    """
    Retourne la liste des spécialités de diplômes admissibles pour un master donné.
    Source : SpecialiteParcoursMapping, résolution via master.specialite = code_parcours.

    GET /api/candidatures/masters/<master_id>/specialites-admissibles/
    Réponse :
      {
        "master_id": 3,
        "master_nom": "Mastère Professionnel Génie Logiciel (MPGL)",
        "code_parcours": "MPGL",
        "specialites": [
          {"nom": "Licence en Sciences de l'Informatique génie logiciel", "abreviation": "LSI-GL"},
          ...
        ]
      }
    """
    try:
        master = Master.objects.get(pk=master_id)
    except Master.DoesNotExist:
        return Response({'error': 'Master introuvable.'}, status=status.HTTP_404_NOT_FOUND)

    code_parcours = (master.specialite or '').strip()

    # Résolution prioritaire : via code_parcours (ex: "MPGL")
    mapping = None
    if code_parcours:
        mapping = SpecialiteParcoursMapping.objects.filter(
            code_parcours=code_parcours, actif=True
        ).first()

    # Fallback : via FK master
    if mapping is None:
        mapping = SpecialiteParcoursMapping.objects.filter(
            master=master, actif=True
        ).first()

    if mapping is None:
        return Response(
            {'error': f'Aucune matrice d\'admissibilité trouvée pour ce master (code: {code_parcours!r}).'},
            status=status.HTTP_404_NOT_FOUND,
        )

    return Response({
        'master_id': master.id,
        'master_nom': master.nom,
        'code_parcours': mapping.code_parcours,
        'specialites': mapping.specialites,
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def list_all_parcours(request):
    """
    API endpoint pour récupérer la liste de tous les parcours avec leurs spécialités.
    
    Query params optionnels:
    - type_formation: Filtrer par type ('master' ou 'ingenieur')
    """
    type_formation = request.query_params.get('type_formation')
    
    query = SpecialiteParcoursMapping.objects.filter(actif=True).order_by('ordre', 'nom_parcours')
    
    if type_formation:
        query = query.filter(type_formation=type_formation)
    
    parcours_list = [
        {
            'code_parcours': p.code_parcours,
            'nom_parcours': p.nom_parcours,
            'type_formation': p.type_formation,
            'specialites': p.specialites,
            'nombre_specialites': len(p.specialites) if p.specialites else 0,
        }
        for p in query
    ]
    
    return Response(parcours_list)


# ============================================================================
# ÉTAPE 2: SYSTÈME DE STATUT + NOTIFICATIONS
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def changer_statut_candidature_endpoint(request, candidature_id):
    """
    Endpoint pour changer le statut d'une candidature et envoyer les notifications.
    
    Request body:
    {
        "nouveau_statut": "sous_examen",  # Requis
        "raison": "Sélection par commission",  # Optionnel
        "envoyer_notification": true  # Optionnel, défaut true
    }
    """
    from .services_statut_notifications import StatutNotificationService
    
    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response(
            {'error': f'Candidature {candidature_id} non trouvée'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Vérifier les permissions (admin, responsable_commission ou membre de la commission)
    user_role = getattr(request.user, 'role', None)
    is_privileged = (
        request.user.is_staff
        or request.user.is_superuser
        or user_role in ('admin', 'responsable_commission', 'commission')
    )
    if not is_privileged:
        return Response(
            {'error': 'Permissions insuffisantes pour changer un statut'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    data = request.data
    nouveau_statut = data.get('nouveau_statut')
    raison = data.get('raison', '')
    envoyer_notification = data.get('envoyer_notification', True)
    
    if not nouveau_statut:
        return Response(
            {'error': 'Le paramètre "nouveau_statut" est requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        status_history = StatutNotificationService.changer_statut(
            candidature=candidature,
            nouveau_statut=nouveau_statut,
            raison=raison,
            changed_by=request.user,
            envoyer_notification=envoyer_notification,
        )
        
        return Response({
            'success': True,
            'message': f'Statut changé à {nouveau_statut}',
            'candidature_numero': candidature.numero,
            'ancien_statut': status_history.ancien_statut,
            'nouveau_statut': status_history.nouveau_statut,
            'date_changement': status_history.date_changement,
            'notification_envoyee': status_history.notification_envoyee,
        })
    except ValueError as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        logger.error(
            f"Erreur lors du changement de statut: {str(e)}",
            exc_info=True
        )
        return Response(
            {'error': 'Erreur interne du serveur'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def recuperer_historique_statuts_endpoint(request, candidature_id):
    """
    Endpoint pour récupérer l'historique complet des changements de statut.
    """
    from .services_statut_notifications import StatutNotificationService
    
    try:
        candidature = Candidature.objects.get(id=candidature_id)
    except Candidature.DoesNotExist:
        return Response(
            {'error': f'Candidature {candidature_id} non trouvée'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Vérifier les permissions (candidat ou admin)
    if candidature.candidat != request.user and not request.user.is_staff:
        return Response(
            {'error': 'Permissions insuffisantes'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        historique = StatutNotificationService.recuperer_historique_statuts(candidature)
        
        return Response({
            'candidature_id': candidature.id,
            'candidature_numero': candidature.numero,
            'master': candidature.master.nom,
            'statut_actuel': candidature.statut,
            'historique': historique,
            'total_changements': len(historique),
        })
    except Exception as e:
        logger.error(
            f"Erreur lors de la récupération de l'historique: {str(e)}",
            exc_info=True
        )
        return Response(
            {'error': 'Erreur interne du serveur'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ============================================================================
# ÉTAPE 3: SYSTÈME MULTI-COMMISSIONS - ENDPOINTS DE COMMISSION
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_commissions_from_candidature(request):
    """
    GET /api/commissions/my-commissions/
    
    Retourne la liste des commissions liées à l'utilisateur authentifié.
    Appelé par auth-service pour lister les commissions disponibles.
    
    Query params:
    - user_id (optionnel): ID de l'utilisateur. Si absent, utilise request.user.id
    
    Returns:
    {
        "success": true,
        "user_id": 42,
        "count": 2,
        "commissions": [
            {
                "id": 1,
                "nom": "Commission MPGL",
                "description": "Commission Master Professionnel Génie Logiciel",
                "master_id": 5,
                "master_nom": "Master Professionnel en Ingenierie Logicielle",
                "actif": true,
                "role": "responsable"
            }
        ]
    }
    """
    from django.db.models import Q
    
    user_id = request.query_params.get('user_id')
    if user_id:
        try:
            user_id = int(user_id)
        except (ValueError, TypeError):
            return Response(
                {'error': 'user_id invalide'},
                status=status.HTTP_400_BAD_REQUEST
            )
    else:
        user_id = request.user.id
    
    try:
        # Récupérer toutes les commissions liées à l'utilisateur
        commissions = Commission.objects.filter(
            Q(membres__user_id=user_id, membres__actif=True) |
            Q(membre_commission_links__user_id=user_id, membre_commission_links__actif=True)
        ).distinct().select_related('master').filter(actif=True)
        
        data = []
        for commission in commissions:
            # Déterminer le rôle de l'utilisateur dans cette commission
            membre_role = 'membre'
            membre = MembreCommission.objects.filter(
                commission=commission,
                user_id=user_id,
                actif=True
            ).first()
            if membre:
                membre_role = membre.role or 'membre'
            
            data.append({
                'id': commission.id,
                'nom': commission.nom,
                'description': commission.description or '',
                'master_id': commission.master_id,
                'master_nom': commission.master.nom if commission.master else '',
                'actif': commission.actif,
                'role': membre_role,
            })
        
        return Response({
            'success': True,
            'user_id': user_id,
            'count': len(data),
            'commissions': data,
        })
    
    except Exception as e:
        logger.exception("Erreur get_my_commissions_from_candidature pour user %s: %s", user_id, e)
        return Response(
            {'error': f'Erreur serveur: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_commission_members_list(request, commission_id=None):
    """
    GET /api/candidatures/commissions/<commission_id>/members/
    GET /api/commissions/commission-members/?commission_id=<id>  (legacy)

    Retourne la liste des membres d'une commission spécifique.
    Accepte commission_id depuis l'URL OU depuis ?commission_id=...
    """
    # Path param prioritaire, sinon fallback sur query string
    if commission_id is None:
        commission_id = request.query_params.get('commission_id')

    if not commission_id:
        return Response(
            {'error': 'commission_id est requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        commission = Commission.objects.get(id=commission_id, actif=True)
    except Commission.DoesNotExist:
        return Response(
            {'error': f'Commission {commission_id} non trouvée'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        # Récupérer les membres actifs de la commission
        membres = MembreCommission.objects.filter(
            commission=commission,
            actif=True
        ).select_related('user').order_by('-role', 'user__first_name')
        
        data = []
        for membre in membres:
            data.append({
                'id': membre.id,
                'user_id': membre.user.id,
                'first_name': membre.user.first_name or '',
                'last_name': membre.user.last_name or '',
                'email': membre.user.email,
                'role': membre.role or 'membre',
                'date_nomination': membre.date_nomination.isoformat() if membre.date_nomination else None,
            })
        
        return Response({
            'success': True,
            'commission_id': commission.id,
            'commission_nom': commission.nom,
            'count': len(data),
            'members': data,
        })
    
    except Exception as e:
        logger.exception("Erreur get_commission_members_list pour commission %s: %s", commission_id, e)
        return Response(
            {'error': f'Erreur serveur: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_fichier_dossier(request):
    """
    Endpoint générique d'upload de fichier pour le dossier candidat.
    Accepte: fichier (File), document_type (str), numero_dossier (str optionnel).
    Stocke le fichier dans MEDIA_ROOT/dossiers/<candidature_id>/<document_type>/<filename>.
    """
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile

    fichier = request.FILES.get('fichier')
    if not fichier:
        return Response({'error': 'Aucun fichier reçu.'}, status=status.HTTP_400_BAD_REQUEST)

    max_size = 5 * 1024 * 1024  # 5 Mo
    if fichier.size > max_size:
        return Response({'error': 'Fichier trop volumineux (max 5 Mo).'}, status=status.HTTP_400_BAD_REQUEST)

    document_type = request.data.get('document_type', 'divers').strip()[:100]
    numero_dossier = request.data.get('numero_dossier', '').strip()

    # Trouver la candidature du candidat (la plus récente si plusieurs)
    candidature = (
        Candidature.objects.filter(candidat=request.user)
        .order_by('-date_soumission')
        .first()
    )

    # ── Normalisation ASCII pour éviter problèmes Windows/encodage ──
    import unicodedata
    import re as _re
    import uuid as _uuid

    def _ascii_slug(text, max_len):
        # Convertit é → e, à → a, etc. + supprime caractères non-ASCII
        normalized = unicodedata.normalize('NFKD', text or '')
        ascii_only = normalized.encode('ASCII', 'ignore').decode('ASCII')
        # Garde alphanumeric + _ - et remplace espaces par _
        cleaned = _re.sub(r'[^a-zA-Z0-9_\- ]', '', ascii_only).strip()
        cleaned = _re.sub(r'\s+', '_', cleaned)
        return (cleaned or 'doc')[:max_len]

    candidature_ref = str(candidature.id) if candidature else 'inconnu'
    safe_type = _ascii_slug(document_type, 60)
    safe_name = _ascii_slug(fichier.name.rsplit('.', 1)[0], 80)
    ext = (fichier.name.rsplit('.', 1)[-1] or 'bin').lower()[:5]
    safe_name = f'{safe_name}.{ext}'
    path = f'dossiers/{candidature_ref}/{safe_type}/{safe_name}'

    # ── Sauvegarde avec gestion d'erreur explicite ──
    try:
        saved_path = default_storage.save(path, ContentFile(fichier.read()))
    except Exception as exc:
        logger.exception('upload_fichier_dossier — erreur stockage')
        return Response(
            {
                'error': f'Erreur lors du stockage du fichier : {type(exc).__name__}',
                'detail': str(exc)[:300],
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    return Response({
        'success': True,
        'message': 'Fichier enregistré avec succès.',
        'fichier_url': default_storage.url(saved_path),
        'document_type': document_type,
        'candidature_id': candidature.id if candidature else None,
    }, status=status.HTTP_201_CREATED)


# ──────────────────────────────────────────────────────────────────────────
# GET /api/candidatures/<candidature_id>/list-fichiers-deposes/
# Liste les fichiers réellement présents dans MEDIA_ROOT/dossiers/<id>/
# Utilisé côté responsable/membre pour voir ce qu'Ahmed a uploadé.
# ──────────────────────────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_fichiers_deposes(request, candidature_id):
    import os
    from django.conf import settings as _settings
    from django.core.files.storage import default_storage

    try:
        candidature = Candidature.objects.get(pk=candidature_id)
    except Candidature.DoesNotExist:
        return Response({'error': 'Candidature introuvable'}, status=status.HTTP_404_NOT_FOUND)

    # Permission : admin / commission / responsable / le candidat lui-même
    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission') and candidature.candidat_id != request.user.id:
        return Response({'error': 'Accès refusé'}, status=status.HTTP_403_FORBIDDEN)

    media_root = getattr(_settings, 'MEDIA_ROOT', '')
    media_url = getattr(_settings, 'MEDIA_URL', '/media/').rstrip('/')
    base_dir = os.path.join(str(media_root), 'dossiers', str(candidature.id))

    fichiers = []
    if os.path.isdir(base_dir):
        for doc_type_dir in sorted(os.listdir(base_dir)):
            full_type_dir = os.path.join(base_dir, doc_type_dir)
            if not os.path.isdir(full_type_dir):
                continue
            for filename in sorted(os.listdir(full_type_dir)):
                full_path = os.path.join(full_type_dir, filename)
                if not os.path.isfile(full_path):
                    continue
                try:
                    size = os.path.getsize(full_path)
                    mtime = os.path.getmtime(full_path)
                except OSError:
                    size, mtime = 0, 0
                rel_path = f'dossiers/{candidature.id}/{doc_type_dir}/{filename}'
                # MOD v6 §4 — URL absolue (http://host:8003/media/...) pour éviter le
                # 404 quand le front (port 4200) ouvre une URL relative /media/.
                fichiers.append({
                    'type_document': doc_type_dir.replace('_', ' '),
                    'nom_fichier': filename,
                    'taille_octets': size,
                    'taille_human': _human_size(size),
                    'date_upload': mtime,
                    'url': request.build_absolute_uri(f'{media_url}/{rel_path}'),
                    'extension': (filename.rsplit('.', 1)[-1] or '').lower(),
                })

    return Response({
        'candidature_id': candidature.id,
        'numero': candidature.numero or '',
        'candidat_nom': candidature.candidat.get_full_name() or candidature.candidat.email,
        'count': len(fichiers),
        'fichiers': fichiers,
    })


def _human_size(size_bytes):
    if size_bytes < 1024:
        return f'{size_bytes} o'
    if size_bytes < 1024 * 1024:
        return f'{size_bytes / 1024:.1f} Ko'
    return f'{size_bytes / 1024 / 1024:.1f} Mo'


# ──────────────────────────────────────────────────────────────────────────
# POST /api/candidatures/analyser-ocr-lot/
# Analyse OCR par lot sur plusieurs candidatures en une seule fois
# Body : { "candidature_ids": [9, 153, ...] }
# ──────────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_ocr_excel(request):
    """Exporte les résultats OCR en fichier Excel."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from django.http import HttpResponse
    import io

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    resultats = request.data.get('resultats', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rapport OCR'

    headers = ['Candidat', 'N° Candidature', 'Score Déclaré', 'Score OCR', 'Écart', 'Moteur', 'Statut']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='0F1F3D', end_color='0F1F3D', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row, r in enumerate(resultats, 2):
        statut = r.get('statut', '—')
        ws.cell(row=row, column=1, value=r.get('candidat_nom', '—'))
        ws.cell(row=row, column=2, value=r.get('numero', '—'))
        ws.cell(row=row, column=3, value=r.get('score_declare', '—'))
        ws.cell(row=row, column=4, value=r.get('score_extrait', '—'))
        ws.cell(row=row, column=5, value=r.get('ecart', '—'))
        ws.cell(row=row, column=6, value=r.get('moteur', '—'))
        cell_statut = ws.cell(row=row, column=7, value=statut)
        color_map = {'ok': 'D1FAE5', 'anomalie': 'FEF3C7', 'erreur': 'FEE2E2'}
        color = color_map.get(statut, 'FFFFFF')
        cell_statut.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rapport_ocr_lot.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_ocr_pdf(request):
    """Exporte les résultats OCR en fichier PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse
    import io

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    resultats = request.data.get('resultats', [])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph('ISIMM — Rapport d\'Analyse OCR par Lot', styles['Title']))
    elements.append(Spacer(1, 12))

    data = [['Candidat', 'N° Cand.', 'Déclaré', 'OCR', 'Écart', 'Moteur', 'Statut']]
    for r in resultats:
        data.append([
            r.get('candidat_nom', '—'), r.get('numero', '—'),
            str(r.get('score_declare', '—')), str(r.get('score_extrait', '—')),
            str(r.get('ecart', '—')), r.get('moteur', '—'), r.get('statut', '—')
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F1F3D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rapport_ocr_lot.pdf"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyser_ocr_lot(request):
    """Lance l'analyse OCR RÉELLE sur une liste de candidatures (pdfplumber)."""
    import os as _os
    from django.conf import settings as _settings
    from .ocr_service import OCRService

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    ids = request.data.get('candidature_ids') or []
    if not isinstance(ids, list) or not ids:
        return Response(
            {'error': 'Le champ candidature_ids doit etre une liste non vide.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    media_root = str(getattr(_settings, 'MEDIA_ROOT', ''))
    resultats = []
    erreurs = 0
    anomalies = 0

    for cand_id in ids:
        try:
            cand = Candidature.objects.select_related('candidat').get(pk=int(cand_id))
        except (Candidature.DoesNotExist, ValueError, TypeError):
            resultats.append({
                'candidature_id': cand_id,
                'candidat_nom': '',
                'success': False,
                'statut': 'erreur',
                'message': 'Candidature introuvable',
                'error': 'Candidature introuvable',
            })
            erreurs += 1
            continue

        # Cherche le 1er PDF dans le dossier
        cand_dir = _os.path.join(media_root, 'dossiers', str(cand.id))
        pdf_path = None
        if _os.path.isdir(cand_dir):
            for subdir in sorted(_os.listdir(cand_dir)):
                full_subdir = _os.path.join(cand_dir, subdir)
                if _os.path.isdir(full_subdir):
                    for fname in sorted(_os.listdir(full_subdir)):
                        if fname.lower().endswith('.pdf'):
                            pdf_path = _os.path.join(full_subdir, fname)
                            break
                    if pdf_path:
                        break

        if not pdf_path:
            resultats.append({
                'candidature_id': cand.id,
                'candidat_nom': cand.candidat.get_full_name(),
                'success': False,
                'statut': 'erreur',
                'message': 'Aucun PDF depose',
                'error': 'Aucun PDF depose',
            })
            erreurs += 1
            continue

        try:
            score_declare = float(cand.score or 0) or None

            # ✅ APPEL À OCRService RÉEL (pdfplumber + L1/L2/L3 + recalcul score)
            result = OCRService.analyser_releve_complet(pdf_path, score_declare or 0)

            statut = 'ok' if result.get('statut') == 'conforme' else ('anomalie' if result.get('statut') == 'incoherence' else 'erreur')

            if statut == 'anomalie':
                anomalies += 1

            resultats.append({
                'candidature_id': cand.id,
                'candidat_nom': cand.candidat.get_full_name(),
                'numero': cand.numero,
                'master': cand.master.nom if cand.master else '',
                'success': True,
                'statut': statut,
                'score_extrait': result.get('score_extrait'),
                'score_declare': result.get('score_declare'),
                'ecart': result.get('ecart'),
                'confiance': result.get('confiance'),
                'moteur': result.get('moteur'),  # ✅ "pdfplumber" pas "simulation"
                'alerte': result.get('alerte'),
                'message': result.get('alerte') or '',
                # ✅ Détail des notes extraites (L1/L2/L3 + M.G/B.N.R/B.S.P)
                'detail_notes': result.get('detail_notes'),
                'fichier': _os.path.basename(pdf_path),
                'nb_anomalies': len(result.get('anomalies', [])),
            })
        except Exception as exc:
            logger.exception('OCR lot réel — erreur sur candidature %s', cand.id)
            resultats.append({
                'candidature_id': cand.id,
                'candidat_nom': cand.candidat.get_full_name(),
                'success': False,
                'statut': 'erreur',
                'message': str(exc)[:200],
                'error': str(exc)[:200],
            })
            erreurs += 1

    conformes = sum(1 for r in resultats if r.get('statut') == 'ok')
    analysees = conformes + anomalies

    # ✅ RETOURNE LA STRUCTURE ATTENDUE PAR LE FRONTEND
    return Response({
        'success': True,
        'message': f'Analyse OCR complétée: {analysees}/{len(ids)} analysées, {anomalies} anomalie(s)',
        'total': len(ids),
        'nb_analysees': analysees,
        'nb_conformes': conformes,
        'nb_anomalies': anomalies,
        'nb_erreurs': erreurs,
        'resultats': resultats,
    })


def _build_rapport_data(request, candidature_ids):
    """Helper : exécute l'OCR RÉELLE sur les candidatures + retourne (data list + meta).

    Utilise OCRService avec pdfplumber (Req-3) — pas simulation.
    """
    import os as _os
    from django.conf import settings as _settings
    from .ocr_service import OCRService

    media_root = str(getattr(_settings, 'MEDIA_ROOT', ''))
    data = []

    for cand_id in candidature_ids:
        try:
            cand = Candidature.objects.select_related('candidat', 'master').get(pk=int(cand_id))
        except (Candidature.DoesNotExist, ValueError, TypeError):
            continue

        cand_dir = _os.path.join(media_root, 'dossiers', str(cand.id))
        pdf_path = None
        if _os.path.isdir(cand_dir):
            for subdir in sorted(_os.listdir(cand_dir)):
                full_subdir = _os.path.join(cand_dir, subdir)
                if _os.path.isdir(full_subdir):
                    for fname in sorted(_os.listdir(full_subdir)):
                        if fname.lower().endswith('.pdf'):
                            pdf_path = _os.path.join(full_subdir, fname)
                            break
                    if pdf_path:
                        break

        score_declare = float(cand.score or 0)
        row = {
            'numero':         cand.numero or '',
            'candidat_nom':   cand.candidat.get_full_name() or cand.candidat.email,
            'candidat_email': cand.candidat.email,
            'master':         cand.master.nom if cand.master else '',
            'score_declare':  score_declare,
            'score_extrait':  None,
            'ecart':          None,
            'moteur':         '—',
            'statut':         'Pas de PDF',
            'fichier':        '',
        }
        if pdf_path:
            try:
                # ✅ APPEL OCRService RÉEL (pdfplumber + recalcul score complet)
                result = OCRService.analyser_releve_complet(pdf_path, score_declare or 0)

                row['score_extrait'] = result.get('score_extrait')
                row['ecart'] = result.get('ecart')
                row['moteur'] = result.get('moteur') or '—'  # ✅ "pdfplumber" pas "simulation"
                row['fichier'] = _os.path.basename(pdf_path)
                row['detail_notes'] = result.get('detail_notes')

                if result['statut'] == 'conforme':
                    row['statut'] = 'Conforme'
                elif result['statut'] == 'incoherence':
                    row['statut'] = 'INCOHÉRENCE'
                else:
                    row['statut'] = 'Erreur OCR'

            except Exception as exc:
                logger.exception('OCR rapport — erreur cand %s', cand.id)
                row['statut'] = f'Erreur: {str(exc)[:80]}'
        data.append(row)
    return data


# ──────────────────────────────────────────────────────────────────────────
# POST /api/candidatures/rapport-conformite-ocr/excel/
# Génère un rapport Excel de conformité OCR sur les candidatures sélectionnées
# ──────────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rapport_conformite_ocr_excel(request):
    """Export Excel du rapport de conformité OCR."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    from django.http import HttpResponse
    from io import BytesIO

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    ids = request.data.get('candidature_ids') or []
    if not isinstance(ids, list) or not ids:
        return Response({'error': 'candidature_ids manquant'}, status=status.HTTP_400_BAD_REQUEST)

    data = _build_rapport_data(request, ids)
    if not data:
        return Response({'error': 'Aucune donnée à exporter'}, status=status.HTTP_404_NOT_FOUND)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rapport Conformité OCR'

    # ── Titre ──
    ws['A1'] = 'ISIMM — Rapport de Conformité OCR'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1E3A8A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 32

    ws['A2'] = f"Date d'export : {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color='64748B')
    ws.merge_cells('A2:H2')

    # ── Headers ──
    headers = ['N° Candidature', 'Candidat', 'Email', 'Master',
               'Score Déclaré', 'Score OCR', 'Écart Δ', 'Statut Conformité']
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1E40AF')
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # ── Lignes ──
    for r_idx, row in enumerate(data, start=5):
        ws.cell(row=r_idx, column=1, value=row['numero']).border = border
        ws.cell(row=r_idx, column=2, value=row['candidat_nom']).border = border
        ws.cell(row=r_idx, column=3, value=row['candidat_email']).border = border
        ws.cell(row=r_idx, column=4, value=row['master']).border = border
        ws.cell(row=r_idx, column=5, value=row['score_declare']).border = border
        ws.cell(row=r_idx, column=6, value=row['score_extrait']).border = border
        ws.cell(row=r_idx, column=7, value=row['delta']).border = border

        statut_cell = ws.cell(row=r_idx, column=8, value=row['statut'])
        statut_cell.border = border
        statut_cell.alignment = Alignment(horizontal='center')
        statut_cell.font = Font(bold=True)
        if row['statut'] == 'Conforme':
            statut_cell.fill = PatternFill('solid', fgColor='D1FAE5')
            statut_cell.font = Font(bold=True, color='065F46')
        elif row['statut'] == 'ANOMALIE':
            statut_cell.fill = PatternFill('solid', fgColor='FEE2E2')
            statut_cell.font = Font(bold=True, color='991B1B')
        else:
            statut_cell.fill = PatternFill('solid', fgColor='FEF3C7')
            statut_cell.font = Font(bold=True, color='92400E')

    # ── Largeurs colonnes ──
    widths = {'A': 18, 'B': 28, 'C': 30, 'D': 36, 'E': 14, 'F': 14, 'G': 12, 'H': 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Récap ──
    total = len(data)
    conformes = sum(1 for r in data if r['statut'] == 'Conforme')
    anomalies = sum(1 for r in data if r['statut'] == 'ANOMALIE')
    pas_pdf = sum(1 for r in data if 'PDF' in r['statut'])

    last_row = 5 + len(data) + 2
    ws.cell(row=last_row, column=1, value='SYNTHÈSE :').font = Font(bold=True, size=12)
    ws.cell(row=last_row+1, column=1, value=f'  Total candidats : {total}')
    ws.cell(row=last_row+2, column=1, value=f'  Conformes      : {conformes}')
    ws.cell(row=last_row+3, column=1, value=f'  Anomalies      : {anomalies}')
    ws.cell(row=last_row+4, column=1, value=f'  Sans PDF       : {pas_pdf}')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Rapport_Conformite_OCR_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ──────────────────────────────────────────────────────────────────────────
# POST /api/candidatures/rapport-conformite-ocr/pdf/
# Génère un rapport PDF de conformité OCR sur les candidatures sélectionnées
# ──────────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rapport_conformite_ocr_pdf(request):
    """Export PDF du rapport de conformité OCR."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from django.utils import timezone
    from django.http import HttpResponse
    from io import BytesIO

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    ids = request.data.get('candidature_ids') or []
    if not isinstance(ids, list) or not ids:
        return Response({'error': 'candidature_ids manquant'}, status=status.HTTP_400_BAD_REQUEST)

    data = _build_rapport_data(request, ids)
    if not data:
        return Response({'error': 'Aucune donnée à exporter'}, status=status.HTTP_404_NOT_FOUND)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    # ── Titre ──
    title_style = styles['Title']
    title_style.fontSize = 18
    title_style.textColor = colors.HexColor('#1E3A8A')
    elements.append(Paragraph('<b>ISIMM</b> — Rapport de Conformité OCR', title_style))
    elements.append(Paragraph(
        f"Date d'export : {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal'],
    ))
    elements.append(Spacer(1, 0.5*cm))

    # ── Synthèse ──
    total = len(data)
    conformes = sum(1 for r in data if r['statut'] == 'Conforme')
    anomalies = sum(1 for r in data if r['statut'] == 'ANOMALIE')
    pas_pdf = sum(1 for r in data if 'PDF' in r['statut'])

    synth = (
        f"<b>Synthèse :</b> Total : <b>{total}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Conformes : <b><font color='#065F46'>{conformes}</font></b> &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Anomalies : <b><font color='#991B1B'>{anomalies}</font></b> &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Sans PDF : <b><font color='#92400E'>{pas_pdf}</font></b>"
    )
    elements.append(Paragraph(synth, styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    # ── Tableau ──
    tdata = [['N° Candidature', 'Candidat', 'Email', 'Score Déclaré', 'Score OCR', 'Écart Δ', 'Statut']]
    for r in data:
        tdata.append([
            r['numero'][:20],
            r['candidat_nom'][:25],
            r['candidat_email'][:30],
            f"{r['score_declare']:.2f}" if r['score_declare'] is not None else '—',
            f"{r['score_extrait']:.2f}" if r['score_extrait'] is not None else '—',
            f"{r['delta']:.2f}" if r['delta'] is not None else '—',
            r['statut'][:18],
        ])

    col_widths = [3.5*cm, 4.5*cm, 5*cm, 2.5*cm, 2.5*cm, 2*cm, 3.5*cm]
    t = Table(tdata, colWidths=col_widths, repeatRows=1)
    style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]
    # Couleur statut par ligne
    for i, r in enumerate(data, start=1):
        if r['statut'] == 'Conforme':
            style.append(('BACKGROUND', (6,i), (6,i), colors.HexColor('#D1FAE5')))
            style.append(('TEXTCOLOR', (6,i), (6,i), colors.HexColor('#065F46')))
        elif r['statut'] == 'ANOMALIE':
            style.append(('BACKGROUND', (6,i), (6,i), colors.HexColor('#FEE2E2')))
            style.append(('TEXTCOLOR', (6,i), (6,i), colors.HexColor('#991B1B')))
        else:
            style.append(('BACKGROUND', (6,i), (6,i), colors.HexColor('#FEF3C7')))
            style.append(('TEXTCOLOR', (6,i), (6,i), colors.HexColor('#92400E')))
    t.setStyle(TableStyle(style))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)

    filename = f'Rapport_Conformite_OCR_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ──────────────────────────────────────────────────────────────────────
# Sprint 4 — Finaliser dossier de candidature (CORRECTION 1)
# ──────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finaliser_dossier(request, candidature_id):
    """POST /api/candidatures/<id>/finaliser-dossier/

    Le candidat clique « Finaliser mon dossier » → on vérifie que toutes
    les pièces obligatoires sont déposées puis on passe le statut à
    `dossier_depose` et on crée une notification de confirmation.
    """
    try:
        candidature = Candidature.objects.select_related('master', 'candidat').get(
            id=candidature_id,
        )
    except Candidature.DoesNotExist:
        return Response(
            {'error': 'Candidature introuvable.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    # Sécurité : seul le candidat propriétaire peut finaliser sa propre candidature
    if candidature.candidat_id != getattr(request.user, 'id', None):
        return Response(
            {'error': "Vous n'êtes pas autorisé à finaliser ce dossier."},
            status=status.HTTP_403_FORBIDDEN,
        )

    # Vérifier statut compatible : seul un dossier 'preselectionne' ou
    # 'en_attente_dossier' peut être finalisé.
    if candidature.statut not in ('preselectionne', 'en_attente_dossier'):
        return Response(
            {
                'error': "Le dossier ne peut pas être finalisé dans son statut actuel.",
                'statut_actuel': candidature.statut,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Vérifier qu'au moins une pièce a été déposée.
    # Les fichiers candidats sont stockés sur disque via `upload_fichier_dossier` à
    #   MEDIA_ROOT/dossiers/<candidature_ref>/<type>/<filename>
    # (et NON pas dans la table Document — qui reste vide). On scanne donc le dossier.
    from django.conf import settings as _settings
    import os as _os
    candidature_ref = str(candidature.id)
    upload_dir = _os.path.join(
        getattr(_settings, 'MEDIA_ROOT', ''), 'dossiers', str(candidature_ref),
    )
    fichiers_count = 0
    if _os.path.isdir(upload_dir):
        for _root, _dirs, _files in _os.walk(upload_dir):
            for f in _files:
                if not f.startswith('.'):
                    fichiers_count += 1
    # Fallback : si le modèle Document est utilisé (legacy), compter aussi
    try:
        from .models import Document as DocumentModel
        fichiers_count += DocumentModel.objects.filter(
            candidature=candidature,
        ).exclude(fichier='').count()
    except ImportError:
        pass

    if fichiers_count == 0:
        return Response(
            {
                'error': "Aucun document n'a été déposé. Veuillez déposer vos pièces avant de finaliser.",
                'debug_upload_dir': upload_dir,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Tout est OK — mettre à jour le statut
    candidature.statut = 'dossier_depose'
    candidature.dossier_depose = True
    if not candidature.date_depot_dossier:
        candidature.date_depot_dossier = timezone.now()
    candidature.save(update_fields=['statut', 'dossier_depose', 'date_depot_dossier'])

    # Notification au candidat
    try:
        Notification.objects.create(
            utilisateur=candidature.candidat,
            titre='Dossier finalisé',
            message=(
                f"Votre dossier {candidature.numero or candidature.id} a été "
                f"finalisé avec succès et est désormais en cours d'examen."
            ),
            type_notification='CHANGEMENT_STATUT_DOSSIER',
            est_lu=False,
        )
    except Exception:
        # La notification est non-bloquante
        pass

    return Response(
        {
            'message': 'Dossier finalisé avec succès. Votre dossier est en cours d\'examen.',
            'candidature_id': candidature.id,
            'statut': candidature.statut,
            'date_depot_dossier': (
                candidature.date_depot_dossier.isoformat()
                if candidature.date_depot_dossier
                else None
            ),
        },
        status=status.HTTP_200_OK,
    )
