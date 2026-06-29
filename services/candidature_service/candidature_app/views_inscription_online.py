# -*- coding: utf-8 -*-


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from candidature_app.models import Candidature
from django.utils import timezone
from django.db.models import Q
import openpyxl
from django.core.exceptions import ValidationError
import logging

logger = logging.getLogger(__name__)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def saisir_numero_inscription(request, candidature_id):
    """
    Candidat saisit son numéro d'inscription universitaire.

    Payload:
        {
            "numero_inscription": "20241234567"
        }

    Réponse:
        {
            "success": true,
            "statut_inscription": "en_attente_verification",
            "message": "Numéro d'inscription enregistré. En attente de vérification."
        }
    """
    try:
        candidature = Candidature.objects.get(pk=candidature_id)
    except Candidature.DoesNotExist:
        return Response(
            {'error': 'Candidature introuvable'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Vérifier que c'est le candidat qui modifie
    if candidature.candidat != request.user:
        return Response(
            {'error': 'Permission refusée'},
            status=status.HTTP_403_FORBIDDEN
        )

    # Extraire le numéro
    numero = request.data.get('numero_inscription', '').strip()
    if not numero:
        return Response(
            {'error': 'Le numéro d\'inscription est obligatoire'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Vérifier que la candidature est sélectionnée
    if candidature.statut not in ['selectionne', 'preselectionne']:
        return Response(
            {'error': 'Vous devez être sélectionné(e) pour saisir votre numéro d\'inscription'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Mettre à jour
    candidature.numero_inscription = numero
    candidature.statut_inscription = 'en_attente_verification'
    candidature.date_saisie_inscription = timezone.now()
    candidature.save(update_fields=['numero_inscription', 'statut_inscription', 'date_saisie_inscription'])

    return Response({
        'success': True,
        'statut_inscription': candidature.statut_inscription,
        'message': 'Numéro d\'inscription enregistré. En attente de vérification par le responsable.'
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def verifier_excel_inscriptions(request):
    """
    Responsable importe un fichier Excel avec la liste officielle des inscrits.
    Compare les N° inscriptions saisis par les candidats avec ceux du fichier.

    Payload (multipart/form-data):
        - fichier: fichier Excel (.xlsx)
        - master_id (optionnel): limiter à un master

    Format Excel attendu:
        Colonne "numero_inscription" avec les numéros officiels

    Réponse:
        {
            "success": true,
            "nb_confirmes": 45,
            "nb_non_trouves": 3,
            "nb_total_verifies": 48,
            "details": {
                "confirmes": ["Ahmed Ben Ali", ...],
                "non_trouves": ["Nom Candidat", ...]
            }
        }
    """
    # Vérifier les permissions
    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'responsable_master'):
        return Response(
            {'error': 'Seul un responsable peut importer les inscriptions'},
            status=status.HTTP_403_FORBIDDEN
        )

    fichier = request.FILES.get('fichier')
    if not fichier:
        return Response(
            {'error': 'Veuillez fournir un fichier Excel'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Charger le workbook
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active

        # Lire les numéros officiels de la colonne "numero_inscription"
        headers = [cell.value for cell in ws[1]]

        if 'numero_inscription' not in headers:
            return Response(
                {'error': 'La colonne "numero_inscription" est obligatoire dans l\'Excel'},
                status=status.HTTP_400_BAD_REQUEST
            )

        col_idx = headers.index('numero_inscription') + 1
        numeros_officiels = set()

        for row in ws.iter_rows(min_row=2):
            val = row[col_idx - 1].value
            if val:
                numeros_officiels.add(str(val).strip())

        logger.info(f"Excel: {len(numeros_officiels)} numéros lus depuis le fichier")

        # Filtrer les candidatures en attente de vérification
        candidatures_verifier = Candidature.objects.filter(
            statut_inscription='en_attente_verification',
            numero_inscription__isnull=False
        )

        # Optionnel: limiter à un master
        master_id = request.data.get('master_id')
        if master_id:
            candidatures_verifier = candidatures_verifier.filter(master_id=master_id)

        confirmes = []
        non_trouves = []

        for candidature in candidatures_verifier:
            if candidature.numero_inscription and str(candidature.numero_inscription).strip() in numeros_officiels:
                # Candidat trouvé dans la liste officielle
                candidature.statut_inscription = 'inscrit'
                candidature.save(update_fields=['statut_inscription'])
                confirmes.append(candidature.candidat.get_full_name())
            else:
                # Candidat NOT trouvé
                non_trouves.append(candidature.candidat.get_full_name())

        logger.info(f"Inscription: {len(confirmes)} confirmées, {len(non_trouves)} non trouvées")

        return Response({
            'success': True,
            'nb_confirmes': len(confirmes),
            'nb_non_trouves': len(non_trouves),
            'nb_total_verifies': candidatures_verifier.count(),
            'details': {
                'confirmes': confirmes,
                'non_trouves': non_trouves,
            }
        })

    except Exception as e:
        logger.exception("Erreur lors de la vérification Excel")
        return Response(
            {'error': f'Erreur lors de la lecture du fichier: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def comparer_inscrits_admis(request):
    """
    v7 §6.5 — Importe la liste officielle des inscrits (Excel) et renvoie les
    candidats ADMIS qui n'y figurent PAS (= « admis mais non inscrits »).

    Matching par Nom + Prénom (robuste : les numéros changent entre seeds).

    Body (multipart/form-data):
        - fichier   : .xlsx (colonnes « Nom » et « Prénom » obligatoires)
        - master_id : (optionnel) limiter à un master
        - statuts   : (optionnel) liste CSV de statuts considérés « admis »
                      (défaut: selectionne,admis,preselectionne)

    Réponse:
        { success, nb_admis, nb_inscrits_fichier, nb_non_inscrits, non_inscrits: [...] }
    """
    import unicodedata

    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    fichier = request.FILES.get('fichier')
    if not fichier:
        return Response({'error': 'Veuillez fournir un fichier Excel'}, status=status.HTTP_400_BAD_REQUEST)

    def _norm(s):
        s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode()
        return ' '.join(s.lower().split())

    try:
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active
        headers = [_norm(c.value) for c in ws[1]]

        def _col(*names):
            for n in names:
                if n in headers:
                    return headers.index(n)
            return None

        i_nom = _col('nom')
        i_prenom = _col('prenom')
        if i_nom is None or i_prenom is None:
            return Response(
                {'error': "Le fichier doit contenir les colonnes « Nom » et « Prénom »."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        inscrits = set()
        for row in ws.iter_rows(min_row=2):
            nom = row[i_nom].value if i_nom < len(row) else ''
            prenom = row[i_prenom].value if i_prenom < len(row) else ''
            if nom or prenom:
                inscrits.add((_norm(nom), _norm(prenom)))
    except Exception as e:
        logger.exception("Erreur lecture Excel inscrits")
        return Response({'error': f'Lecture du fichier impossible: {e}'}, status=status.HTTP_400_BAD_REQUEST)

    statuts_param = str(request.data.get('statuts') or '').strip()
    statuts = [s.strip() for s in statuts_param.split(',') if s.strip()] or \
        ['selectionne', 'admis', 'preselectionne']

    qs = Candidature.objects.select_related('candidat', 'master').filter(statut__in=statuts)
    master_id = request.data.get('master_id')
    if master_id:
        qs = qs.filter(master_id=master_id)

    admis = []
    non_inscrits = []
    for c in qs:
        nom = _norm(getattr(c.candidat, 'last_name', ''))
        prenom = _norm(getattr(c.candidat, 'first_name', ''))
        item = {
            'candidature_id': c.id,
            'numero': c.numero,
            'nom': getattr(c.candidat, 'last_name', ''),
            'prenom': getattr(c.candidat, 'first_name', ''),
            'master_nom': c.master.nom if c.master else '',
            'statut': c.statut,
        }
        admis.append(item)
        if (nom, prenom) not in inscrits:
            non_inscrits.append(item)

    return Response({
        'success': True,
        'nb_admis': len(admis),
        'nb_inscrits_fichier': len(inscrits),
        'nb_non_inscrits': len(non_inscrits),
        'non_inscrits': non_inscrits,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def liste_inscriptions_saisies(request):
    """
    Liste, pour l'espace responsable, les candidats ayant SAISI leur numéro
    d'inscription universitaire (en attente de vérification ou déjà inscrits).
    Alimente la page « Inscriptions » côté responsable avec des données RÉELLES
    (remplace l'ancien mock du composant consulter-inscriptions).

    Query params (optionnels):
        - master_id : limiter à un master
    """
    role = getattr(request.user, 'role', None)
    if role not in ('admin', 'responsable_commission', 'commission', 'responsable_master'):
        return Response({'error': 'Permission refusee'}, status=status.HTTP_403_FORBIDDEN)

    qs = (
        Candidature.objects
        .select_related('candidat', 'master')
        .filter(numero_inscription__isnull=False)
        .exclude(numero_inscription='')
        .order_by('-date_saisie_inscription', '-id')
    )

    master_id = request.query_params.get('master_id')
    if master_id and str(master_id).isdigit():
        qs = qs.filter(master_id=int(master_id))

    def _categorie(master):
        label = (
            (getattr(master, 'nom', '') or '') + ' ' + (getattr(master, 'specialite', '') or '')
        ).lower()
        if 'ingenieur' in label or 'ingénieur' in label or 'cycle' in label:
            return 'ingenieur'
        if 'data' in label or 'science' in label or 'donnees' in label or 'données' in label:
            return 'master-ds'
        return 'master-gl'

    statut_map = {
        'inscrit': 'Validée',
        'en_attente_verification': 'En attente',
        'rejete': 'Rejetée',
        'rejetee': 'Rejetée',
    }

    rows = []
    for c in qs:
        cand = c.candidat
        si = (getattr(c, 'statut_inscription', '') or '').lower()
        nom_complet = (cand.get_full_name() if cand else '') or getattr(c, 'candidat_nom', '') or '—'
        rows.append({
            'id': c.id,
            'candidat': nom_complet.strip() or '—',
            'email': (getattr(cand, 'email', '') if cand else '') or '',
            'specialite': (
                (getattr(c.master, 'specialite', '') or getattr(c.master, 'nom', ''))
                if c.master else ''
            ),
            'commissionCategory': _categorie(c.master),
            'statut': statut_map.get(si, 'En attente'),
            'paiement': 'Payé' if si == 'inscrit' else 'En attente',
            'dateDepot': (
                c.date_saisie_inscription.strftime('%Y-%m-%d')
                if getattr(c, 'date_saisie_inscription', None) else ''
            ),
            'matricule': str(getattr(c, 'numero_inscription', '') or ''),
        })

    return Response({'success': True, 'count': len(rows), 'inscriptions': rows})
